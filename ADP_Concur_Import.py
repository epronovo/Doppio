"""
ADP_Concur_Import - load Kelly's workbook into the database.

One workbook carries everything: the ADP export on a data sheet, six lookup
tabs, and the three Concur record templates. Nothing here depends on the tab
names, because a re-cut of the report will not keep them - each sheet is
recognised by what its header row says, and the header row is found by looking
for it rather than assumed to be row 1. That matters for the ADP sheet, which
has three rows of Kelly's notes above the real headings.

The lookup tabs are a full refresh: dropping a workbook replaces the maps with
what it carries, because the maps are the workbook's job. The employees merge
on File Number, so re-dropping a fresh ADP cut refreshes the people already
held without disturbing the ones keyed in by hand.
"""
from __future__ import annotations

import argparse
import json
import sqlite3
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ADP_Concur_Db import ADP_COLUMNS, DEFAULT_DB_PATH, connect, resolve_db_path

# How each sheet is recognised: every one of these headings has to be present
# in the candidate header row. Deliberately short lists - enough to be sure,
# few enough that an extra or renamed column elsewhere does not break the match.
SHEET_SIGNATURES: list[tuple[str, list[str]]] = [
    ("employees",      ["payroll company code", "file number", "position status"]),
    ("status_map",     ["position status", "concur status"]),
    ("country_map",    ["adp country"]),
    ("org_map",        ["business unit description", "home department code"]),
    ("language_map",   ["language", "adp language"]),
    ("salary_map",     ["pay grade code", "expense map"]),
    ("supervisor_map", ["exception employee id", "supervisor id"]),
]

# The record templates announce themselves in their first cell.
RECORD_SIGNATURES = {"trx type (305)": "305",
                     "trx type (350)": "350",
                     "trx type (360)": "360"}

# How far down a sheet to look for the header row.
HEADER_SCAN_ROWS = 8


def _norm(value) -> str:
    """A heading reduced to something comparable - case and whitespace out."""
    if value is None:
        return ""
    return " ".join(str(value).split()).strip().lower()


def _cell(value) -> str:
    """A data cell as the string the database stores."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d") if (value.hour, value.minute, value.second) == (0, 0, 0) \
            else value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def find_header_row(rows: list[tuple], wanted: list[str]) -> int | None:
    """
    The 1-based row that carries every heading in `wanted`.

    Matching is on a leading substring, so 'language' finds 'Language' and
    'supervisor id' finds 'Supervisor ID (from ADP)'. The first row that has
    them all wins.
    """
    for i, row in enumerate(rows[:HEADER_SCAN_ROWS], 1):
        heads = [_norm(c) for c in row]
        if all(any(h.startswith(w) or w in h for h in heads if h) for w in wanted):
            return i
    return None


def identify_sheet(rows: list[tuple]) -> tuple[str | None, int | None]:
    """What a worksheet is, and which row its headings are on."""
    if rows and _norm(rows[0][0] if rows[0] else "") in RECORD_SIGNATURES:
        return "record_" + RECORD_SIGNATURES[_norm(rows[0][0])], 1
    # The 350/360 templates put two rows above their headings.
    for i, row in enumerate(rows[:HEADER_SCAN_ROWS], 1):
        if row and _norm(row[0]) in RECORD_SIGNATURES:
            return "record_" + RECORD_SIGNATURES[_norm(row[0])], i

    for kind, wanted in SHEET_SIGNATURES:
        row = find_header_row(rows, wanted)
        if row:
            return kind, row
    return None, None


def column_index(headings: list[str], *wanted: str) -> int | None:
    """Position of the first heading that starts with any of `wanted`."""
    normed = [_norm(h) for h in headings]
    for w in wanted:
        w = _norm(w)
        for i, h in enumerate(normed):
            if h.startswith(w):
                return i
    return None


# --------------------------------------------------------------- the sheets


# Position Status values that mean the person is still employed. Used only to
# pick between two rows for the same person - what Concur is told comes from
# the Status Map, not from here.
LIVE_STATUSES = {"active", "leave", "leave of absence"}


def _row_rank(row: dict) -> tuple:
    """
    How good a candidate one ADP row is for being *the* row for a person.

    ADP writes one row per employment record, so an internal transfer arrives
    as the same File Number twice - terminated under the old payroll company
    and active under the new one. Concur wants one profile per employee, and
    it wants the live one. Ranking, most significant first:

      1. a live Position Status beats a terminated one
      2. no Termination Date beats having one
      3. the later Rehire Date, then the later Hire Date
      4. having a Supervisor ID beats not having one
      5. the later row in the file

    Every rejected row is recorded on the employee, so the choice is visible
    rather than silent.
    """
    status = (row.get("position_status") or "").strip().lower()
    return (
        1 if status in LIVE_STATUSES else 0,
        0 if (row.get("termination_date") or "").strip() else 1,
        (row.get("rehire_date") or "").strip(),
        (row.get("hire_date") or "").strip(),
        1 if (row.get("supervisor_id_raw") or "").strip() else 0,
        row["__row"],
    )


def import_employees(conn: sqlite3.Connection, rows: list[tuple], header_row: int,
                     import_id: int) -> dict:
    """
    Merge the ADP export into ADP_Concur_Employees on File Number.

    Only the ADP columns are written. The derived ones are left to
    ADP_Concur_Map.ADP_Concur_derive(), which the caller runs afterwards -
    a workbook that carries stale lookup results should not put them in the
    database.

    A row already held keeps its key, its include flags and its manual edits to
    fields the file does not carry; a person the file does not mention is left
    alone rather than deleted, because a partial cut of ADP is a normal thing
    to be handed.
    """
    headings = [h for h in rows[header_row - 1]]
    positions = {}
    missing = []
    for heading, column in ADP_COLUMNS:
        idx = column_index(headings, heading)
        if idx is None:
            missing.append(heading)
        positions[column] = idx

    if positions.get("file_number") is None:
        raise ValueError("The ADP sheet has no 'File Number' column - that is "
                         "the key every record is built on.")

    cols = [c for c in positions if positions[c] is not None]

    # Read the whole sheet first, then resolve the duplicates, then write. It
    # has to be done in that order: which row wins is a property of the group,
    # not of the row, so there is nothing to insert until the group is known.
    by_file: dict[str, list[dict]] = {}
    skipped = 0
    for n, row in enumerate(rows[header_row:], header_row + 1):
        file_number = _cell(row[positions["file_number"]]
                            if positions["file_number"] < len(row) else None)
        if not file_number:
            # A wholly blank row is the end of the data, not an error; a row
            # with data but no file number is one we cannot key.
            if any(_cell(c) for c in row):
                skipped += 1
            continue
        record = {"__row": n}
        for c in cols:
            idx = positions[c]
            record[c] = _cell(row[idx]) if idx < len(row) else ""
        record["file_number"] = file_number
        by_file.setdefault(file_number, []).append(record)

    insert_sql = (
        "INSERT INTO ADP_Concur_Employees (source, import_id, source_row, "
        "duplicate_rows, duplicate_note, " + ", ".join(cols) + ") "
        "VALUES ('adp', ?, ?, ?, ?, " + ", ".join("?" for _ in cols) + ") "
        "ON CONFLICT (file_number) DO UPDATE SET "
        + ", ".join(f"{c} = excluded.{c}" for c in cols if c != "file_number")
        + ", import_id = excluded.import_id, source_row = excluded.source_row"
        + ", duplicate_rows = excluded.duplicate_rows"
        + ", duplicate_note = excluded.duplicate_note"
        + ", modified_at = datetime('now')"
        + ", row_state = CASE WHEN ADP_Concur_Employees.row_state = 'new' "
          "THEN 'new' ELSE 'unchanged' END"
    )

    cur = conn.cursor()
    loaded = 0
    duplicates = 0
    for file_number, group in by_file.items():
        note = None
        if len(group) > 1:
            duplicates += 1
            group = sorted(group, key=_row_rank, reverse=True)
            taken, rejected = group[0], group[1:]
            note = (f"{len(group)} ADP rows. Took row {taken['__row']} "
                    f"({taken.get('payroll_company_code') or '?'}, "
                    f"{taken.get('position_status') or '?'}); ignored "
                    + ", ".join(f"row {r['__row']} ({r.get('payroll_company_code') or '?'}, "
                                f"{r.get('position_status') or '?'})" for r in rejected)
                    + ".")
        else:
            taken = group[0]
        cur.execute(insert_sql,
                    [import_id, taken["__row"], len(group), note]
                    + [taken[c] for c in cols])
        loaded += 1

    return {"loaded": loaded, "skipped": skipped, "duplicates": duplicates,
            "source_rows": sum(len(g) for g in by_file.values()),
            "missing_columns": missing}


def _refresh(conn: sqlite3.Connection, table: str, columns: list[str],
             records: list[list]) -> int:
    """Replace a lookup table with what the workbook carries."""
    cur = conn.cursor()
    cur.execute(f"DELETE FROM {table}")
    if not records:
        return 0
    cur.executemany(
        f"INSERT OR REPLACE INTO {table} ({', '.join(columns)}) "
        f"VALUES ({', '.join('?' for _ in columns)})", records)
    return len(records)


def import_status_map(conn, rows, header_row) -> int:
    h = rows[header_row - 1]
    a = column_index(h, "position status")
    b = column_index(h, "concur status")
    out = [[_cell(r[a]), _cell(r[b]) if b is not None and b < len(r) else ""]
           for r in rows[header_row:] if a is not None and a < len(r) and _cell(r[a])]
    return _refresh(conn, "ADP_Concur_StatusMap", ["position_status", "concur_status"], out)


def import_country_map(conn, rows, header_row) -> int:
    h = rows[header_row - 1]
    a = column_index(h, "legal / preferred address", "legal /", "adp country code", "country")
    b = column_index(h, "adp country")
    if a is not None and a == b:
        a, b = 0, 1
    out = [[_cell(r[a]), _cell(r[b]) if b is not None and b < len(r) else ""]
           for r in rows[header_row:] if a is not None and a < len(r) and _cell(r[a])]
    return _refresh(conn, "ADP_Concur_CountryMap", ["adp_country", "concur_country"], out)


def import_org_map(conn, rows, header_row) -> int:
    h = rows[header_row - 1]
    idx = {
        "business_unit_desc": column_index(h, "business unit description"),
        "home_department_desc": column_index(h, "home department description"),
        "home_department_code": column_index(h, "home department code"),
        "org_unit_1": column_index(h, "concur org 1"),
        "org_unit_2": column_index(h, "concur / erp code org 2", "concur/erp code org 2"),
        "default_language": column_index(h, "default org language"),
        "currency": column_index(h, "reimbursement currency", "reumbursement currency"),
    }
    key = idx["business_unit_desc"]
    out = []
    for r in rows[header_row:]:
        if key is None or key >= len(r) or not _cell(r[key]):
            continue
        out.append([_cell(r[i]) if i is not None and i < len(r) else ""
                    for i in idx.values()])
    return _refresh(conn, "ADP_Concur_OrgMap", list(idx), out)


def import_language_map(conn, rows, header_row) -> int:
    """
    Only the first three columns are read.

    The tab carries a second, wider block off to the right - the full Concur
    locale list, 'English (Australia)' and friends - which is reference
    material for choosing the stem, not a lookup the workbook performs.
    """
    h = rows[header_row - 1]
    a = column_index(h, "language")
    b = column_index(h, "language code")
    c = column_index(h, "adp language")
    seen = set()
    out = []
    for r in rows[header_row:]:
        if a is None or a >= len(r):
            continue
        desc = _cell(r[a])
        if not desc or desc in seen:
            continue
        seen.add(desc)
        out.append([desc,
                    _cell(r[b]) if b is not None and b < len(r) else "",
                    _cell(r[c]) if c is not None and c < len(r) else ""])
    return _refresh(conn, "ADP_Concur_LanguageMap",
                    ["language_desc", "language_code", "adp_language"], out)


def import_salary_map(conn, rows, header_row) -> int:
    h = rows[header_row - 1]
    idx = {"pay_grade_code": column_index(h, "pay grade code"),
           "pay_grade_desc": column_index(h, "pay grade description"),
           "expense_map": column_index(h, "expense map"),
           "travel_map": column_index(h, "travel map")}
    key = idx["pay_grade_code"]
    out = []
    for r in rows[header_row:]:
        if key is None or key >= len(r) or not _cell(r[key]):
            continue
        out.append([_cell(r[i]) if i is not None and i < len(r) else ""
                    for i in idx.values()])
    return _refresh(conn, "ADP_Concur_SalaryMap", list(idx), out)


def import_supervisor_map(conn, rows, header_row) -> int:
    """
    The supervisor exceptions.

    A duplicated employee row is not an error - the workbook has one - so the
    first entry for a file number wins and the rest are dropped, which is what
    VLOOKUP does. A blank supervisor is kept: it is how 'top of the food chain'
    is expressed, and the note column says so.
    """
    h = rows[header_row - 1]
    idx = {"file_number": column_index(h, "exception employee id"),
           "employee_name": column_index(h, "exception employee name"),
           "supervisor_name": column_index(h, "supervisor employee name"),
           "supervisor_id": column_index(h, "supervisor id")}
    note_at = 5  # the workbook keeps its free-text note in column F
    key = idx["file_number"]
    out, seen = [], set()
    for r in rows[header_row:]:
        if key is None or key >= len(r):
            continue
        fn = _cell(r[key])
        if not fn or fn in seen:
            continue
        seen.add(fn)
        row = [_cell(r[i]) if i is not None and i < len(r) else "" for i in idx.values()]
        row.append(_cell(r[note_at]) if note_at < len(r) else "")
        out.append(row)
    return _refresh(conn, "ADP_Concur_SupervisorMap",
                    list(idx) + ["note"], out)


def import_layout(conn: sqlite3.Connection, record_type: str,
                  rows: list[tuple], header_row: int) -> int:
    """
    Capture a record template's columns.

    The heading row is the one starting 'Trx Type (nnn)'. Where the template
    also carries a field-width row underneath - the 350 and 360 tabs do - it is
    stored beside the heading, so the extract can be checked against the widths
    Concur publishes without going back to the spreadsheet.
    """
    headings = rows[header_row - 1]
    widths = rows[header_row] if len(rows) > header_row else []
    # A width row is all numbers-ish; a data row starts with the record type.
    if widths and _cell(widths[0]) == record_type:
        widths = []

    cur = conn.cursor()
    cur.execute("DELETE FROM ADP_Concur_Layouts WHERE record_type = ?", (record_type,))
    out = []
    for i, heading in enumerate(headings):
        if heading is None and i > 0 and all(h is None for h in headings[i:]):
            break
        out.append((record_type, i + 1, get_column_letter(i + 1),
                    " ".join(str(heading).split()) if heading is not None else "",
                    _cell(widths[i]) if i < len(widths) else ""))
    cur.executemany(
        "INSERT INTO ADP_Concur_Layouts "
        "(record_type, position, column_ref, heading, max_width) VALUES (?, ?, ?, ?, ?)",
        out)
    return len(out)


# ------------------------------------------------------------------- driver


HANDLERS = {
    "status_map": ("ADP_Concur_StatusMap", import_status_map),
    "country_map": ("ADP_Concur_CountryMap", import_country_map),
    "org_map": ("ADP_Concur_OrgMap", import_org_map),
    "language_map": ("ADP_Concur_LanguageMap", import_language_map),
    "salary_map": ("ADP_Concur_SalaryMap", import_salary_map),
    "supervisor_map": ("ADP_Concur_SupervisorMap", import_supervisor_map),
}


def ADP_Concur_import_workbook(file_path: str | Path,
                               conn: sqlite3.Connection | None = None,
                               db_path: str | None = None,
                               derive: bool = True) -> dict:
    """
    Load one workbook. The whole thing runs in a single transaction.

    Returns what happened per worksheet, including the ones it did not
    recognise - a sheet that is not reported as loaded is one to look at.
    """
    path = Path(file_path)
    own_conn = conn is None
    conn = conn or connect(db_path)

    wb = load_workbook(path, read_only=True, data_only=True)
    cur = conn.cursor()
    cur.execute("INSERT INTO ADP_Concur_Imports (file_name, file_path) VALUES (?, ?)",
                (path.name, str(path)))
    import_id = cur.lastrowid

    sheets, unknown = [], []
    employees = None
    try:
        for ws in wb.worksheets:
            rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
            if not rows:
                continue
            kind, header_row = identify_sheet(rows)
            if kind is None:
                unknown.append(ws.title)
                continue

            if kind == "employees":
                employees = import_employees(conn, rows, header_row, import_id)
                sheets.append({"sheet": ws.title, "kind": "ADP export",
                               "rows": employees["loaded"],
                               "source_rows": employees["source_rows"],
                               "duplicates": employees["duplicates"],
                               "skipped": employees["skipped"],
                               "missing_columns": employees["missing_columns"]})
            elif kind.startswith("record_"):
                record_type = kind.split("_", 1)[1]
                n = import_layout(conn, record_type, rows, header_row)
                sheets.append({"sheet": ws.title,
                               "kind": f"{record_type} layout", "rows": n})
            else:
                table, handler = HANDLERS[kind]
                n = handler(conn, rows, header_row)
                sheets.append({"sheet": ws.title, "kind": table, "rows": n})

        cur.execute(
            "UPDATE ADP_Concur_Imports SET sheet_counts = ?, row_count = ? "
            "WHERE import_id = ?",
            (json.dumps(sheets), employees["loaded"] if employees else 0, import_id))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        wb.close()

    result = {"file": path.name, "import_id": import_id, "sheets": sheets,
              "unknown_sheets": unknown,
              "employees": employees["loaded"] if employees else 0}

    if derive:
        from ADP_Concur_Map import ADP_Concur_derive, load_config
        result["derive"] = ADP_Concur_derive(conn, load_config())

    if own_conn:
        conn.close()
    return result


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Load an ADP/Concur workbook.")
    ap.add_argument("files", nargs="+", help="One or more .xlsx workbooks")
    ap.add_argument("--db", default=None, help=f"SQLite path (default {DEFAULT_DB_PATH})")
    ap.add_argument("--no-derive", action="store_true",
                    help="Skip rebuilding the derived columns afterwards")
    args = ap.parse_args(argv)

    conn = connect(args.db)
    print(f"Database: {resolve_db_path(args.db)}")
    for f in args.files:
        res = ADP_Concur_import_workbook(f, conn=conn, derive=not args.no_derive)
        print(f"\n{res['file']}  (import {res['import_id']})")
        for s in res["sheets"]:
            extra = ""
            if s.get("duplicates"):
                extra += (f" from {s['source_rows']} ADP row(s), "
                          f"{s['duplicates']} employee(s) had more than one")
            if s.get("skipped"):
                extra += f", {s['skipped']} unkeyed row(s) skipped"
            if s.get("missing_columns"):
                extra += f", missing: {', '.join(s['missing_columns'])}"
            print(f"  {s['sheet']:<18} {s['kind']:<26} {s['rows']:>5} row(s){extra}")
        for u in res["unknown_sheets"]:
            print(f"  {u:<18} not recognised - ignored")
        if "derive" in res:
            d = res["derive"]
            print(f"  derived {d['employees']} employee(s): "
                  f"{d['errors']} error(s), {d['warnings']} warning(s)")
    conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
