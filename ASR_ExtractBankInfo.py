import datetime
import openpyxl
import pandas as pd
import re
import requests
import time
import xlsxwriter
from pathlib import Path


# EVS100 output directory (mirrors the pattern used in ASR_UpdateBasePrice.py)
EVS100_TO_PROCESS = Path(__file__).parent / "evs100" / "ToProcess"


def _load_google_api_key() -> str:
    """Read the Google Maps API key from config.json next to this script."""
    config_path = Path(__file__).parent / "config.json"
    if config_path.is_file():
        return config_path.read_text().strip()
    return ""



# =============================================================================
# Bank-routing mapping helpers
# =============================================================================

def _is_alpha_start(value: str) -> bool:
    """Return True when the first 2 characters of *value* are both alphabetic."""
    v = value.strip()
    return len(v) >= 2 and v[:2].isalpha()


def _is_iban(value: str) -> bool:
    """
    Return True when *value* matches the basic IBAN structure:
    2-letter country code followed by 2 numeric check digits.
    Spaces are ignored (IBANs are sometimes formatted with spaces).
    """
    v = value.strip().replace(" ", "")
    return len(v) >= 4 and v[:2].isalpha() and v[2:4].isdigit()


def _map_bank_routing(routing_val: str) -> tuple[str, str, str]:
    """
    Apply the BACN field mapping rules from the API_CRS692MI_AddBankAccount2
    workbook to the "Bank Routing (Local [ABA, Sort Code] or SWIFT Code)" field.

    Rules
    -----
    • If the value contains "/" → split into two parts:
        – Both numeric            → part1 → SWBI,  part2 → BACN
        – part1 alpha, part2 num  → part1 → SWBI,  part2 → BACN
        – part1 alpha, part2 alpha→ part1 → SWBI,  part2 → IBAN
        – part1 numeric, part2 alpha → part1 → BACN, part2 → IBAN
    • If no "/":
        – Starts with alpha → SWBI
        – Numeric           → BACN

    Returns (swbi, bacn, iban).
    """
    val = routing_val.strip()
    if not val:
        return "", "", ""

    if "/" in val:
        p1, p2 = (p.strip() for p in val.split("/", 1))
        p1_alpha = _is_alpha_start(p1)
        p2_iban  = _is_iban(p2)

        # p2 is a valid IBAN → always route it there
        if p2_iban:
            iban_part = p2.replace(" ", "")
            # p1 alpha → SWBI; p1 numeric → BACN
            if p1_alpha:
                return p1, "", iban_part
            return "", p1, iban_part
        # No IBAN in p2 — both numeric → p1 → SWBI, p2 → BACN
        if not p1_alpha:
            return p1, p2, ""
        # p1 alpha, p2 numeric → p1 → SWBI, p2 → BACN
        return p1, p2, ""
    else:
        if _is_alpha_start(val):
            return val, "", ""
        return "", val, ""


# =============================================================================
# EVS100 export — CRS692MI.AddBankAccount2
# =============================================================================

def export_crs692mi_xlsx(results: list[dict], out_dir: Path) -> Path:
    """
    Write extracted bank-info records in the EVS100 import format expected by
    CRS692MI.AddBankAccount2.

    Layout
    ------
    Control sheet:
        Row 1  – Worksheet | Description | Data
        Row 2  – API_CRS692MI_AddBankAccount2 | Bank Accounts | x

    Data sheet (API_CRS692MI_AddBankAccount2):
        Row 1  – field names
        Row 2  – field descriptions
        Row 3  – required flags
        Row 4+ – one data row per supplier

    Field mapping applied
    --------------------
    MESSAGE  → (blank)
    BKTP     → hardcoded "03"
    ACHO     → SupplierNumber
    BKIN     → hardcoded "US2"
    BKID     → hardcoded "001"
    CBPY     → hardcoded "01"
    STAT     → hardcoded "20"
    BANA     → Beneficiary Name  (truncated to 36 characters)
    LNCD     → hardcoded "GB"
    IBAN     → Bank Account Number (IBAN if available)
                 – first 2 chars alphabetic → IBAN
                 – first 2 chars numeric    → BACN instead
    SWBI     → SWIFT Code (payment outside U.S.)
                 – also receives alpha-prefix values split from Bank Routing
    BACN     → Bank Routing (Local [ABA, Sort Code] or SWIFT Code)
                 – see _map_bank_routing() for full split/alpha logic

    Returns the path of the written file.
    """
    DATA_SHEET = "API_CRS692MI_AddBankAccount2"

    COLUMNS = [
        "MESSAGE", "BKTP", "ACHO", "BKIN", "BKID",
        "CBPY",    "STAT", "BANA", "LNCD", "IBAN", "SWBI", "BACN",
    ]
    DESCRIPTIONS = [
        None,
        "Bank account type (N:2)",
        "Account holder (A:10)",
        "Bank account indicator (A:4)",
        "Bank account identity (A:5)",
        "Bank priority (N:2)",
        "Status (A:2)",
        "Bank account name (A:36)",
        "Language (A:2)",
        "International bank account number (A:34)",
        "SWIFT/BIC code (A:34)",
        "Bank account number (A:35)",
    ]
    REQUIRED = ["no", "yes", "yes", "yes", "yes", "yes", "yes", "yes", "yes", "yes", "yes", "yes"]

    out_dir.mkdir(parents=True, exist_ok=True)
    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"API_CRS692MI_{ts}.xlsx"

    wb = xlsxwriter.Workbook(str(out_path))

    # ── Control sheet ──────────────────────────────────────────────────────
    ws_ctrl = wb.add_worksheet("Control")
    for col, val in enumerate(["Worksheet", "Description", "Data"]):
        ws_ctrl.write(0, col, val)
    for col, val in enumerate([DATA_SHEET, "Bank Accounts", "x"]):
        ws_ctrl.write(1, col, val)

    # ── Data sheet ─────────────────────────────────────────────────────────
    ws_data = wb.add_worksheet(DATA_SHEET)

    # Row 1 – field names
    for col, name in enumerate(COLUMNS):
        ws_data.write(0, col, name)

    # Row 2 – descriptions
    for col, desc in enumerate(DESCRIPTIONS):
        if desc is not None:
            ws_data.write(1, col, desc)

    # Row 3 – required flags
    for col, req in enumerate(REQUIRED):
        ws_data.write(2, col, req)

    # Rows 4+ – one row per supplier
    for row_idx, rec in enumerate(results, start=3):
        supplier_num = rec.get("SupplierNumber", "")

        # BANA – truncate at 36 characters
        bana = rec.get("Beneficiary Name", "")[:36]

        # IBAN field: proper IBAN format (2-letter country + 2 check digits) → IBAN;
        # numeric → BACN; contains "/" → apply same split logic as routing field;
        # anything else (e.g. plain account codes) is dropped
        bank_acct = rec.get("Bank Account Number (IBAN if available)", "").strip()
        if _is_iban(bank_acct):
            iban_val  = bank_acct.replace(" ", "")
            acct_bacn = ""
            acct_swbi = ""
        elif "/" in bank_acct:
            acct_swbi, acct_bacn, iban_val = _map_bank_routing(bank_acct)
        elif bank_acct and not _is_alpha_start(bank_acct):
            iban_val  = ""
            acct_bacn = bank_acct
            acct_swbi = ""
        else:
            iban_val  = ""
            acct_bacn = ""
            acct_swbi = ""

        # SWBI: explicit SWIFT Code field
        swbi_val = rec.get("SWIFT Code (payment outside U.S.)", "").strip()

        # Bank Routing → split into SWBI / BACN / IBAN components
        routing_raw = rec.get("Bank Routing (Local [ABA, Sort Code] or SWIFT Code)", "").strip()
        routing_swbi, routing_bacn, routing_iban = _map_bank_routing(routing_raw)

        # Merge: explicit SWIFT Code takes priority; then routing-derived; then acct_swbi
        final_swbi = swbi_val or routing_swbi or acct_swbi
        # BACN: numeric bank_acct takes priority; fall back to routing-derived BACN
        final_bacn = acct_bacn or routing_bacn
        # IBAN: alpha bank_acct takes priority; fall back to routing split (p2 alpha)
        final_iban = iban_val or routing_iban

        row_values = [
            None,          # MESSAGE  (always blank)
            "03",          # BKTP
            supplier_num,  # ACHO
            "US2",         # BKIN
            "001",         # BKID
            "01",          # CBPY
            "20",          # STAT
            bana,          # BANA
            "GB",          # LNCD
            final_iban,    # IBAN
            final_swbi,    # SWBI
            final_bacn,    # BACN
        ]

        for col, val in enumerate(row_values):
            if val:
                ws_data.write(row_idx, col, val)

    wb.close()
    return out_path


# =============================================================================
# Country name → ISO 3166-1 alpha-2 lookup
# =============================================================================

_COUNTRY_CODES: dict[str, str] = {
    # A
    "afghanistan": "AF", "albania": "AL", "algeria": "DZ", "argentina": "AR",
    "armenia": "AM", "australia": "AU", "austria": "AT", "azerbaijan": "AZ",
    # B
    "bahrain": "BH", "bangladesh": "BD", "belarus": "BY", "belgium": "BE",
    "bolivia": "BO", "bosnia": "BA", "bosnia and herzegovina": "BA",
    "brazil": "BR", "brasil": "BR", "bulgaria": "BG",
    # C
    "cambodia": "KH", "cameroon": "CM", "canada": "CA", "chile": "CL",
    "china": "CN", "people's republic of china": "CN", "prc": "CN",
    "colombia": "CO", "costa rica": "CR", "croatia": "HR",
    "czech republic": "CZ", "czechia": "CZ",
    # D
    "denmark": "DK",
    # E
    "ecuador": "EC", "egypt": "EG", "el salvador": "SV", "estonia": "EE",
    "ethiopia": "ET",
    # F
    "finland": "FI", "france": "FR",
    # G
    "georgia": "GE", "germany": "DE", "ghana": "GH", "greece": "GR",
    "guatemala": "GT",
    # H
    "honduras": "HN", "hong kong": "HK", "hungary": "HU",
    # I
    "iceland": "IS", "india": "IN", "indonesia": "ID", "iran": "IR",
    "iraq": "IQ", "ireland": "IE", "israel": "IL", "italy": "IT",
    # J
    "jamaica": "JM", "japan": "JP", "jordan": "JO",
    # K
    "kazakhstan": "KZ", "kenya": "KE", "korea": "KR", "south korea": "KR",
    "republic of korea": "KR", "kuwait": "KW",
    # L
    "latvia": "LV", "lebanon": "LB", "lithuania": "LT", "luxembourg": "LU",
    # M
    "malaysia": "MY", "mexico": "MX", "méxico": "MX",
    "moldova": "MD", "morocco": "MA", "mozambique": "MZ",
    # N
    "nepal": "NP", "netherlands": "NL", "new zealand": "NZ",
    "nigeria": "NG", "norway": "NO",
    # O
    "oman": "OM",
    # P
    "pakistan": "PK", "panama": "PA", "paraguay": "PY", "peru": "PE",
    "philippines": "PH", "poland": "PL", "portugal": "PT",
    # Q
    "qatar": "QA",
    # R
    "romania": "RO", "russia": "RU", "russian federation": "RU",
    # S
    "saudi arabia": "SA", "senegal": "SN", "serbia": "RS",
    "singapore": "SG", "slovakia": "SK", "slovenia": "SI",
    "south africa": "ZA", "spain": "ES", "sri lanka": "LK",
    "sweden": "SE", "switzerland": "CH",
    # T
    "taiwan": "TW", "republic of china": "TW",
    "tanzania": "TZ", "thailand": "TH", "tunisia": "TN",
    "turkey": "TR", "türkiye": "TR", "turkiye": "TR",
    # U
    "ukraine": "UA", "united arab emirates": "AE", "uae": "AE",
    "united kingdom": "GB", "uk": "GB", "great britain": "GB", "england": "GB",
    "united states": "US", "united states of america": "US",
    "usa": "US", "u.s.a.": "US", "u.s.": "US",
    "uruguay": "UY", "uzbekistan": "UZ",
    # V
    "venezuela": "VE", "vietnam": "VN", "viet nam": "VN",
    # Z
    "zambia": "ZM", "zimbabwe": "ZW",
}

# Also accept bare ISO alpha-2 codes as-is (e.g. "CN", "GB")
_ISO_ALPHA2 = {v for v in _COUNTRY_CODES.values()}


def _lookup_country(token: str) -> str:
    """Return ISO alpha-2 code for *token*, or '' if not recognised."""
    t = token.strip()
    # Already a 2-letter ISO code
    if t.upper() in _ISO_ALPHA2:
        return t.upper()
    # Already a 3-letter ISO-like code — try stripping to 2 if it maps
    if re.match(r'^[A-Z]{3}$', t.upper()):
        two = t.upper()[:2]
        if two in _ISO_ALPHA2:
            return two
    key = re.sub(r'[^\w\s]', '', t).strip().lower()
    return _COUNTRY_CODES.get(key, "")


# =============================================================================
# Google Places address lookup
# =============================================================================

def _lookup_address_from_maps(name: str, api_key: str, country_hint: str = "") -> dict | None:
    """
    Look up a supplier address using the Google Geocoding API and return
    structured address fields.

    Combines *name* + *country_hint* (the full raw Beneficiary Address) into
    a single query string — the Geocoding API returns fully typed
    address_components in one call, including the ISO alpha-2 country code.

    Returns a dict with keys: adr1, adr2, town, ecar, pono, cscd,
    plus _formatted (Google's formatted address, for display only),
    or None if the lookup fails or returns no usable result.
    """
    query = f"{name} {country_hint}".strip() if country_hint else name

    url = "https://maps.googleapis.com/maps/api/geocode/json"
    params = {
        "address": query,
        "key":     api_key,
    }

    try:
        resp = requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        data = resp.json()
    except Exception as exc:
        print(f"    ⚠️   Maps lookup failed for {repr(name)}: {exc}")
        return None

    status = data.get("status", "")
    if status != "OK":
        print(f"    ⚠️   Maps status {status} for {repr(name)}")
        return None

    results = data.get("results", [])
    if not results:
        return None

    result     = results[0]
    components = result.get("address_components", [])
    formatted  = result.get("formatted_address", "")

    if not components:
        return None

    # Index components by their primary type for easy retrieval
    comp: dict[str, dict] = {}
    for c in components:
        for t in c.get("types", []):
            comp.setdefault(t, c)

    def get(type_name: str, short: bool = False) -> str:
        c = comp.get(type_name, {})
        return c.get("short_name" if short else "long_name", "")

    # Build street line from available components
    street_parts = [get("subpremise"), get("premise"), get("street_number"), get("route")]
    street = " ".join(p for p in street_parts if p)

    # City: locality preferred; fall back to sub-locality or admin level 2
    town = (
        get("locality")
        or get("sublocality_level_1")
        or get("administrative_area_level_2")
    )

    # ECAR: only keep short_name when it is already a proper 2-letter code
    # (e.g. "MH" for Maharashtra, "TX" for Texas).  Long names like "Greater
    # London", "Liaoning", or "Östergötland" are not useful in a 2-char field.
    ecar_short = get("administrative_area_level_1", short=True)
    ecar_clean = ecar_short.replace(".", "").replace(" ", "")
    ecar = ecar_short.replace(".", "")[:2] if len(ecar_clean) <= 2 else ""

    # ADR1 fallback: when no street_number / route component exists, use the
    # first (most-specific) segment of the formatted address as ADR1 and the
    # second segment (if it is not the city or country) as ADR2.
    if not street and formatted:
        fa_parts = [p.strip() for p in formatted.split(",")]
        street = fa_parts[0] if fa_parts else ""
        if len(fa_parts) > 1:
            p2 = fa_parts[1].strip()
            # Skip segment if it matches the town or looks like a bare postal code
            if p2 and p2.lower() != town.lower() and not re.match(r'^\d[\d\s\-]+$', p2):
                street = (street + ", " + p2) if street else p2

    return {
        "adr1":       street[:36],
        "adr2":       street[36:72].strip(", "),
        "adr3":       street[72:108].strip(", "),
        "town":       town[:20],
        "ecar":       ecar,
        "pono":       get("postal_code")[:10],
        "cscd":       get("country", short=True)[:3],
        "_formatted": formatted,   # display only — not written to sheet
    }


# =============================================================================
# Address parsing helper
# =============================================================================

_POSTAL_RE = re.compile(
    r'\b(\d{5}(?:-\d{4})?'                   # US ZIP (12345 or 12345-6789)
    r'|[A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2}'   # UK postcode  (EC1A 1BB)
    r'|\d{4,8})\b',
    re.IGNORECASE,
)
_STATE_RE = re.compile(r'(?:^|[\s,])([A-Z]{2})(?:[\s,]|$)')


def _parse_address(address_str: str) -> dict:
    """
    Parse a free-form Beneficiary Address into CRS620MI fields.

    Algorithm (back-to-front)
    -------------------------
    1. Split into parts by newline; if single line, split by comma.
    2. From the END, identify and remove the country → CSCD (ISO alpha-2).
    3. From the new END, extract postal code (regex) → PONO, and
       state/province (2-letter abbreviation) → ECAR, from that token.
    4. The new last part becomes the city → TOWN (max 20 chars).
    5. Rejoin the remaining parts with ", " and character-split into
       ADR1 (chars 1-36) and ADR2 (chars 37-72).

    Returns dict with keys: adr1, adr2, town, ecar, pono, cscd.
    """
    out = {"adr1": "", "adr2": "", "adr3": "", "town": "", "ecar": "", "pono": "", "cscd": ""}
    if not address_str:
        return out

    # Split into parts
    if re.search(r'[\r\n]', address_str):
        parts = [p.strip() for p in re.split(r'[\r\n]+', address_str.strip()) if p.strip()]
    else:
        parts = [p.strip() for p in address_str.split(',') if p.strip()]

    if not parts:
        return out

    # ── Step 1: country from the end ─────────────────────────────────────
    # The last part may be just a country ("China"), or country + postal
    # code together ("CHINA 215164").  Try both forms.
    last = parts[-1]
    code = _lookup_country(last)
    if not code:
        # Strip any embedded postal code and retry
        pm_last = _POSTAL_RE.search(last)
        if pm_last:
            country_candidate = (last[:pm_last.start()] + last[pm_last.end():]).strip().strip(",").strip()
            code = _lookup_country(country_candidate)
            if code and not out["pono"]:
                out["pono"] = pm_last.group(1).strip()[:10]
    if code:
        out["cscd"] = code
        parts.pop()

    if not parts:
        return out

    # ── Step 2: postal code + state from the new last part ───────────────
    city_candidate = parts[-1]

    if not out["pono"]:          # postal may already have been found in step 1
        pm = _POSTAL_RE.search(city_candidate)
        if pm:
            out["pono"] = pm.group(1).strip()[:10]
            city_candidate = (city_candidate[:pm.start()] + city_candidate[pm.end():]).strip().strip(",").strip()

    sm = _STATE_RE.search(city_candidate)
    if sm:
        out["ecar"] = sm.group(1)
        city_candidate = (city_candidate[:sm.start()] + city_candidate[sm.end():]).strip().strip(",").strip()

    # ── Step 3: city is what's left of that last part ────────────────────
    if city_candidate:
        out["town"] = city_candidate[:20].strip()
        parts.pop()
    else:
        # Entire last part consumed by postal/state → pop it, then take
        # the new last part as the city
        parts.pop()
        if parts:
            out["town"] = parts[-1][:20].strip()
            parts.pop()

    if not parts:
        return out

    # ── Step 4: character-split the rest into ADR1 / ADR2 ────────────────
    remaining = ", ".join(parts)
    out["adr1"] = remaining[:36]
    out["adr2"] = remaining[36:72].strip(", ")
    out["adr3"] = remaining[72:108].strip(", ")

    return out


# =============================================================================
# EVS100 export — CRS620MI.AddAddress
# =============================================================================

def export_crs620mi_xlsx(results: list[dict], out_dir: Path) -> Path:
    """
    Write extracted supplier address records in the EVS100 import format
    expected by CRS620MI.AddAddress.

    Layout
    ------
    Control sheet:
        Row 1  – Worksheet | Description | Data
        Row 2  – API_CRS620MI_AddAddress | Supplier Address | x

    Data sheet (API_CRS620MI_AddAddress):
        Row 1  – field names
        Row 2  – field descriptions
        Row 3  – required flags
        Row 4+ – one data row per supplier

    Field mapping applied
    --------------------
    SUNO  → SupplierNumber
    ADTE  → hardcoded "10"
    ADID  → hardcoded "001"
    STDT  → today's date (YYYYMMDD)
    SUNM  → Beneficiary Name (truncated to 36 characters)
    ADR1  → first 36 characters of Beneficiary Address
    ADR2  → next 36 characters of Beneficiary Address
    TOWN  → city parsed from Beneficiary Address (max 20 chars)
    ECAR  → state/province parsed from Beneficiary Address (2 chars)
    PONO  → postal code parsed from Beneficiary Address (max 10 chars)
    CSCD  → country parsed from Beneficiary Address (max 3 chars)
    PRIA  → hardcoded "0"

    Returns the path of the written file.
    """
    DATA_SHEET = "API_CRS620MI_AddAddress"
    TODAY      = datetime.date.today().strftime("%Y%m%d")

    COLUMNS = [
        "MESSAGE", "SUNO", "ADTE", "ADID", "STDT",
        "SUNM",    "ADR1", "ADR2", "ADR3", "TOWN", "ECAR", "PONO", "CSCD", "PRIA",
        "GOOGLE_ADDRESS",
        "SOURCE_ADDRESS",
    ]
    DESCRIPTIONS = [
        None,
        "Supplier number (A:10)",
        "Address type (N:2)",
        "Address number (A:6)",
        "Start date (D:10)",
        "Supplier name (A:36)",
        "Address line 1 (A:36)",
        "Address line 2 (A:36)",
        "Address line 3 (A:36)",
        "City (A:20)",
        "State (A:2)",
        "Postal code (A:10)",
        "Country (A:3)",
        "Printout code (N:1)",
        "Google formatted address (reference only)",
        "Beneficiary address used for lookup (reference only)",
    ]
    REQUIRED = [
        "no", "yes", "yes", "yes", "yes",
        "yes", "yes", "yes", "yes", "yes", "yes", "yes", "yes", "yes",
        "no", "no",
    ]

    out_dir.mkdir(parents=True, exist_ok=True)
    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"API_CRS620MI_{ts}.xlsx"

    wb = xlsxwriter.Workbook(str(out_path))

    # ── Control sheet ──────────────────────────────────────────────────────
    ws_ctrl = wb.add_worksheet("Control")
    for col, val in enumerate(["Worksheet", "Description", "Data"]):
        ws_ctrl.write(0, col, val)
    for col, val in enumerate([DATA_SHEET, "Supplier Address", "x"]):
        ws_ctrl.write(1, col, val)

    # ── Data sheet ─────────────────────────────────────────────────────────
    ws_data = wb.add_worksheet(DATA_SHEET)

    for col, name in enumerate(COLUMNS):
        ws_data.write(0, col, name)

    for col, desc in enumerate(DESCRIPTIONS):
        if desc is not None:
            ws_data.write(1, col, desc)

    for col, req in enumerate(REQUIRED):
        ws_data.write(2, col, req)

    # Load API key once; if present, use Google Places for each supplier
    api_key = _load_google_api_key()
    if api_key:
        print(f"  🗺️   Google Maps API key loaded — using Places lookup")
    else:
        print(f"  ⚠️   No Google Maps API key found — falling back to local address parser")

    # Rows 4+ – one row per supplier
    for row_idx, rec in enumerate(results, start=3):
        name     = rec.get("Beneficiary Name", "")
        addr_raw = rec.get("Beneficiary Address ", "").strip()

        addr = None
        google_formatted = ""
        if api_key and name:
            addr = _lookup_address_from_maps(name, api_key, country_hint=addr_raw)
            if addr:
                google_formatted = addr.pop("_formatted", "")
                # Google gives clean structured fields for the tail (city, state,
                # postal, country).  For the street lines, the original wire-form
                # address is more precise (floor, unit, building) so parse it
                # locally and use those ADR1/ADR2/ADR3 values instead.
                # If the original address is blank, keep Google's street components.
                if addr_raw:
                    original_parsed = _parse_address(addr_raw)
                    addr["adr1"] = original_parsed["adr1"]
                    addr["adr2"] = original_parsed["adr2"]
                    addr["adr3"] = original_parsed["adr3"]
                print(f"    ✅  [{rec.get('SupplierNumber','')}] {name}")
                print(f"         Google  : {google_formatted}")
                print(f"         ADR1    : {addr['adr1']}")
                print(f"         ADR2    : {addr['adr2']}")
                print(f"         TOWN    : {addr['town']}  ECAR: {addr['ecar']}  PONO: {addr['pono']}  CSCD: {addr['cscd']}")
            else:
                print(f"    ⚠️   [{rec.get('SupplierNumber','')}] {name} → no Maps result, using local parser")
            time.sleep(0.05)   # stay well within 50 req/s rate limit

        if addr is None:
            addr = _parse_address(addr_raw)

        row_values = [
            None,                                        # MESSAGE
            rec.get("SupplierNumber", ""),               # SUNO
            "10",                                        # ADTE
            "001",                                       # ADID
            TODAY,                                       # STDT
            rec.get("Beneficiary Name", "")[:36],        # SUNM
            addr["adr1"],                                # ADR1
            addr["adr2"],                                # ADR2
            addr["adr3"],                                # ADR3
            addr["town"],                                # TOWN
            addr["ecar"],                                # ECAR
            addr["pono"],                                # PONO
            addr["cscd"],                                # CSCD
            "0",                                         # PRIA
            google_formatted,                            # GOOGLE_ADDRESS (reference)
            addr_raw,                                    # SOURCE_ADDRESS (reference)
        ]

        for col, val in enumerate(row_values):
            if val:
                ws_data.write(row_idx, col, val)

    wb.close()
    return out_path


# =============================================================================
# Extraction
# =============================================================================

def process_green_sheets_to_excel(file_path, output_excel="extracted_green_sheets.xlsx"):
    # Load the workbook with data_only=True to read values instead of formulas
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Target fields (each will become its own separate column)
    target_fields = [
        "Entity",
        "Beneficiary Name",
        "Beneficiary Address ", # Kept the space to match the sheet exactly
        "Beneficiary Account Number",
        "Bank Name",
        "Bank Address",
        "Bank Routing or ABA#",
        "SWIFT Code (payment outside U.S.)",
        "Bank Account Number (IBAN if available)",
        "Bank Routing (Local [ABA, Sort Code] or SWIFT Code)"
    ]

    # Store the results here
    results = []

    # Common HEX codes for green tabs in Excel
    green_colors = ['FF92D050', 'FF00B050', '92D050', '00B050', 'FF00FF00', '00FF00']

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        color = ws.sheet_properties.tabColor

        # Check if the tab has a color and if it matches any green shade
        if color is not None and getattr(color, 'rgb', None) in green_colors:

            # Extract only the Supplier Number from the sheet name
            match = re.search(r"^(\d+)", sheet_name.strip())
            supplier_num = match.group(1) if match else sheet_name

            # Initialize the row dictionary with the Supplier Number and empty strings for fields
            row_data = {"SupplierNumber": supplier_num}
            for field in target_fields:
                row_data[field] = ""

            # Scan each row up to the maximum row used in the sheet
            for r in range(1, max(150, ws.max_row + 1)):
                for c in range(1, 5):
                    cell_val = ws.cell(row=r, column=c).value

                    if isinstance(cell_val, str):
                        cleaned_cell = cell_val.strip()

                        # Check if the cell text matches any of our target fields
                        for field in target_fields:
                            if cleaned_cell == field.strip():
                                # Look for the first non-empty value to the right
                                for v_col in range(c + 1, 11):
                                    val = ws.cell(row=r, column=v_col).value
                                    if val is not None and str(val).strip() != "":
                                        row_data[field] = str(val).strip()
                                        break
                                break

            results.append(row_data)

    # Convert results list to a DataFrame
    df_results = pd.DataFrame(results)

    # Rearrange the columns to strictly follow your requested order
    columns_order = ["SupplierNumber"] + target_fields
    df_results = df_results[columns_order]

    # Export results ONLY to Excel
    df_results.to_excel(output_excel, index=False)
    print(f"Extraction complete! Found and processed {len(results)} green sheets into {output_excel}.")

    return results


# =============================================================================
# Main
# =============================================================================

def main() -> None:
    print(f"\n{'═' * 60}")
    print(f"  ExtractBankInfo")
    print(f"{'═' * 60}")

    # ------------------------------------------------------------------ #
    # Run extraction                                                      #
    # ------------------------------------------------------------------ #
    results = process_green_sheets_to_excel("Wire Request Forms - FMG.xlsm")

    if not results:
        print("⚠️   No green sheets found. Exiting.")
        return

    # ------------------------------------------------------------------ #
    # Export to EVS100 format for CRS692MI.AddBankAccount2               #
    # ------------------------------------------------------------------ #
    print(f"\n▶   Exporting {len(results):,} record(s) to CRS692MI EVS100 file …")
    xlsx_path = export_crs692mi_xlsx(results, EVS100_TO_PROCESS)
    print(f"  📄  Saved to: evs100/ToProcess/{xlsx_path.name}")

    # ------------------------------------------------------------------ #
    # Export to EVS100 format for CRS620MI.AddAddress                    #
    # ------------------------------------------------------------------ #
    print(f"\n▶   Exporting {len(results):,} record(s) to CRS620MI EVS100 file …")
    xlsx_path_addr = export_crs620mi_xlsx(results, EVS100_TO_PROCESS)
    print(f"  📄  Saved to: evs100/ToProcess/{xlsx_path_addr.name}")


if __name__ == "__main__":
    main()
