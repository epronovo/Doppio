# ASR_UpdateBasePrice.py
"""
ASR_UpdateBasePrice.py
----------------------
Identifies base price records in OIS017 whose LVDT (last valid date) would be
set to today's date, but only for items listed in ASR_ItemFilter.xlsx.
Instead of calling OIS017MI.UpdBasePrice directly, the candidate records are
written to an EVS100-compatible Excel file for review and optional submission.

Steps
-----
1.  Load item filter from ASR_ItemFilter.xlsx (sheet 1, column "ITEM").
2.  Call EXPORTMI.Select to list all price lists
    (OJPRRF, OJCUCD, OJFVDT from OPRICH).
3.  For each price list, call OIS017MI.LstBasePrice to retrieve
    PRRF, CUCD, FVDT, ITNO, VFDT, LVDT.
4.  For each matching item (in filter), check if today > LVDT.
      - Already expired → skip.
      - Otherwise → add to candidates with LVDT = today.
5.  Export candidate records to an EVS100-format Excel file
    (evs100/ToProcess/API_OIS017MI_<timestamp>.xlsx) with a Control sheet
    and a data sheet named API_OIS017MI_UpdBasePrice.
6.  Print summary, then optionally:
      a. Upload the file to M3 via the File Management REST API (PUT).
      b. Trigger processing via EVS100MI.ImportFile.

Usage:
    python ASR_UpdateBasePrice.py
"""

import datetime
import os
import requests
from pathlib import Path

from tqdm import tqdm
import xlsxwriter
from openpyxl import load_workbook

from InforMI import (
    CONFIG,
    get_ion_token,
    post_to_m3,
)
from UserDefaults import load_user_defaults, save_user_defaults

# =============================================================================
# Constants
# =============================================================================

EXPORTMI_SEP     = "^"
ITEM_FILTER_FILE = Path(__file__).parent / "ASR_ItemFilter.xlsx"
EVS100_TO_PROCESS = Path(__file__).parent / "evs100" / "ToProcess"
# Fields used in the candidate report
UPDATE_FIELDS = ["PRRF", "CUCD", "FVDT", "ITNO", "VFDT", "LVDT"]


# =============================================================================
# Tenant setup — always prompts, bypasses the 1-hour cache in InforMI
# =============================================================================

def _select_ionapi_forced(ionapi_dir: Path) -> Path:
    files = sorted(
        [f for f in os.listdir(ionapi_dir) if f.endswith(".ionapi")],
        key=str.lower,
    )
    if not files:
        raise FileNotFoundError(f"No .ionapi files found in: {ionapi_dir}")

    print("\n  🔹  Select ION API file:")
    for i, f in enumerate(files, start=1):
        print(f"    {i}. {f}")

    while True:
        try:
            choice = int(input(f"\n  Enter choice (1-{len(files)}): "))
            if 1 <= choice <= len(files):
                return ionapi_dir / files[choice - 1]
            print("  ❌  Invalid choice. Try again.")
        except ValueError:
            print("  ❌  Please enter a number.")


def _prompt_company_division_forced() -> None:
    defaults = load_user_defaults()
    default_company  = defaults.get("company",  "100")
    default_division = defaults.get("division", "500")

    company  = input(f"    Company  (default: {default_company}):  ").strip()
    division = input(f"    Division (default: {default_division}): ").strip()

    CONFIG["company"]  = company  if company  else default_company
    CONFIG["division"] = division if division else default_division

    defaults["company"]  = CONFIG["company"]
    defaults["division"] = CONFIG["division"]
    save_user_defaults(defaults)


# =============================================================================
# Step 1 – Load filter criteria from Excel
# =============================================================================

def load_item_filter(path: Path) -> set[str]:
    """
    Reads the first sheet of ASR_ItemFilter.xlsx and returns a set of ITEM
    values.  The first row is treated as a header and is skipped.
    """
    if not path.is_file():
        raise FileNotFoundError(f"Item filter workbook not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    items: set[str] = set()
    first_row = True
    for row in ws.iter_rows(values_only=True):
        if first_row:
            first_row = False
            continue  # skip header
        val = row[0]
        if val is not None and str(val).strip():
            items.add(str(val).strip())

    wb.close()
    print(f"  ✅  Loaded {len(items):,} items from filter: {path.name}")
    return items


# =============================================================================
# Step 2 – EXPORTMI.Select → parse REPL rows
# =============================================================================

def fetch_exportmi(
    session: requests.Session,
    query: str,
    label: str,
) -> list[dict]:
    """
    Calls EXPORTMI.Select with the given query and returns a list of dicts,
    one per data row, keyed by the column names found in the HDRS row.
    Handles 401 token refresh automatically.
    """
    list_url = CONFIG["api_url"]

    payload = {
        "program": "EXPORTMI",
        "transactions": [{
            "transaction": "Select",
            "record": {
                "QERY": query,
                "SEPC": EXPORTMI_SEP,
                "HDRS": "1",
            },
            "selectedColumns": ["QERY", "SEPC", "HDRS", "REPL"],
        }],
    }

    headers = {
        "Authorization": f"Bearer {CONFIG['access_token']}",
        "Content-Type": "application/json",
    }

    response = session.post(list_url, json=payload, headers=headers)
    if response.status_code == 401:
        get_ion_token()
        headers["Authorization"] = f"Bearer {CONFIG['access_token']}"
        response = session.post(list_url, json=payload, headers=headers)

    response.raise_for_status()
    data = response.json()

    repl_rows: list[str] = []
    for result in data.get("results", []):
        for record in result.get("records", []):
            repl_val = record.get("REPL", "")
            if repl_val:
                repl_rows.append(repl_val)

    if not repl_rows:
        print(f"  ⚠️   No rows returned from EXPORTMI ({label}).")
        return []

    # First row is the column header (HDRS:1)
    col_names = [c.strip() for c in repl_rows[0].rstrip(EXPORTMI_SEP).split(EXPORTMI_SEP)]

    parsed: list[dict] = []
    for raw in repl_rows[1:]:
        values = raw.rstrip(EXPORTMI_SEP).split(EXPORTMI_SEP)
        values += [""] * (len(col_names) - len(values))
        row = {col_names[i]: values[i].strip() for i in range(len(col_names))}
        parsed.append(row)

    print(f"  ✅  {len(parsed):,} records retrieved ({label}).")
    return parsed


# =============================================================================
# Step 3 – OIS017MI.LstBasePrice per price list
# =============================================================================

def fetch_lst_base_price(
    session: requests.Session,
    prrf: str,
    cucd: str,
    fvdt: str,
) -> list[dict]:
    """
    Calls OIS017MI.LstBasePrice for a single (PRRF, CUCD, FVDT) combination.
    Returns a list of dicts with keys: PRRF, CUCD, FVDT, ITNO, VFDT, LVDT.
    Returns an empty list on error (error is printed but not raised).
    """
    payload = {
        "program": "OIS017MI",
        "transactions": [{
            "transaction": "LstBasePrice",
            "record": {
                "PRRF": prrf,
                "CUCD": cucd,
                "FVDT": fvdt,
            },
            "selectedColumns": ["PRRF", "CUCD", "FVDT", "ITNO", "VFDT", "LVDT"],
        }],
    }

    try:
        result = post_to_m3(payload, session)
    except Exception as exc:
        tqdm.write(
            f"  ❌  LstBasePrice failed – PRRF={prrf} CUCD={cucd} FVDT={fvdt}: {exc}"
        )
        return []

    rows: list[dict] = []
    for res in result.get("results", []):
        for rec in res.get("records", []):
            rows.append({
                "PRRF": rec.get("PRRF", ""),
                "CUCD": rec.get("CUCD", ""),
                "FVDT": rec.get("FVDT", ""),
                "ITNO": rec.get("ITNO", ""),
                "VFDT": rec.get("VFDT", ""),
                "LVDT": rec.get("LVDT", ""),
            })
    return rows


# =============================================================================
# Step 4b – Upload file to M3 File Management (PUT)
# =============================================================================

def upload_file_to_m3(
    file_path: Path,
    session: requests.Session,
) -> bool:
    """
    Uploads file_path to the M3 FileImport area via the File Management REST API.

    PUT {iu}/{ti}/M3/foundation-rest/file-management/v1/file/FileImport/{filename}

    Returns True on success, False on failure.
    """
    filename   = file_path.name
    upload_url = (
        f"{CONFIG['iu']}/{CONFIG['ti']}"
        f"/M3/foundation-rest/file-management/v1/file/FileImport/{filename}"
    )

    headers = {
        "Authorization": f"Bearer {CONFIG['access_token']}",
        "Content-Type": "application/octet-stream",
    }

    with open(file_path, "rb") as fh:
        file_bytes = fh.read()

    response = session.put(upload_url, data=file_bytes, headers=headers)

    if response.status_code == 401:
        get_ion_token()
        headers["Authorization"] = f"Bearer {CONFIG['access_token']}"
        response = session.put(upload_url, data=file_bytes, headers=headers)

    if response.status_code in (200, 201, 204):
        print(f"  ✅  Upload succeeded (HTTP {response.status_code}): {filename}")
        return True
    else:
        print(
            f"  ❌  Upload failed (HTTP {response.status_code}): "
            f"{response.text[:200]}"
        )
        return False


# =============================================================================
# Step 4c – Trigger EVS100MI.ImportFile
# =============================================================================

def process_file_in_m3(
    filename: str,
    session: requests.Session,
) -> bool:
    """
    Calls EVS100MI.ImportFile with FNAM=filename so M3 processes the uploaded
    spreadsheet through the EVS100 interface.

    Returns True on success, False on failure.
    """
    payload = {
        "program": "EVS100MI",
        "transactions": [{
            "transaction": "ImportFile",
            "record": {"FNAM": filename},
        }],
    }

    try:
        result = post_to_m3(payload, session)
        # Check for an API-level error in the response
        for res in result.get("results", []):
            err = res.get("errorMessage", "").strip() if isinstance(res, dict) else ""
            if err:
                print(f"  ❌  EVS100MI.ImportFile error: {err}")
                return False
        print(f"  ✅  EVS100MI.ImportFile processed successfully: {filename}")
        return True
    except Exception as exc:
        print(f"  ❌  EVS100MI.ImportFile failed: {exc}")
        return False


# =============================================================================
# Summary and Candidate Export
# =============================================================================

def print_summary(
    total_price_lists: int,
    total_checked: int,
    total_skipped_not_in_filter: int,
    total_skipped_already_expired: int,
    candidates: list[dict],
) -> None:
    print(f"\n{'═' * 60}")
    print(f"  ASR_UpdateBasePrice – Run Summary")
    print(f"{'═' * 60}")
    print(f"  📋  Price lists processed        : {total_price_lists:,}")
    print(f"  🔍  Items checked                : {total_checked:,}")
    print(f"  ⏭️   Skipped (not in filter)      : {total_skipped_not_in_filter:,}")
    print(f"  ⏭️   Skipped (already expired)    : {total_skipped_already_expired:,}")
    print(f"  📄  Candidates written to Excel  : {len(candidates):,}")
    print(f"{'═' * 60}")


def export_evs100_xlsx(
    candidates: list[dict],
    out_dir: Path,
) -> Path:
    """
    Writes candidate records in the EVS100 import format expected by EVS100.py.

    Layout
    ------
    Control sheet (tab name: "Control"):
        Row 1 header : Worksheet | Description | Data
        Row 2 entry  : API_OIS017MI_UpdBasePrice | Sales Price List interface | x

    Data sheet (tab name: "API_OIS017MI_UpdBasePrice"):
        Row 1  – field names  : MESSAGE | PRRF | CUCD | FVDT | ITNO | VFDT | LVDT
        Row 2  – descriptions : <blank> | Pricelist (A:10) | Currency (A:3) | ...
        Row 3  – required     : no | yes | yes | yes | yes | yes | yes
        Row 4+ – data rows    : <blank> | <PRRF> | <CUCD> | <FVDT> | <ITNO> | ...

    Returns the path of the written file.
    """
    DATA_SHEET   = "API_OIS017MI_UpdBasePrice"
    DESCRIPTIONS = [
        None,
        "Pricelist (A:10)",
        "Currency (A:3)",
        "Validfrom (D:10)",
        "Itemnumber (A:15)",
        "Valid from (D:10)",
        "Validto (D:10)",
    ]
    REQUIRED = ["no", "yes", "yes", "yes", "yes", "yes", "yes"]
    COLUMNS  = ["MESSAGE"] + UPDATE_FIELDS   # MESSAGE first, then PRRF CUCD FVDT ITNO VFDT LVDT

    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"API_OIS017MI_{ts}.xlsx"

    wb = xlsxwriter.Workbook(str(out_path))

    # ── Control sheet ─────────────────────────────────────────────────────
    ws_ctrl = wb.add_worksheet("Control")
    for col, val in enumerate(["Worksheet", "Description", "Data"]):
        ws_ctrl.write(0, col, val)
    for col, val in enumerate([DATA_SHEET, "Sales Price List interface", "x"]):
        ws_ctrl.write(1, col, val)

    # ── Data sheet ────────────────────────────────────────────────────────
    ws_data = wb.add_worksheet(DATA_SHEET)

    # Row 1 – column names
    for col, name in enumerate(COLUMNS):
        ws_data.write(0, col, name)

    # Row 2 – field descriptions
    for col, desc in enumerate(DESCRIPTIONS):
        if desc is not None:
            ws_data.write(1, col, desc)

    # Row 3 – required flags
    for col, req in enumerate(REQUIRED):
        ws_data.write(2, col, req)

    # Rows 4+ – data (MESSAGE column is blank / None)
    for row_idx, rec in enumerate(candidates, start=3):
        for col, field in enumerate(UPDATE_FIELDS, start=1):
            val = rec.get(field, "")
            if val:
                ws_data.write(row_idx, col, val)

    wb.close()
    return out_path


# =============================================================================
# Main
# =============================================================================

def update_base_price() -> None:
    ionapi_dir = Path(__file__).parent / "ionapi"
    today_str  = datetime.date.today().strftime("%Y%m%d")   # e.g. "20260325"
    today_int  = int(today_str)

    print(f"\n{'═' * 60}")
    print(f"  ASR_UpdateBasePrice")
    print(f"  Today : {today_str}")
    print(f"{'═' * 60}")

    # ------------------------------------------------------------------ #
    # Step 1 – Load item filter                                           #
    # ------------------------------------------------------------------ #
    print("\n📂  Step 1 – Loading item filter …")
    item_filter = load_item_filter(ITEM_FILTER_FILE)

    if not item_filter:
        print("⚠️   Item filter is empty. Exiting.")
        return

    # ------------------------------------------------------------------ #
    # Configure SOURCE tenant                                             #
    # ------------------------------------------------------------------ #
    CONFIG["tenant"] = _select_ionapi_forced(ionapi_dir)
    _prompt_company_division_forced()
    get_ion_token()

    # ------------------------------------------------------------------ #
    # Step 2 – EXPORTMI: list all price lists                             #
    # ------------------------------------------------------------------ #
    print("\n🔍  Step 2 – EXPORTMI.Select: listing all price lists …")
    with requests.Session() as session:
        price_lists = fetch_exportmi(
            session,
            "OJPRRF,OJCUCD,OJFVDT from OPRICH",
            label="OPRICH price lists",
        )

    if not price_lists:
        print("⚠️   No price lists found. Exiting.")
        return

    # ------------------------------------------------------------------ #
    # Step 3 – OIS017MI.LstBasePrice for each price list                 #
    # Collect candidates matching the item filter and not yet expired     #
    # ------------------------------------------------------------------ #
    print(
        f"\n🔍  Step 3 – OIS017MI.LstBasePrice for "
        f"{len(price_lists):,} price list(s) …"
    )

    candidates: list[dict]  = []
    total_checked           = 0
    total_skipped_filter    = 0
    total_skipped_expired   = 0

    with requests.Session() as session:
        for pl in tqdm(price_lists, desc="Listing base prices", unit="price list"):
            prrf = pl.get("OJPRRF", "")
            cucd = pl.get("OJCUCD", "")
            fvdt = pl.get("OJFVDT", "")

            rows = fetch_lst_base_price(session, prrf, cucd, fvdt)
            total_checked += len(rows)

            for row in rows:
                itno = row.get("ITNO", "")
                lvdt = row.get("LVDT", "")

                # Only process items present in the filter
                if itno not in item_filter:
                    total_skipped_filter += 1
                    continue

                # Skip records already expired (today is past LVDT)
                if lvdt and int(lvdt) < today_int:
                    total_skipped_expired += 1
                    continue

                # Queue for update: set LVDT to today
                candidates.append({
                    "PRRF": prrf,
                    "CUCD": cucd,
                    "FVDT": fvdt,
                    "ITNO": itno,
                    "VFDT": row.get("VFDT", ""),
                    "LVDT": today_str,
                })

    print(f"\n  📊  Total records checked               : {total_checked:,}")
    print(f"  📊  Skipped – not in filter             : {total_skipped_filter:,}")
    print(f"  📊  Skipped – already expired           : {total_skipped_expired:,}")
    print(f"  📊  Candidates to update (LVDT={today_str}): {len(candidates):,}")

    if not candidates:
        print("\n✅  No records require updating. Done.")
        return

    # Preview the first 10 records
    fields  = UPDATE_FIELDS
    col_w   = {
        f: max(len(f), max((len(r.get(f, "")) for r in candidates), default=0))
        for f in fields
    }
    hdr_line = "  " + "  ".join(f.ljust(col_w[f]) for f in fields)
    sep_line = "  " + "  ".join("─" * col_w[f] for f in fields)

    print(f"\n{'═' * len(hdr_line)}")
    print(f"  📋  Candidates to export – {len(candidates):,} record(s)")
    print(f"{'═' * len(hdr_line)}")
    print(hdr_line)
    print(sep_line)
    for rec in candidates[:10]:
        print("  " + "  ".join(rec.get(f, "").ljust(col_w[f]) for f in fields))
    if len(candidates) > 10:
        print(f"  … and {len(candidates) - 10:,} more record(s)")
    print(f"{'═' * len(hdr_line)}")

    # ------------------------------------------------------------------ #
    # Step 4 – Export candidates to EVS100 Excel file                    #
    # ------------------------------------------------------------------ #
    print(f"\n▶   Step 4 – Exporting {len(candidates):,} candidate record(s) to EVS100 Excel file …")
    xlsx_path = export_evs100_xlsx(candidates, EVS100_TO_PROCESS)
    print(f"  📄  Saved to: evs100/ToProcess/{xlsx_path.name}")

    # ------------------------------------------------------------------ #
    # Step 5 – Summary                                                    #
    # ------------------------------------------------------------------ #
    print_summary(
        total_price_lists=len(price_lists),
        total_checked=total_checked,
        total_skipped_not_in_filter=total_skipped_filter,
        total_skipped_already_expired=total_skipped_expired,
        candidates=candidates,
    )

    # ------------------------------------------------------------------ #
    # Step 6 – Optional: upload and process in M3                        #
    # ------------------------------------------------------------------ #
    send = input("\n  Send this file to M3? [y/N]: ").strip().lower()
    if send != "y":
        print("  ℹ️   File not sent. Done.")
        return

    get_ion_token()  # refresh token before upload
    with requests.Session() as session:
        print(f"\n▶   Step 6a – Uploading {xlsx_path.name} to M3 …")
        uploaded = upload_file_to_m3(xlsx_path, session)

        if uploaded:
            trigger = input(
                "\n  Process file in M3 (EVS100MI.ImportFile)? [y/N]: "
            ).strip().lower()
            if trigger == "y":
                print(f"\n▶   Step 6b – Triggering EVS100MI.ImportFile …")
                process_file_in_m3(xlsx_path.name, session)
            else:
                print("  ℹ️   File uploaded but not processed. Done.")


# =============================================================================
# Entry Point
# =============================================================================

if __name__ == "__main__":
    update_base_price()
