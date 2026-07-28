import openpyxl
import os
import re
import sqlite3

from openpyxl.utils import get_column_letter
from config import BASE_DIR, EVS100_DIR, get_sqlite_db_path
from InforMI import prompt_for_company_division

# initialize field
print(f"")
QUERIES_DIR = BASE_DIR / "queries/evs"
prompt_for_company_division()
SQLITE_DB_PATH = get_sqlite_db_path()

def run_query(conn, query):
    cursor = conn.cursor()
    cursor.execute(query)

    # Only return data if it's a SELECT-like query
    if cursor.description is None:
        return None, None  # Indicates no result set
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    return columns, rows

def write_to_excel(columns, rows, output_path, sheet_name):
    wb = openpyxl.Workbook()

    # Clean sheet name
    sheet_name_cleaned = re.sub(r'_\[[^]]*\]$', '', sheet_name)[:31]
    ws_data = wb.active
    ws_data.title = sheet_name_cleaned

    # === Row 1: Header ===
    for col_idx, col_name in enumerate(columns, 1):
        ws_data.cell(row=1, column=col_idx, value=col_name)

    # === Row 2: Blank row ===
    # (nothing to do — openpyxl will leave it blank if you skip writing)

    # === Row 3: Blank under MESSAGE, "no" under others ===
    for col_idx, col_name in enumerate(columns, 1):
        value = "no"
        if col_name != "MESSAGE":
            value = "yes"
        ws_data.cell(row=3, column=col_idx, value=value)

    # === Data rows starting from row 4 ===
    for row_idx, row in enumerate(rows, start=4):
        for col_idx, value in enumerate(row, 1):
            ws_data.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns for the data sheet
    for col_idx, col_name in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        # Adjust width based on header length, with a minimum
        width = max(12, len(str(col_name)) + 2) 
        ws_data.column_dimensions[col_letter].width = width

    # --- Control Sheet ---
    
    # Add Control sheet
    ws_control = wb.create_sheet("Control")

    # Move it to the first position
    wb.move_sheet(ws_control, offset=-len(wb.sheetnames) + 1)

    # Write control content
    ws_control.cell(row=1, column=1, value="Worksheet")
    ws_control.cell(row=1, column=2, value="Description")
    ws_control.cell(row=1, column=3, value="Data")
    ws_control.cell(row=2, column=1, value=sheet_name_cleaned)
    ws_control.cell(row=2, column=2, value=sheet_name_cleaned)
    ws_control.cell(row=2, column=3, value="x")

    # Auto-size columns for the control sheet
    for col_idx in range(1, 4):
        col_letter = get_column_letter(col_idx)
        ws_control.column_dimensions[col_letter].width = 20
        
    wb.save(output_path)

def reverse_string(s):
    return s[::-1] if s else s

def process_sql_files(conn):
    archive_dir = QUERIES_DIR / "_archive"
    os.makedirs(archive_dir, exist_ok=True)
    
    for filename in os.listdir(QUERIES_DIR):
        if not filename.lower().endswith(".sql"):
            continue

        sql_path = os.path.join(QUERIES_DIR, filename)
        with open(sql_path, 'r', encoding='utf-8') as f:
            sql = f.read()

        base_name = os.path.splitext(filename)[0]
        output_dir = EVS100_DIR / "ToProcess"

        os.makedirs(output_dir, exist_ok=True)  # Ensure the folder exists
        output_path = os.path.join(output_dir, base_name + ".xlsx")

        print(f"Processing: {filename}")
        try:
            columns, rows = run_query(conn, sql)

            if columns is None or rows is None:
                print(f"ℹ️ Query did not return a result set (e.g., CREATE/INSERT/UPDATE), skipping Excel export.")
            else:
                print(f"#️⃣  {len(rows)} rows returned")
                write_to_excel(columns, rows, output_path, base_name)
                print(f"💾 Saved: {output_path}")

            # ✅ Move .sql file to _archive
            archived_path = archive_dir / filename
            os.rename(sql_path, archived_path)
            print(f"📦 Archived: {archived_path}")

        except Exception as e:
            print(f"❌ Error: {e}")

    print("✅ All queries processed.")

if __name__ == "__main__":
    conn = sqlite3.connect(SQLITE_DB_PATH)
    conn.create_function("REVERSE", 1, reverse_string)  # Register REVERSE in the same connection
    process_sql_files(conn)
    conn.close()