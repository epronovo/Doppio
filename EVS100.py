# EVS100.py
import pandas as pd
import requests
import shutil
import time
from math import ceil
from openpyxl import load_workbook
from tqdm import tqdm
from tqdm.contrib.concurrent import thread_map
from pathlib import Path

from config import EVS100_DIR
from InforMI import CONFIG, get_ion_token, post_to_m3, select_ionapi_file, prompt_for_company_division
from UserDefaults import load_user_defaults, save_user_defaults

INPUT_DIR = EVS100_DIR / "ToProcess"
OUTPUT_DIR = EVS100_DIR / "Complete"

# =============================================================================
# Helper Functions
# =============================================================================
def prompt_with_default(prompt_text, default_value, cast_func=str, valid_func=None):
    while True:
        user_input = input(f"{prompt_text} (default: {default_value}): ").strip()
        if not user_input:
            return default_value
        try:
            value = cast_func(user_input)
            if valid_func and not valid_func(value):
                tqdm.write("Invalid input. Try again.")
                continue
            return value
        except Exception as e:
            tqdm.write(f"Invalid input: {e}. Try again.")

def read_control_sheet(path):
    df = pd.read_excel(path, sheet_name="Control", header=0, dtype=str)
    return df[df["Data"].str.lower() == 'x']

def read_data_sheet_with_headers(path, sheet_name):
    df_all = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=str).fillna("")
    header_row_1 = df_all.iloc[0:1].copy()
    header_row_2 = df_all.iloc[1:2].copy()
    header_row_3 = df_all.iloc[2:3].copy()
    df_data = df_all.iloc[3:].copy()
    df_data.columns = df_all.iloc[0]
    header_row_1.columns = df_data.columns
    header_row_2.columns = df_data.columns
    header_row_3.columns = df_data.columns
    return df_data.reset_index(drop=True), header_row_1.reset_index(drop=True), header_row_2.reset_index(drop=True), header_row_3.reset_index(drop=True)

def build_payload(file_path, sheet_name):
    program, transaction = sheet_name.replace("API_", "").split("_", 1)
    wb = load_workbook(file_path, read_only=True)
    ws = wb[sheet_name]
    all_columns = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    yes_no_row = [cell.value for cell in next(ws.iter_rows(min_row=3, max_row=3))]
    wb.close()
    selected_columns = [col for col, yn in zip(all_columns, yes_no_row) if str(yn).strip().lower() == "yes"]
    df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None, skiprows=3, dtype=str)
    df_raw = df_raw.iloc[:, [i for i, yn in enumerate(yes_no_row) if str(yn).strip().lower() == "yes"]]
    df_raw.columns = selected_columns
    df_raw = df_raw.fillna("")
    records = df_raw.to_dict(orient="records")
    return program, transaction, records, selected_columns

def annotate_errors(df, results):
    df = df.copy()
    error_messages = []
    for idx, res in enumerate(results):
        error_messages.append(res.get("errorMessage", "") if isinstance(res, dict) else "")
        if not isinstance(res, dict):
            continue
        records = res.get("records", [])
        if records and isinstance(records[0], dict):
            for key, val in records[0].items():
                if key not in df.columns:
                    df[key] = ""
                df.at[idx, key] = val
    if "MESSAGE" in df.columns:
        df["MESSAGE"] = error_messages
    else:
        df.insert(0, "MESSAGE", error_messages)
    return df, len(error_messages), sum(bool(msg) for msg in error_messages)

def write_output_xlsx(path, sheets):
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

# =============================================================================
# Core Workbook Processing
# =============================================================================
def process_workbook(file_path, errors_only=True, thread_index=0):
    control_df = read_control_sheet(file_path)
    output_sheets = {"Control": control_df}
    total_records = 0
    total_errors = 0

    with requests.Session() as session:
        for _, row in control_df.iterrows():
            sheet_name = row["Worksheet"]
            df_original, hdr1, hdr2, hdr3 = read_data_sheet_with_headers(file_path, sheet_name)
            program, transaction, records, selected_columns = build_payload(file_path, sheet_name)
            total_records += len(records)

            batch_results = []
            num_batches = ceil(len(records) / CONFIG["batch_size"])

            for i in tqdm(
                range(0, len(records), CONFIG["batch_size"]),
                total=num_batches,
                desc=f"{file_path.name} - {sheet_name}",
                leave=False,
                position=thread_index
            ):
                batch = records[i:i + CONFIG["batch_size"]]
                transactions = [{"transaction": transaction, "record": r, "selectedColumns": selected_columns} for r in batch]
                payload = {"program": program, "transactions": transactions}
                result = post_to_m3(payload, session)
                batch_results.extend(result["results"])

            output_df, _, error_count = annotate_errors(df_original.copy(), batch_results)
            columns = output_df.columns

            if "MESSAGE" not in hdr2.columns:
                hdr2.insert(0, "MESSAGE", "Error Message")
                hdr3.insert(0, "MESSAGE", "yes")

            hdr2 = hdr2.reindex(columns=columns, fill_value="")
            hdr3 = hdr3.reindex(columns=columns, fill_value="")
            output_df = pd.concat([hdr2, hdr3, output_df], ignore_index=True)

            if errors_only:
                output_df = output_df[output_df["MESSAGE"].astype(str).str.strip() != ""].reset_index(drop=True)

            total_errors += error_count
            output_sheets[sheet_name] = output_df

    postfix = f"_p{total_records}_e{total_errors}"
    out_path = OUTPUT_DIR / f"{file_path.stem}{postfix}{file_path.suffix}"
    write_output_xlsx(out_path, output_sheets)
    shutil.move(str(file_path), OUTPUT_DIR / file_path.name)

def process_workbook_with_progress(file_path, errors_only=True, thread_index=0):
    tqdm.write(f"🛠️  Processing {file_path.name}...")
    process_workbook(file_path, errors_only=errors_only, thread_index=thread_index)
    tqdm.write(f"✅ Completed {file_path.name}")

# =============================================================================
# Main Processing Function
# =============================================================================
def process_all_workbooks():
    ionapi_dir = Path(__file__).parent / "ionapi"
    CONFIG["tenant"] = select_ionapi_file(ionapi_dir)
    prompt_for_company_division()

    defaults = load_user_defaults()
    max_workers = int(input(f"Enter number of workers (default {defaults.get('max_workers',5)}): ") or defaults.get("max_workers",5))
    errors_only = input(f"Show only errors? (true/false, default {defaults.get('errors_only',True)}): ").lower()
    errors_only = {"true": True, "false": False}.get(errors_only, defaults.get("errors_only", True))

    defaults["max_workers"] = max_workers
    defaults["errors_only"] = errors_only
    save_user_defaults(defaults)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    files = sorted(
        [f for f in INPUT_DIR.glob("*.xlsx") if not f.name.startswith("~")],
        key=lambda f: f.name.lower()
    )
    if not files:
        tqdm.write("📂 No Excel files to process.")
        return

    max_workers = min(max_workers, len(files))
    start_time = time.time()
    token = get_ion_token()

    tqdm.write("\n🔍 Processing workbooks...")

    def wrapper_with_index(args):
        file_path, idx = args
        return process_workbook_with_progress(file_path, errors_only=errors_only, thread_index=idx)

    # Thread-safe concurrent processing with progress bar
    thread_map(wrapper_with_index, [(f, i) for i, f in enumerate(files)],
               max_workers=max_workers, desc="Workbooks")

    elapsed = time.time() - start_time
    hours, rem = divmod(elapsed, 3600)
    minutes, seconds = divmod(rem, 60)
    tqdm.write(f"\n✅ Total run time: {int(hours):02d}:{int(minutes):02d}:{int(seconds):02d} (hh:mm:ss)")

# =============================================================================
# Entry Point
# =============================================================================
if __name__ == "__main__":
    process_all_workbooks()