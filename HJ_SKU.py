# HJ_SKU.py
# -----------------------------------------------------------------------
# PURPOSE
#   Reads a Syndigo SKU CSV export from the input folder,
#   applies field mappings for MMS200MI.AddItmBasic,
#   and writes a styled multi-tab Excel workbook to the output folder.
#
#   Tabs produced
#     API_MMS200MI_AddItmBasic  — Item master (Basic)
#     API_MMS200MI_AddItmWhs    — Item warehouse
#     API_MMS200MI_UpdItmFac    — Item facility
#
#   The workbook follows the MpGd_Build_Template layout:
#     Row 1 – API field names          (light blue)
#     Row 2 – Field description        (light blue)
#     Row 3 – Required flag            (light blue)
#     Row 4 – Mapping / source note    (gray, bold)
#     Row 5 – Example (first record)   (italic)
#     Row 6+ – Data rows from CSV
#
# INPUTS
#   ./input/*.csv   — Syndigo SKU export
#
# OUTPUTS
#   ./output/HJ_SKU_{YYYYMMDD}_{HHMMSS}.xlsx
#
# USAGE
#   python HJ_SKU.py [--input-folder PATH] [--output-folder PATH] [--db PATH]
# -----------------------------------------------------------------------

import argparse
import glob
import logging
import os
import sqlite3
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
SCRIPT_DIR            = os.path.dirname(os.path.abspath(__file__))
DEFAULT_INPUT_FOLDER  = os.path.join(SCRIPT_DIR, "input")
DEFAULT_OUTPUT_FOLDER = os.path.join(SCRIPT_DIR, "output")
DEFAULT_DB_PATH       = "/Users/ericpronovost/sqlite/doppio.db"

# SQL template written to SCRIPT_DIR for each API tab
_SQL_FILES = {
    "API_MMS200MI_AddItmBasic": (
        "SELECT 'MMS200MI'[minm],'AddItmBasic'[trnm],* FROM (\n"
        "\tSELECT thgskuid[ITNO],\n"
        "\tthgskuname[ITDS],\n"
        "\tthgskudescription[FUDS],\n"
        "\tthgskustatus[STAT],\n"
        "\t'HJDATALOAD'[RESP],\n"
        "\tthguom[UNMS],\n"
        "\titemtype[ITTY],\n"
        "\tthgprimarymakebuy[MABU],\n"
        "\t'Y1'[PRVG],\n"
        "\titemgroup[ITGR],\n"
        "\tproductgroup[ITCL],\n"
        "\tthghjlegacyerpproductnumber[ITNE]\n"
        "\tFROM SyndigoSKU\n"
        ") as temp\n"
        "WHERE 1=1\n"
        "AND ITNO <> '' AND ITDS <> ''\n"
    ),
    "API_MMS200MI_AddItmWhs": (
        "SELECT 'MMS200MI'[minm],'AddItmWhs'[trnm],* FROM (\n"
        "\tSELECT '100'[WHLO],\n"
        "\tthgskuid[ITNO],\n"
        "\t'2'[ALMT],\n"
        "\t'2'[STMT],\n"
        "\t'1'[SPMT],\n"
        "\t'HJDATALOAD'[RESP],\n"
        "\t'M1'[PLCD],\n"
        "\t'A'[ABCD],\n"
        "\t'A'[ABFC],\n"
        "\t'A'[ACOC],\n"
        "\t'1'[PUIT],\n"
        "\t'G10'[ORTY],\n"
        "\t'RI1'[FACI],\n"
        "\t'1'[STRL],\n"
        "\t'20'[STAT],\n"
        "\t'E01'[CPCD],\n"
        "\t'Y40001'[SUNO],\n"
        "\t'A1'[WHTY],\n"
        "\t'L'[WHSY],\n"
        "\t'1'[INSC],\n"
        "\t'APPROVED'[WHSL],\n"
        "\t'A'[MABC],\n"
        "\t'1'[ABCM],\n"
        "\t'2'[DPID],\n"
        "\t'DP'[PRCD],\n"
        "\t'01'[FCCM],\n"
        "\t'30'[PFTM],\n"
        "\t'2'[SVEI],\n"
        "\t'100'[PLHZ],\n"
        "\t'1'[OPLC],\n"
        "\t'1'[MSCH]\n"
        "\tFROM SyndigoSKU\n"
        ") as temp\n"
        "WHERE 1=1\n"
        "AND ITNO <> ''\n"
    ),
    "API_MMS200MI_UpdItmFac": (
        "SELECT 'MMS200MI'[minm],'UpdItmFac'[trnm],* FROM (\n"
        "\tSELECT 'RI1'[FACI],\n"
        "\tthgskuid[ITNO],\n"
        "\t'0'[LEA4],\n"
        "\tthghtcs[CSNO],\n"
        "\t'HJPCM1'[WSCA],\n"
        "\t'0'[CPL0],\n"
        "\t'0'[CPDC],\n"
        "\t'1'[VAMT],\n"
        "\t'0'[ALTS],\n"
        "\t'100'[REWH],\n"
        "\t'1'[DLET],\n"
        "\t'1'[MARC],\n"
        "\t'1'[FATM]\n"
        "\tFROM SyndigoSKU\n"
        ") as temp\n"
        "WHERE 1=1\n"
        "AND ITNO <> ''\n"
    ),
}

# (api_field, description, required, mapping_note)
FIELDS_MMS200 = [
    ("ITNO", "Item number",              "Yes", "thgskuid"),
    ("ITDS", "Name",                     "Yes", "thgskuname"),
    ("FUDS", "Description",              "No",  "thgskudescription"),
    ("STAT", "Status",                   "Yes", "thgskustatus"),
    ("RESP", "Responsible",              "Yes", '"HJDATALOAD" — hardcoded'),
    ("UNMS", "Unit of measure",          "Yes", "thguom"),
    ("ITTY", "Item type",                "Yes", "itemtype"),
    ("MABU", "Make/buy code",            "Yes", "thgprimarymakebuy"),
    ("PRVG", "Price group",              "No",  '"Y1" — hardcoded'),
    ("ITGR", "Item group",               "Yes", "itemgroup"),
    ("ITCL", "Product group",            "Yes", "productgroup"),
    ("ITNE", "Legacy ERP product number","No",  "thghjlegacyerpproductnumber"),
]

FIELDS_MMS200_WHS = [
    ("WHLO", "Warehouse",               "Yes", '"100" — hardcoded'),
    ("ITNO", "Item number",             "Yes", "thgskuid"),
    ("ALMT", "Allocation method",       "Yes", '"2" — hardcoded'),
    ("STMT", "Storage method",          "Yes", '"2" — hardcoded'),
    ("SPMT", "Picking method",          "Yes", '"1" — hardcoded'),
    ("RESP", "Responsible",             "Yes", '"HJDATALOAD" — hardcoded'),
    ("PLCD", "Planning method",         "Yes", '"M1" — hardcoded'),
    ("ABCD", "ABC class demand",        "Yes", '"A" — hardcoded'),
    ("ABFC", "ABC class frequency",     "Yes", '"A" — hardcoded'),
    ("ACOC", "ABC class contribution",  "Yes", '"A" — hardcoded'),
    ("PUIT", "Reorder method",          "Yes", '"1" — hardcoded'),
    ("ORTY", "Order type",              "Yes", '"G10" — hardcoded'),
    ("FACI", "Facility",                "Yes", '"RI1" — hardcoded'),
    ("STRL", "Storage location",        "Yes", '"1" — hardcoded'),
    ("STAT", "Status",                  "Yes", '"20" — hardcoded'),
    ("CPCD", "CTP policy",              "Yes", '"E01" — hardcoded'),
    ("SUNO", "Supplier number",         "Yes", '"Y40001" — hardcoded'),
    ("WHTY", "Warehouse type",          "Yes", '"A1" — hardcoded'),
    ("WHSY", "Warehouse subtype",       "Yes", '"L" — hardcoded'),
    ("INSC", "Inspection code",         "Yes", '"1" — hardcoded'),
    ("WHSL", "Location",                "Yes", '"APPROVED" — hardcoded'),
    ("MABC", "Manual ABC class",        "Yes", '"A" — hardcoded'),
    ("ABCM", "ABC method",              "Yes", '"1" — hardcoded'),
    ("DPID", "Display info",            "Yes", '"2" — hardcoded'),
    ("PRCD", "Price method",            "Yes", '"DP" — hardcoded'),
    ("FCCM", "Forecast control",        "Yes", '"01" — hardcoded'),
    ("PFTM", "Planning fence time",     "Yes", '"30" — hardcoded'),
    ("SVEI", "Shelf life method",       "Yes", '"2" — hardcoded'),
    ("PLHZ", "Planning horizon",        "Yes", '"100" — hardcoded'),
    ("OPLC", "OP planning method",      "Yes", '"1" — hardcoded'),
    ("MSCH", "Master scheduled",        "Yes", '"1" — hardcoded'),
]

FIELDS_MMS200_FAC = [
    ("FACI", "Facility",                "Yes", '"RI1" — hardcoded'),
    ("ITNO", "Item number",             "Yes", "thgskuid"),
    ("LEA4", "Administrative lead time","Yes", '"0" — hardcoded'),
    ("CSNO", "Customs number",          "Yes", "thghtcs"),
    ("WSCA", "Costing model",           "Yes", '"HJPCM1" — hardcoded'),
    ("CPL0", "Costing price",           "Yes", '"0" — hardcoded'),
    ("CPDC", "Costing price method",    "Yes", '"0" — hardcoded'),
    ("VAMT", "Costing value",           "Yes", '"1" — hardcoded'),
    ("ALTS", "Alternate structure",     "Yes", '"0" — hardcoded'),
    ("REWH", "Receiving warehouse",     "Yes", '"100" — hardcoded'),
    ("DLET", "Delete flag",             "Yes", '"1" — hardcoded'),
    ("MARC", "Manual update flag",      "Yes", '"1" — hardcoded'),
    ("FATM", "Facility item method",    "Yes", '"1" — hardcoded'),
]

# ---------------------------------------------------------------------------
# Field-value transformation lookups
# ---------------------------------------------------------------------------

_ITTY_MAP = {
    "component":                      "C01",
    "raw material":                   "C02",
    "packaging & presentation":       "C03",
    "tooling":                        "C04",
    "commercial items":               "C05",
    "component item":                 "D09",
    "branch (planning) items":        "D16",
    "manufacturing die":              "DIE",
    "fine paper":                     "F01",
    "fashion product - manufactured": "FS1",
    "jewelry":                        "J01",
    "jewelry phantoms":               "J02",
    "precious metals":                "M01",
    "regalia":                        "R01",
}

_ITGR_MAP = {
    "accessories":                 "IL030",
    "bags and sleeves":            "C3050",
    "capital equipment":           "C5090",
    "chains and neck ribbons":     "C1030",
    "cords":                       "IL020",
    "corregate":                   "C3010",
    "credentials & documents":     "IN010",
    "die":                         "C4010",
    "facilities & utilities":      "C5030",
    "fashion jewelry":             "GE040",
    "fees & charges":              "C5100",
    "findings":                    "C1010",
    "gifts & keepsakes":           "GE060",
    "hr & people services":        "C5060",
    "inserts and padding":         "C3040",
    "insignia":                    "GE010",
    "it & technology":             "C5050",
    "item group example":          "ZZZZZ",
    "mailers":                     "C3020",
    "marketing & promotional":     "C5070",
    "medals":                      "GE050",
    "metal":                       "C2020",
    "mold":                        "C4020",
    "mountings":                   "C1020",
    "non-inventory supplies":      "C5020",
    "operations support":          "C5040",
    "pins":                        "GE020",
    "plaque and trophy components":"C1040",
    "presentation box":            "C3030",
    "rings":                       "GE030",
    "services":                    "C5010",
    "stoles":                      "IL010",
    "stones":                      "C2010",
    "travel & events":             "C5080",
}

_ITCL_MAP = {
    "pins/posts":                   "C1011",
    "joint":                        "C1012",
    "connectors":                   "C1013",
    "closures":                     "C1014",
    "backs":                        "C1015",
    "wire/ stock":                  "C1016",
    "crimps/terminations":          "C1017",
    "stone settings":               "C1021",
    "mountings":                    "C1022",
    "crest mounting tubing":        "C1023",
    "pegs":                         "C1024",
    "bails":                        "C1025",
    "neck chains":                  "C1031",
    "neck ribbons":                 "C1032",
    "ribbon bar":                   "C1033",
    "boards":                       "C1041",
    "plates":                       "C1042",
    "seals & castings":             "C1043",
    "genuine stones":               "C2011",
    "synthetic stones":             "C2012",
    "grain":                        "C2021",
    "sheet/ flat stock":            "C2022",
    "wire":                         "C2023",
    "coil":                         "C2024",
    "solder paste":                 "C2025",
    "corregate":                    "C3011",
    "mailers":                      "C3021",
    "presentation box":             "C3031",
    "inserts & padding":            "C3041",
    "bags & sleeves":               "C3051",
    "top die":                      "C4011",
    "back die":                     "C4012",
    "trim tool":                    "C4013",
    "pierce tool":                  "C4014",
    "mold":                         "C4021",
    "professional services":        "C5011",
    "staffing & labor":             "C5012",
    "safety & compliance":          "C5013",
    "office supplies":              "C5021",
    "operations & mro":             "C5022",
    "safety supplies":              "C5023",
    "janitorial & maintenance":     "C5031",
    "utilities & overhead":         "C5032",
    "equipment & rental":           "C5033",
    "other":                        "C5034",
    "operations equipment":         "C5041",
    "external operations services": "C5042",
    "warehouse & fulfillment":      "C5043",
    "samples":                      "C5044",
    "hardware & equipment":         "C5051",
    "software & subscriptions":     "C5052",
    "it services":                  "C5053",
    "employee relations":           "C5061",
    "recruiting & staffing":        "C5062",
    "training & development":       "C5063",
    "printed materials":            "C5071",
    "promo materials":              "C5072",
    "marketing collateral":         "C5073",
    "catering":                     "C5081",
    "events":                       "C5082",
    "capital equipment":            "C5091",
    "fees & charges":               "C5101",
    "badges":                       "GE011",
    "guards":                       "GE012",
    "dangles":                      "GE013",
    "medical pins":                 "GE021",
    "professional pins":            "GE022",
    "education pins":               "GE023",
    "greek pins":                   "GE024",
    "championship rings":           "GE031",
    "class rings":                  "GE032",
    "fashion rings":                "GE033",
    "professional rings":           "GE034",
    "pendants":                     "GE041",
    "charms":                       "GE042",
    "cufflinks":                    "GE043",
    "tie bars / tie tac":           "GE044",
    "necklaces":                    "GE045",
    "bracelets":                    "GE046",
    "earrings":                     "GE047",
    "other medals":                 "GE051",
    "grad medals":                  "GE052",
    "ornaments":                    "GE061",
    "coins":                        "GE066",
    "key rings/ key chain":         "GE069",
    "lamps":                        "GE06C",
    "plaques":                      "GE06D",
    "trophies":                     "GE06E",
    "candles":                      "GE06F",
    "misc":                         "GE06G",
    "stoles":                       "IL011",
    "cords":                        "IL021",
    "accessories":                  "IL031",
    "certificates":                 "IN011",
    "product group example":        "ZZZ",
}


def _format_htcs(val: str) -> str:
    """Normalise an HTS code to ####.##.#### format; return '' for blank/zero."""
    v = val.strip()
    if not v or v.lstrip("0") == "":
        return ""
    digits = v.replace(".", "")
    if len(digits) == 10 and digits.isdigit():
        return f"{digits[:4]}.{digits[4:6]}.{digits[6:]}"
    return v  # already formatted or unexpected shape — pass through


def _lookup_itty(val: str) -> str:
    return _ITTY_MAP.get(val.lower(), val)


def _lookup_itgr(val: str) -> str:
    return _ITGR_MAP.get(val.lower(), "ZZZZZ")


def _lookup_itcl(val: str, itgr_raw: str) -> str:
    key = val.lower()
    # Rule 1: direct match
    if key in _ITCL_MAP:
        return _ITCL_MAP[key]
    # Rule 2: ITCL + ' ' + ITGR (raw source values)
    combined = (val.strip() + " " + itgr_raw.strip()).lower()
    if combined in _ITCL_MAP:
        return _ITCL_MAP[combined]
    # Rule 3: val is a substring of any CTTX40 key
    for cttx40_key, ctstky in _ITCL_MAP.items():
        if key and key in cttx40_key:
            return ctstky
    return "ZZZ"


# Colours (match MpGd_Build_Template.py)
FILL_BLUE  = PatternFill("solid", start_color="ADD8E6", end_color="ADD8E6")
FILL_GRAY  = PatternFill("solid", start_color="D3D3D3", end_color="D3D3D3")
FILL_WHITE = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
FILL_ALT   = PatternFill("solid", start_color="F5F5F5", end_color="F5F5F5")

FONT_DEFAULT = Font(name="Arial", size=10)
FONT_BOLD    = Font(name="Arial", size=10, bold=True)
FONT_ITALIC  = Font(name="Arial", size=10, italic=True)
FONT_LABEL   = Font(name="Arial", size=10, bold=True, color="444444")

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(
            os.path.join(SCRIPT_DIR, "HJ_SKU.log"),
            encoding="utf-8",
        ),
    ],
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def map_row_mms200(csv_row: pd.Series) -> dict:
    """Map one CSV row to MMS200MI.AddItmBasic fields."""
    unms_raw = (csv_row.get("thguom")            or "").strip()
    mabu_raw = (csv_row.get("thgprimarymakebuy") or "").strip()
    itty_raw = (csv_row.get("itemtype")          or "").strip()
    itgr_raw = (csv_row.get("itemgroup")         or "").strip()
    itcl_raw = (csv_row.get("productgroup")      or "").strip()

    unms = "EA" if unms_raw.lower() == "each" else unms_raw

    mabu_lower = mabu_raw.lower()
    if mabu_lower == "manufactured":
        mabu = "1"
    elif mabu_lower == "purchased":
        mabu = "2"
    else:
        mabu = mabu_raw

    return {
        "ITNO": (csv_row.get("thgskuid")                    or "").strip(),
        "ITDS": (csv_row.get("thgskuname")                  or "").strip(),
        "FUDS": (csv_row.get("thgskudescription")           or "").strip(),
        "STAT": (csv_row.get("thgskustatus")                or "").strip(),
        "RESP": "HJDATALOAD",
        "UNMS": unms,
        "ITTY": _lookup_itty(itty_raw),
        "MABU": mabu,
        "PRVG": "Y1",
        "ITGR": _lookup_itgr(itgr_raw),
        "ITCL": _lookup_itcl(itcl_raw, itgr_raw),
        "ITNE": (csv_row.get("thghjlegacyerpproductnumber") or "").strip(),
    }


def map_row_itmwhs(csv_row: pd.Series) -> dict:
    """Map one CSV row to MMS200MI.AddItmWhs fields (all hardcoded except ITNO)."""
    return {
        "WHLO": "100",
        "ITNO": (csv_row.get("thgskuid") or "").strip(),
        "ALMT": "2",
        "STMT": "2",
        "SPMT": "1",
        "RESP": "HJDATALOAD",
        "PLCD": "M1",
        "ABCD": "A",
        "ABFC": "A",
        "ACOC": "A",
        "PUIT": "1",
        "ORTY": "G10",
        "FACI": "RI1",
        "STRL": "1",
        "STAT": "20",
        "CPCD": "E01",
        "SUNO": "Y40001",
        "WHTY": "A1",
        "WHSY": "L",
        "INSC": "1",
        "WHSL": "APPROVED",
        "MABC": "A",
        "ABCM": "1",
        "DPID": "2",
        "PRCD": "DP",
        "FCCM": "01",
        "PFTM": "30",
        "SVEI": "2",
        "PLHZ": "100",
        "OPLC": "1",
        "MSCH": "1",
    }


def map_row_upditmfac(csv_row: pd.Series) -> dict:
    """Map one CSV row to MMS200MI.UpdItmFac fields."""
    return {
        "FACI": "RI1",
        "ITNO": (csv_row.get("thgskuid") or "").strip(),
        "LEA4": "0",
        "CSNO": _format_htcs(csv_row.get("thghtcs") or ""),
        "WSCA": "HJPCM1",
        "CPL0": "0",
        "CPDC": "0",
        "VAMT": "1",
        "ALTS": "0",
        "REWH": "100",
        "DLET": "1",
        "MARC": "1",
        "FATM": "1",
    }


def upsert_to_db(df: pd.DataFrame, db_path: str, source_file: str) -> int:
    """Upsert all rows into SyndigoSKU with values already transformed to M3 format."""
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    loaded_at = datetime.now().isoformat()
    rows = []
    for _, row in df.iterrows():
        unms_raw = (row.get("thguom")            or "").strip()
        mabu_raw = (row.get("thgprimarymakebuy") or "").strip()
        itgr_raw = (row.get("itemgroup")         or "").strip()
        itcl_raw = (row.get("productgroup")      or "").strip()

        unms = "EA" if unms_raw.lower() == "each" else unms_raw

        mabu_lower = mabu_raw.lower()
        if mabu_lower == "manufactured":
            mabu = "1"
        elif mabu_lower == "purchased":
            mabu = "2"
        else:
            mabu = mabu_raw

        rows.append((
            (row.get("thgskuid")                     or "").strip(),
            (row.get("thgskuname")                   or "").strip(),
            (row.get("thgskudescription")            or "").strip(),
            (row.get("thgskustatus")                 or "").strip(),
            unms,
            _lookup_itty((row.get("itemtype")        or "").strip()),
            mabu,
            _lookup_itgr(itgr_raw),
            _lookup_itcl(itcl_raw, itgr_raw),
            (row.get("thghjlegacyerpproductnumber")  or "").strip(),
            _format_htcs(row.get("thghtcs")          or ""),
            os.path.basename(source_file),
            loaded_at,
        ))
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS SyndigoSKU (
            thgskuid                        TEXT PRIMARY KEY,
            thgskuname                      TEXT,
            thgskudescription               TEXT,
            thgskustatus                    TEXT,
            thguom                          TEXT,
            itemtype                        TEXT,
            thgprimarymakebuy               TEXT,
            itemgroup                       TEXT,
            productgroup                    TEXT,
            thghjlegacyerpproductnumber     TEXT,
            thghtcs                         TEXT,
            _source_file                    TEXT,
            _loaded_at                      TEXT
        )
    """)
    try:
        conn.execute("ALTER TABLE SyndigoSKU ADD COLUMN thghtcs TEXT")
    except sqlite3.OperationalError:
        pass  # column already exists
    conn.executemany(
        """INSERT OR REPLACE INTO SyndigoSKU
           (thgskuid, thgskuname, thgskudescription, thgskustatus,
            thguom, itemtype, thgprimarymakebuy, itemgroup, productgroup,
            thghjlegacyerpproductnumber, thghtcs, _source_file, _loaded_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        rows,
    )
    conn.commit()
    conn.close()
    return len(rows)


def write_sql_files() -> None:
    """Write one SKU_{tab}.sql per API tab to the project root."""
    for tab_name, sql in _SQL_FILES.items():
        path = os.path.join(SCRIPT_DIR, f"SKU_{tab_name}.sql")
        with open(path, "w", encoding="utf-8") as fh:
            fh.write(sql)
        log.info(f"  Wrote: {os.path.basename(path)}")


def autofit(ws):
    """Set column widths based on content; wrap all cells."""
    for col_cells in ws.columns:
        max_len    = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")


def style_cell(cell, fill=None, font=None, h_align="left"):
    cell.fill      = fill or FILL_WHITE
    cell.font      = font or FONT_DEFAULT
    cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Sheet builder
# ---------------------------------------------------------------------------

def add_api_sheet(wb: Workbook, ws_name: str, fields: list, data_rows: list[dict], example_row: dict):
    """Add one styled API tab to wb."""
    ws = wb.create_sheet(title=ws_name)

    field_names  = [f[0] for f in fields]
    descriptions = [f[1] for f in fields]
    required     = [f[2] for f in fields]
    mappings     = [f[3] for f in fields]
    examples     = [example_row.get(fn, "") for fn in field_names]

    header_rows = [
        ("FIELD",    field_names,  FILL_BLUE,  FONT_BOLD,    "center"),
        ("Message",  descriptions, FILL_BLUE,  FONT_DEFAULT, "left"),
        ("Required", required,     FILL_BLUE,  FONT_DEFAULT, "center"),
        ("Mapping",  mappings,     FILL_GRAY,  FONT_BOLD,    "left"),
        ("Example",  examples,     FILL_WHITE, FONT_ITALIC,  "left"),
    ]

    for row_idx, (label, values, fill, font, align) in enumerate(header_rows, start=1):
        cell_a = ws.cell(row=row_idx, column=1, value=label)
        style_cell(cell_a, fill=fill, font=FONT_LABEL, h_align="right")
        for col_idx, val in enumerate(values, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            style_cell(cell, fill=fill, font=font, h_align=align)

    ws.freeze_panes = ws.cell(row=4, column=2)

    for i, record in enumerate(data_rows):
        row_num  = i + len(header_rows) + 1
        row_fill = FILL_ALT if i % 2 else FILL_WHITE
        ws.cell(row=row_num, column=1, value=i + 1).fill = row_fill
        for col_idx, fn in enumerate(field_names, start=2):
            cell = ws.cell(row=row_num, column=col_idx, value=record.get(fn, ""))
            style_cell(cell, fill=row_fill, font=FONT_DEFAULT)

    ws.cell(row=1, column=1).value = "#"
    autofit(ws)
    ws.column_dimensions["A"].width = 6


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------

def build_workbook(
    mapped_mms200: list[dict],
    mapped_itmwhs: list[dict],
    mapped_upditmfac: list[dict],
) -> Workbook:
    wb = Workbook()

    ctrl      = wb.active
    ctrl.title = "Control"
    hdr_fill  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")

    for c, label in enumerate(["Worksheet", "Description", "Data"], 1):
        cell = ctrl.cell(row=1, column=c, value=label)
        cell.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ctrl_entries = [
        ("API_MMS200MI_AddItmBasic", "Item master — Basic", "x"),
        ("API_MMS200MI_AddItmWhs",   "Item warehouse",      "x"),
        ("API_MMS200MI_UpdItmFac",   "Item facility",       "x"),
    ]
    for entry in ctrl_entries:
        ctrl.append(list(entry))
    autofit(ctrl)

    add_api_sheet(
        wb,
        "API_MMS200MI_AddItmBasic",
        FIELDS_MMS200,
        mapped_mms200,
        mapped_mms200[0] if mapped_mms200 else {},
    )
    add_api_sheet(
        wb,
        "API_MMS200MI_AddItmWhs",
        FIELDS_MMS200_WHS,
        mapped_itmwhs,
        mapped_itmwhs[0] if mapped_itmwhs else {},
    )
    add_api_sheet(
        wb,
        "API_MMS200MI_UpdItmFac",
        FIELDS_MMS200_FAC,
        mapped_upditmfac,
        mapped_upditmfac[0] if mapped_upditmfac else {},
    )

    return wb


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def process(input_folder: str, output_folder: str, db_path: str) -> None:
    csv_files = sorted(glob.glob(os.path.join(input_folder, "*.csv")))
    if not csv_files:
        log.error(f"No CSV files found in: {input_folder}")
        sys.exit(1)

    log.info(f"Found {len(csv_files)} CSV file(s) → using: {os.path.basename(csv_files[0])}")

    try:
        df = pd.read_csv(csv_files[0], dtype=str, keep_default_na=False, encoding="utf-8")
    except UnicodeDecodeError:
        df = pd.read_csv(csv_files[0], dtype=str, keep_default_na=False, encoding="latin-1")

    df = df.apply(lambda col: col.str.strip() if col.dtype == object else col)
    log.info(f"  Read {len(df):,} rows, {len(df.columns)} columns")

    mapped_mms200 = [map_row_mms200(row) for _, row in df.iterrows()]
    log.info(f"  Mapped {len(mapped_mms200):,} records → MMS200MI.AddItmBasic")

    mapped_itmwhs = [map_row_itmwhs(row) for _, row in df.iterrows()]
    log.info(f"  Mapped {len(mapped_itmwhs):,} records → MMS200MI.AddItmWhs")

    mapped_upditmfac = [map_row_upditmfac(row) for _, row in df.iterrows()]
    log.info(f"  Mapped {len(mapped_upditmfac):,} records → MMS200MI.UpdItmFac")

    wb = build_workbook(mapped_mms200, mapped_itmwhs, mapped_upditmfac)

    os.makedirs(output_folder, exist_ok=True)
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_folder, f"HJ_SKU_{timestamp}.xlsx")
    wb.save(output_path)
    log.info(f"  Saved: {output_path}")

    count = upsert_to_db(df, db_path, csv_files[0])
    log.info(f"  Upserted {count:,} rows → SyndigoSKU ({db_path})")

    write_sql_files()
    log.info(f"  SQL files written to: {SCRIPT_DIR}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate MMS200MI.AddItmBasic Excel template from Syndigo SKU CSV"
    )
    parser.add_argument(
        "--input-folder",
        default=DEFAULT_INPUT_FOLDER,
        help=f"Folder containing *.csv files (default: {DEFAULT_INPUT_FOLDER})",
    )
    parser.add_argument(
        "--output-folder",
        default=DEFAULT_OUTPUT_FOLDER,
        help=f"Folder for output .xlsx files (default: {DEFAULT_OUTPUT_FOLDER})",
    )
    parser.add_argument(
        "--db",
        default=DEFAULT_DB_PATH,
        help=f"Path to doppio.db (default: {DEFAULT_DB_PATH})",
    )
    args = parser.parse_args()

    log.info("=" * 60)
    log.info("HJ_SKU  —  MMS200MI.AddItmBasic / AddItmWhs / UpdItmFac")
    log.info("=" * 60)

    if not os.path.isdir(args.input_folder):
        log.error(f"Input folder not found: {args.input_folder}")
        sys.exit(1)

    process(args.input_folder, args.output_folder, args.db)
    log.info("Done.")


if __name__ == "__main__":
    main()
