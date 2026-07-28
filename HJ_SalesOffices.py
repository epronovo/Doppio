# HJ_SalesOffices.py
# -----------------------------------------------------------------------
# PURPOSE
#   Reads a Syndigo Sales Office CSV export from the input folder,
#   applies field mappings for each M3 API in the Sales Office sequence,
#   and writes a styled multi-tab Excel workbook to the output folder.
#
#   Tabs produced
#     API_CRS610MI_Add      — Customer master (Sales Office)
#     API_CRS620MI_Add      — Supplier master
#
#   The workbook follows the MpGd_Build_Template layout:
#     Row 1 – API field names          (light blue)
#     Row 2 – Field description        (light blue)
#     Row 3 – Required flag            (light blue)
#     Row 4 – Mapping / source note    (gray, bold)
#     Row 5 – Example (first record)   (italic)
#     Row 6+ – Data rows from CSV
#
# INPUTS
#   ./input/*.csv   — Syndigo Sales Office export
#
# OUTPUTS
#   ./output/HJ_SalesOffices_{YYYYMMDD}_{HHMMSS}.xlsx
#
# USAGE
#   python HJ_SalesOffices.py [--input-folder PATH] [--output-folder PATH] [--db PATH]
# -----------------------------------------------------------------------

import argparse
import glob
import json
import logging
import os
import sqlite3
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
SCRIPT_DIR            = os.path.dirname(os.path.abspath(__file__))
DEFAULT_INPUT_FOLDER  = os.path.join(SCRIPT_DIR, "input")
DEFAULT_OUTPUT_FOLDER = os.path.join(SCRIPT_DIR, "output")
DEFAULT_DB_PATH       = "/Users/ericpronovost/sqlite/doppio.db"

SUNO_START = 200001   # first supplier number; increments by 1 per row

# SQL template written to SCRIPT_DIR for each API tab
_SQL_FILES = {
    "API_CRS610MI_Add": (
        "SELECT 'CRS610MI'[minm],'Add'[trnm],* FROM (\n"
        "\tSELECT 'Z00003'[CUTM],\n"
        "\tSUBSTR(ptysalesofficepublicname,1,36)[CUNM],\n"
        "\tSUBSTR(street,1,36)[CUA1],\n"
        "\t'GB'[LNCD],\n"
        "\tptysalesofficeid[CUNO],\n"
        "\tCASE WHEN LENGTH(street) > 36 THEN SUBSTR(street,37) ELSE '' END[CUA2],\n"
        "\tpostalCode[PONO],\n"
        "\tcountryCode[CSCD],\n"
        "\tstateCode[ECAR],\n"
        "\t'10'[STAT],\n"
        "\tSUBSTR(city,1,20)[TOWN]\n"
        "\tFROM SyndigoSalesOffices\n"
        ") as temp\n"
        "WHERE 1=1\n"
        "AND CUNM <> '' AND CUA1 <> ''\n"
        "AND NOT (LENGTH(CUNM) > 36 OR LENGTH(CUA1) > 36 OR LENGTH(TOWN) > 20)\n"
    ),
    "API_CRS620MI_AddSupplier": (
        "SELECT 'CRS620MI'[minm],'AddSupplier'[trnm],* FROM (\n"
        "\tSELECT CAST((ROW_NUMBER() OVER (ORDER BY ptysalesofficeid) + 200000) AS TEXT)[SUNO],\n"
        "\tptysalesofficeid[SCNO],\n"
        "\tSUBSTR(ptysalesofficepublicname,1,36)[SUNM],\n"
        "\t'0'[SUTY],\n"
        "\tcountryCode[CSCD],\n"
        "\t'MDY'[DTFM],\n"
        "\t'D01'[ORTY],\n"
        "\t'1'[DT4T],\n"
        "\t'1'[DTCD],\n"
        "\tcountryCode||'D'[CUCD],\n"
        "\t'1'[CRTP],\n"
        "\t'1'[ATPR],\n"
        "\t'ISP'[SUCL],\n"
        "\t'FOB'[TEDL],\n"
        "\t'001'[MODL],\n"
        "\t'FOB'[TEAF],\n"
        "\t'GB'[LNCD],\n"
        "\t'001'[TEPA],\n"
        "\t'N30'[TEPY],\n"
        "\t'EFT'[PYME]\n"
        "\tFROM SyndigoSalesOffices\n"
        ") as temp\n"
        "WHERE 1=1\n"
        "AND SUNM <> ''\n"
    ),
    "API_CRS620MI_AddAddress": (
        "SELECT 'CRS620MI'[minm],'AddAddress'[trnm],* FROM (\n"
        "\tSELECT CAST((ROW_NUMBER() OVER (ORDER BY ptysalesofficeid) + 200000) AS TEXT)[SUNO],\n"
        "\t'1'[ADTE],\n"
        "\t'001'[ADID],\n"
        "\tSUBSTR(ptysalesofficepublicname,1,36)[SUNM],\n"
        "\tSUBSTR(street,1,36)[ADR1],\n"
        "\tCASE WHEN LENGTH(street) > 36 THEN SUBSTR(street,37) ELSE '' END[ADR2],\n"
        "\tSUBSTR(city,1,20)[TOWN],\n"
        "\tstateCode[ECAR],\n"
        "\tpostalCode[PONO],\n"
        "\tcountryCode[CSCD]\n"
        "\tFROM SyndigoSalesOffices\n"
        ") as temp\n"
        "WHERE 1=1\n"
        "AND SUNM <> '' AND ADR1 <> ''\n"
    ),
    "API_CRS335MI_AddCtrlObj": (
        "SELECT 'CRS335MI'[minm],'AddCtrlObj'[trnm],* FROM (\n"
        "\tSELECT CAST((ROW_NUMBER() OVER (ORDER BY ptysalesofficeid) + 200000) AS TEXT)[ACRF],\n"
        "\tSUBSTR(ptysalesofficepublicname,1,40)[TX40],\n"
        "\tSUBSTR(ptysalesofficepublicname,1,15)[TX15]\n"
        "\tFROM SyndigoSalesOffices\n"
        ") as temp\n"
        "WHERE 1=1\n"
        "AND TX40 <> ''\n"
    ),
    "API_MNS150MI_Add": (
        "SELECT 'MNS150MI'[minm],'Add'[trnm],* FROM (\n"
        "\tSELECT CAST((ROW_NUMBER() OVER (ORDER BY ptysalesofficeid) + 200000) AS TEXT)[USID],\n"
        "\tSUBSTR(ptysalesofficepublicname,1,36)[NAME],\n"
        "\t'300'[DFCO],\n"
        "\t'4'[ULTP]\n"
        "\tFROM SyndigoSalesOffices\n"
        ") as temp\n"
        "WHERE 1=1\n"
        "AND NAME <> ''\n"
    ),
    "API_CRS100MI_Add": (
        "SELECT 'CRS100MI'[minm],'Add'[trnm],* FROM (\n"
        "\tSELECT CAST((ROW_NUMBER() OVER (ORDER BY ptysalesofficeid) + 200000) AS TEXT)[SMCD],\n"
        "\tSUBSTR(ptysalesofficepublicname,1,40)[TX40],\n"
        "\tSUBSTR(ptysalesofficepublicname,1,15)[TX15]\n"
        "\tFROM SyndigoSalesOffices\n"
        ") as temp\n"
        "WHERE 1=1\n"
        "AND TX40 <> ''\n"
    ),
}

# (api_field, description, required, mapping_note)
FIELDS_610 = [
    ("CUTM", "Customer type",        "Yes", '"Z00003" — hardcoded'),
    ("CUNM", "Name",                 "Yes", "ptysalesofficepublicname (max 36)"),
    ("CUA1", "Address line 1",       "Yes", "BillingAddress: street (chars 1–36)"),
    ("LNCD", "Language code",        "Yes", '"GB" — hardcoded'),
    ("CUNO", "Customer number",      "Yes", "ptysalesofficeid"),
    ("CUA2", "Address line 2",       "Yes", "BillingAddress: street (chars 37+, if any)"),
    ("PONO", "Postal code",          "Yes", "BillingAddress: postalCode"),
    ("CSCD", "Country code",         "Yes", "BillingAddress: countryCode"),
    ("ECAR", "State / province",     "Yes", "BillingAddress: stateCode"),
    ("STAT", "Status",               "Yes", '"10" — hardcoded'),
    ("TOWN", "City",                 "Yes", "BillingAddress: city (max 20)"),
]

FIELDS_622 = [
    ("SUNO", "Supplier number",      "Yes", f"Sequence starting at {SUNO_START} (matches CRS620MI)"),
    ("ADTE", "Address type",         "Yes", '"1" — hardcoded'),
    ("ADID", "Address ID",           "Yes", '"001" — hardcoded'),
    ("SUNM", "Supplier name",        "Yes", "ptysalesofficepublicname (max 36)"),
    ("ADR1", "Address line 1",       "Yes", "BillingAddress: street (chars 1–36)"),
    ("ADR2", "Address line 2",       "Yes", "BillingAddress: street (chars 37+, if any)"),
    ("TOWN", "City",                 "Yes", "BillingAddress: city"),
    ("ECAR", "State / province",     "Yes", "BillingAddress: stateCode"),
    ("PONO", "Postal code",          "Yes", "BillingAddress: postalCode"),
    ("CSCD", "Country code",         "Yes", "BillingAddress: countryCode"),
]

FIELDS_335 = [
    ("ACRF",  "Attribute reference", "Yes", f"SUNO — sequence starting at {SUNO_START}"),
    ("TX40",  "Description (40)",    "Yes", "ptysalesofficepublicname (max 40)"),
    ("TX15",  "Description (15)",    "Yes", "ptysalesofficepublicname (max 15)"),
]

FIELDS_MNS150 = [
    ("USID",  "User ID",             "Yes", f"SUNO — sequence starting at {SUNO_START}"),
    ("NAME",  "Name",                "Yes", "ptysalesofficepublicname (max 36)"),
    ("DFCO",  "Default company",     "Yes", '"300" — hardcoded'),
    ("ULTP",  "User type",           "Yes", '"4" — hardcoded'),
]

FIELDS_100 = [
    ("SMCD",  "Salesperson code",    "Yes", f"SUNO — sequence starting at {SUNO_START}"),
    ("TX40",  "Description (40)",    "Yes", "ptysalesofficepublicname (max 40)"),
    ("TX15",  "Description (15)",    "Yes", "ptysalesofficepublicname (max 15)"),
]

FIELDS_620 = [
    ("SUNO", "Supplier number",      "Yes", f"Sequence starting at {SUNO_START}"),
    ("SCNO", "Customer number (ref)","Yes", "ptysalesofficeid"),
    ("SUNM", "Supplier name",        "Yes", "ptysalesofficepublicname"),
    ("SUTY", "Supplier type",        "Yes", '"0" — hardcoded'),
    ("CSCD", "Country code",         "Yes", "BillingAddress: countryCode"),
    ("DTFM", "Date format",          "Yes", '"MDY" — hardcoded'),
    ("ORTY", "Order type",           "Yes", '"D01" — hardcoded'),
    ("DT4T", "4-digit year flag",    "Yes", '"1" — hardcoded'),
    ("DTCD", "Date separator code",  "Yes", '"1" — hardcoded'),
    ("CUCD", "Currency code",        "Yes", 'BillingAddress: countryCode + "D" (e.g. USD)'),
    ("CRTP", "Exchange rate type",   "Yes", '"1" — hardcoded'),
    ("ATPR", "Authorised to pay",    "Yes", '"1" — hardcoded'),
    ("SUCL", "Supplier class",       "Yes", '"ISP" — hardcoded'),
    ("TEDL", "Delivery terms",       "Yes", '"FOB" — hardcoded'),
    ("MODL", "Delivery method",      "Yes", '"001" — hardcoded'),
    ("TEAF", "Freight terms",        "Yes", '"FOB" — hardcoded'),
    ("LNCD", "Language code",        "Yes", '"GB" — hardcoded'),
    ("TEPA", "Packaging terms",      "Yes", '"001" — hardcoded'),
    ("TEPY", "Payment terms",        "Yes", '"N30" — hardcoded'),
    ("PYME", "Payment method",       "Yes", '"EFT" — hardcoded'),
]

# Colours (match MpGd_Build_Template.py)
FILL_BLUE  = PatternFill("solid", start_color="ADD8E6", end_color="ADD8E6")
FILL_GRAY  = PatternFill("solid", start_color="D3D3D3", end_color="D3D3D3")
FILL_WHITE = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
FILL_ALT   = PatternFill("solid", start_color="F5F5F5", end_color="F5F5F5")

FONT_DEFAULT = Font(name="Arial", size=10)
FONT_BOLD    = Font(name="Arial", size=10, bold=True)
FONT_ITALIC  = Font(name="Arial", size=10, italic=True)
FONT_LABEL   = Font(name="Arial", size=10, bold=True, color="444444")

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(
            os.path.join(SCRIPT_DIR, "HJ_SalesOffices.log"),
            encoding="utf-8",
        ),
    ],
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_billing_address(raw: str) -> dict:
    """Parse the composite BillingAddress JSON blob; return {} on failure."""
    try:
        return json.loads(raw) if raw else {}
    except (json.JSONDecodeError, TypeError):
        return {}


def map_row_610(csv_row: pd.Series) -> dict:
    """Map one CSV row to CRS610MI.Add fields."""
    addr   = parse_billing_address(csv_row.get("BillingAddress", ""))
    street = (addr.get("street") or "").strip()

    return {
        "CUTM": "Z00003",
        "CUNM": (csv_row.get("ptysalesofficepublicname") or "").strip()[:36],
        "CUA1": street[:36],
        "LNCD": "GB",
        "CUNO": (csv_row.get("ptysalesofficeid") or "").strip(),
        "CUA2": street[36:] if len(street) > 36 else "",
        "PONO": (addr.get("postalCode") or "").strip(),
        "CSCD": (addr.get("countryCode") or "").strip(),
        "ECAR": (addr.get("stateCode") or "").strip(),
        "STAT": "10",
        "TOWN": (addr.get("city") or "").strip()[:20],
    }


def map_row_622(csv_row: pd.Series, row_index: int) -> dict:
    """Map one CSV row to CRS622MI.AddAddress fields."""
    addr   = parse_billing_address(csv_row.get("BillingAddress", ""))
    street = (addr.get("street") or "").strip()

    return {
        "SUNO": str(SUNO_START + row_index),
        "ADTE": "1",
        "ADID": "001",
        "SUNM": (csv_row.get("ptysalesofficepublicname") or "").strip()[:36],
        "ADR1": street[:36],
        "ADR2": street[36:] if len(street) > 36 else "",
        "TOWN": (addr.get("city") or "").strip()[:20],
        "ECAR": (addr.get("stateCode") or "").strip(),
        "PONO": (addr.get("postalCode") or "").strip(),
        "CSCD": (addr.get("countryCode") or "").strip(),
    }


def map_row_620(csv_row: pd.Series, row_index: int) -> dict:
    """Map one CSV row to CRS620MI.Add fields."""
    addr        = parse_billing_address(csv_row.get("BillingAddress", ""))
    country_code = (addr.get("countryCode") or "").strip()

    return {
        "SUNO": str(SUNO_START + row_index),
        "SCNO": (csv_row.get("ptysalesofficeid") or "").strip(),
        "SUNM": (csv_row.get("ptysalesofficepublicname") or "").strip()[:36],
        "SUTY": "0",
        "CSCD": country_code,
        "DTFM": "MDY",
        "ORTY": "D01",
        "DT4T": "1",
        "DTCD": "1",
        "CUCD": country_code + "D",
        "CRTP": "1",
        "ATPR": "1",
        "SUCL": "ISP",
        "TEDL": "FOB",
        "MODL": "001",
        "TEAF": "FOB",
        "LNCD": "GB",
        "TEPA": "001",
        "TEPY": "N30",
        "PYME": "EFT",
    }


def map_row_335(csv_row: pd.Series, row_index: int) -> dict:
    """Map one CSV row to CRS335MI.AddCtrlObj fields."""
    name = (csv_row.get("ptysalesofficepublicname") or "").strip()
    return {
        "ACRF": str(SUNO_START + row_index),
        "TX40": name[:40],
        "TX15": name[:15],
    }


def map_row_mns150(csv_row: pd.Series, row_index: int) -> dict:
    """Map one CSV row to MNS150MI.Add fields."""
    return {
        "USID": str(SUNO_START + row_index),
        "NAME": (csv_row.get("ptysalesofficepublicname") or "").strip()[:36],
        "DFCO": "300",
        "ULTP": "4",
    }


def map_row_100(csv_row: pd.Series, row_index: int) -> dict:
    """Map one CSV row to CRS100MI.Add fields."""
    name = (csv_row.get("ptysalesofficepublicname") or "").strip()
    return {
        "SMCD": str(SUNO_START + row_index),
        "TX40": name[:40],
        "TX15": name[:15],
    }


def upsert_to_db(df: pd.DataFrame, db_path: str, source_file: str) -> int:
    """Parse BillingAddress and upsert all rows into SyndigoSalesOffices."""
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    loaded_at = datetime.now().isoformat()
    rows = []
    for _, row in df.iterrows():
        addr = parse_billing_address(row.get("BillingAddress", ""))
        rows.append((
            (row.get("ptysalesofficeid")          or "").strip(),
            (row.get("ptysalesofficepublicname")  or "").strip(),
            (addr.get("street")                   or "").strip(),
            (addr.get("postalCode")               or "").strip(),
            (addr.get("countryCode")              or "").strip(),
            (addr.get("stateCode")                or "").strip(),
            (addr.get("city")                     or "").strip(),
            os.path.basename(source_file),
            loaded_at,
        ))
    conn = sqlite3.connect(db_path)
    conn.executemany(
        """INSERT OR REPLACE INTO SyndigoSalesOffices
           (ptysalesofficeid, ptysalesofficepublicname, street, postalCode,
            countryCode, stateCode, city, _source_file, _loaded_at)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        rows,
    )
    conn.commit()
    conn.close()
    return len(rows)


def write_sql_files() -> None:
    """Write one SalesOffices_{tab}.sql per API tab to the project root."""
    for tab_name, sql in _SQL_FILES.items():
        path = os.path.join(SCRIPT_DIR, f"SalesOffices_{tab_name}.sql")
        with open(path, "w", encoding="utf-8") as fh:
            fh.write(sql)
        log.info(f"  Wrote: {os.path.basename(path)}")


def autofit(ws):
    """Set column widths based on content; wrap all cells."""
    for col_cells in ws.columns:
        max_len    = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")


def style_cell(cell, fill=None, font=None, h_align="left"):
    cell.fill      = fill or FILL_WHITE
    cell.font      = font or FONT_DEFAULT
    cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Sheet builder (shared by all API tabs)
# ---------------------------------------------------------------------------

def add_api_sheet(wb: Workbook, ws_name: str, fields: list, data_rows: list[dict], example_row: dict):
    """Add one styled API tab to wb."""
    ws = wb.create_sheet(title=ws_name)

    field_names  = [f[0] for f in fields]
    descriptions = [f[1] for f in fields]
    required     = [f[2] for f in fields]
    mappings     = [f[3] for f in fields]
    examples     = [example_row.get(fn, "") for fn in field_names]

    header_rows = [
        ("FIELD",    field_names,  FILL_BLUE,  FONT_BOLD,    "center"),
        ("Message",  descriptions, FILL_BLUE,  FONT_DEFAULT, "left"),
        ("Required", required,     FILL_BLUE,  FONT_DEFAULT, "center"),
        ("Mapping",  mappings,     FILL_GRAY,  FONT_BOLD,    "left"),
        ("Example",  examples,     FILL_WHITE, FONT_ITALIC,  "left"),
    ]

    for row_idx, (label, values, fill, font, align) in enumerate(header_rows, start=1):
        cell_a = ws.cell(row=row_idx, column=1, value=label)
        style_cell(cell_a, fill=fill, font=FONT_LABEL, h_align="right")
        for col_idx, val in enumerate(values, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            style_cell(cell, fill=fill, font=font, h_align=align)

    # Freeze rows 1–3, keep col A visible
    ws.freeze_panes = ws.cell(row=4, column=2)

    # Data rows
    for i, record in enumerate(data_rows):
        row_num  = i + len(header_rows) + 1
        row_fill = FILL_ALT if i % 2 else FILL_WHITE
        ws.cell(row=row_num, column=1, value=i + 1).fill = row_fill
        for col_idx, fn in enumerate(field_names, start=2):
            cell = ws.cell(row=row_num, column=col_idx, value=record.get(fn, ""))
            style_cell(cell, fill=row_fill, font=FONT_DEFAULT)

    ws.cell(row=1, column=1).value = "#"
    autofit(ws)
    ws.column_dimensions["A"].width = 6


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------

def build_workbook(
    mapped_610: list[dict],
    mapped_620: list[dict],
    mapped_622: list[dict],
    mapped_335: list[dict],
    mapped_mns150: list[dict],
    mapped_100: list[dict],
) -> Workbook:
    wb = Workbook()

    # ── Control sheet ──────────────────────────────────────────────────────
    ctrl     = wb.active
    ctrl.title = "Control"
    hdr_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")

    for c, label in enumerate(["Worksheet", "Description", "Data"], 1):
        cell = ctrl.cell(row=1, column=c, value=label)
        cell.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ctrl_entries = [
        ("API_CRS610MI_Add",         "Customer interface", "x"),
        ("API_CRS620MI_AddSupplier", "Supplier Interface", "x"),
        ("API_CRS620MI_AddAddress",  "Supplier Interface", "x"),
        ("API_CRS335MI_AddCtrlObj",  "Control object",     "x"),
        ("API_MNS150MI_Add",         "User",               "x"),
        ("API_CRS100MI_Add",         "Salesperson",        "x"),
    ]
    for entry in ctrl_entries:
        ctrl.append(list(entry))
    autofit(ctrl)

    # ── API tabs ───────────────────────────────────────────────────────────
    add_api_sheet(wb, "API_CRS610MI_Add",         FIELDS_610,    mapped_610,    mapped_610[0]    if mapped_610    else {})
    add_api_sheet(wb, "API_CRS620MI_AddSupplier", FIELDS_620,    mapped_620,    mapped_620[0]    if mapped_620    else {})
    add_api_sheet(wb, "API_CRS620MI_AddAddress",  FIELDS_622,    mapped_622,    mapped_622[0]    if mapped_622    else {})
    add_api_sheet(wb, "API_CRS335MI_AddCtrlObj",  FIELDS_335,    mapped_335,    mapped_335[0]    if mapped_335    else {})
    add_api_sheet(wb, "API_MNS150MI_Add", FIELDS_MNS150, mapped_mns150, mapped_mns150[0] if mapped_mns150 else {})
    add_api_sheet(wb, "API_CRS100MI_Add", FIELDS_100,   mapped_100,   mapped_100[0]   if mapped_100   else {})

    return wb


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def process(input_folder: str, output_folder: str, db_path: str) -> None:
    csv_files = sorted(glob.glob(os.path.join(input_folder, "*.csv")))
    if not csv_files:
        log.error(f"No CSV files found in: {input_folder}")
        sys.exit(1)

    log.info(f"Found {len(csv_files)} CSV file(s) → using: {os.path.basename(csv_files[0])}")

    try:
        df = pd.read_csv(csv_files[0], dtype=str, keep_default_na=False, encoding="utf-8")
    except UnicodeDecodeError:
        df = pd.read_csv(csv_files[0], dtype=str, keep_default_na=False, encoding="latin-1")

    df = df.apply(lambda col: col.str.strip() if col.dtype == object else col)
    log.info(f"  Read {len(df):,} rows, {len(df.columns)} columns")

    SyndigoSalesOffices = df.to_dict(orient="records")
    log.info(f"  Loaded {len(SyndigoSalesOffices):,} raw records into SyndigoSalesOffices")

    mapped_610 = [map_row_610(row) for _, row in df.iterrows()]
    log.info(f"  Mapped {len(mapped_610):,} records → CRS610MI.Add")

    mapped_620 = [map_row_620(row, i) for i, (_, row) in enumerate(df.iterrows())]
    log.info(f"  Mapped {len(mapped_620):,} records → CRS620MI.Add  (SUNO {SUNO_START}–{SUNO_START + len(mapped_620) - 1})")

    mapped_622   = [map_row_622(row, i)    for i, (_, row) in enumerate(df.iterrows())]
    log.info(f"  Mapped {len(mapped_622):,} records → CRS622MI.AddAddress")

    mapped_335   = [map_row_335(row, i)    for i, (_, row) in enumerate(df.iterrows())]
    log.info(f"  Mapped {len(mapped_335):,} records → CRS335MI.AddCtrlObj")

    mapped_mns150 = [map_row_mns150(row, i) for i, (_, row) in enumerate(df.iterrows())]
    log.info(f"  Mapped {len(mapped_mns150):,} records → MNS150MI.Add")

    mapped_100   = [map_row_100(row, i)    for i, (_, row) in enumerate(df.iterrows())]
    log.info(f"  Mapped {len(mapped_100):,} records → CRS100MI.Add")

    wb = build_workbook(mapped_610, mapped_620, mapped_622, mapped_335, mapped_mns150, mapped_100)

    os.makedirs(output_folder, exist_ok=True)
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_folder, f"HJ_SalesOffices_{timestamp}.xlsx")
    wb.save(output_path)
    log.info(f"  Saved: {output_path}")

    count = upsert_to_db(df, db_path, csv_files[0])
    log.info(f"  Upserted {count:,} rows → SyndigoSalesOffices ({db_path})")

    write_sql_files()
    log.info(f"  SQL files written to: {SCRIPT_DIR}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate CRS610MI/CRS620MI Excel templates from Syndigo Sales Office CSV"
    )
    parser.add_argument(
        "--input-folder",
        default=DEFAULT_INPUT_FOLDER,
        help=f"Folder containing *.csv files (default: {DEFAULT_INPUT_FOLDER})",
    )
    parser.add_argument(
        "--output-folder",
        default=DEFAULT_OUTPUT_FOLDER,
        help=f"Folder for output .xlsx files (default: {DEFAULT_OUTPUT_FOLDER})",
    )
    parser.add_argument(
        "--db",
        default=DEFAULT_DB_PATH,
        help=f"Path to doppio.db (default: {DEFAULT_DB_PATH})",
    )
    args = parser.parse_args()

    log.info("=" * 60)
    log.info("HJ_SalesOffices  —  CRS610MI.Add + CRS620MI.Add")
    log.info("=" * 60)

    if not os.path.isdir(args.input_folder):
        log.error(f"Input folder not found: {args.input_folder}")
        sys.exit(1)

    process(args.input_folder, args.output_folder, args.db)
    log.info("Done.")


if __name__ == "__main__":
    main()
