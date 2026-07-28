# HJ_Syndigo_Organization.py
# -----------------------------------------------------------------------
# PURPOSE
#   Reads Syndigo Organization CSV exports from the input folder,
#   upserts all records into the SyndigoOrg table in doppio.db,
#   applies field mappings for each M3 API in the CRS610MI sequence,
#   and writes a styled multi-tab Excel workbook to the output folder.
#
#   Tabs produced
#     API_CRS610MI_Add          — Customer master (Add)
#     API_CRS610MI_ChgFinancial — Customer master (ChgFinancial)
#     API_CRS610MI_ChgOrderInfo — Customer master (ChgOrderInfo)
#
#   The workbook follows the MpGd_Build_Template layout:
#     Row 1 – API field names          (light blue)
#     Row 2 – Field description        (light blue)
#     Row 3 – Required flag            (light blue)
#     Row 4 – Mapping / source note    (gray, bold)
#     Row 5 – Example (first record)   (italic)
#     Row 6+ – Data rows from CSV
#
# INPUTS
#   ./input/*.csv   — Syndigo Organization export
#
# OUTPUTS
#   ./output/HJ_Customers_{YYYYMMDD}_{HHMMSS}.xlsx
#   ./Customers_{tab}.sql  — one per API tab
#
# USAGE
#   python HJ_Syndigo_Organization.py [--input-folder PATH] [--output-folder PATH] [--db PATH]
# -----------------------------------------------------------------------

import argparse
import glob
import logging
import os
import re
import sqlite3
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
SCRIPT_DIR            = os.path.dirname(os.path.abspath(__file__))
DEFAULT_INPUT_FOLDER  = os.path.join(SCRIPT_DIR, "input")
DEFAULT_OUTPUT_FOLDER = os.path.join(SCRIPT_DIR, "output")
DEFAULT_DB_PATH       = "/Users/ericpronovost/sqlite/doppio.db"
TABLE_NAME            = "SyndigoOrg"

# SQL templates written to SCRIPT_DIR for each API tab
_SQL_FILES = {
    "API_CRS610MI_Add": (
        "SELECT 'CRS610MI'[minm],'Add'[trnm],* FROM (\n"
        "\tSELECT 'Z00001'[CUTM],\n"
        "\t'300'[CONO],\n"
        "\t''[DIVI],\n"
        "\t'GB'[LNCD],\n"
        "\tptyorgid[CUNO],\n"
        "\tSUBSTR(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(UPPER(ptyorgpublicname), \"'\", ''), '*', ''), '\"', ''), ' & ', '&'), '& ', '&'), ' &', '&'), '&', ' & '),1,36) as [CUNM],\n"
        "\tcase when mdr_dept_name_full = '' and ptyorgmailtostreet1 = '' then UPPER(ptyorgmdrmstreet)\n"
        "\t     when mdr_dept_name_full <> '' then UPPER('dept:' || mdr_dept_name_full) \n"
        "\t     else ptyorgmailtostreet1 end as [CUA1],\n"
        "\tcase when mdr_dept_name_full = '' then ptyorgmailtostreet2 else ptyorgmdrmstreet end as [CUA2],\n"
        "\tcase when ptyorgmailtoaddresspostalcode = '' then ptyorgmdrmzipcode else ptyorgmailtoaddresspostalcode end as [PONO],\n"
        "\tPHONE[PHNO],\n"
        "\t''[PHN2],\n"
        "\tFAX[TFNO],\n"
        "\t'0'[CUTP],\n"
        "\t''[YREF],\n"
        "\t''[YRE2],\n"
        "\tCSCD[CSCD],\n"
        "\tcase when ptyorgmailtoaddressstate = '' then ptyorgmdrmstate else ptyorgmailtoaddressstate end as [ECAR],\n"
        "\t'20'[STAT],\n"
        "\tcase when ptyorgmailtoaddresscity = '' then UPPER(ptyorgmdrmcity) else UPPER(ptyorgmailtoaddresscity) end as [TOWN],\n"
        "\tCSCD||ptyorgmailtoaddressstate[EDES]\n"
        "\tFROM SyndigoOrg\n"
        "\tWHERE mdr_pid = ''\n"
        ") as temp\n"
        "WHERE 1=1 \n"
        "AND CUNM <> '' AND CUA1 <> ''\n"
        "AND NOT (LENGTH(CUNM) > 36 OR LENGTH(CUA1) > 36 OR LENGTH(TOWN) > 20)"
    ),
    "API_CRS610MI_ChgFinancial": (
        "SELECT 'CRS610MI'[minm],'ChgFinancial'[trnm],* FROM (\n"
        "\tSELECT \n"
        "\tptyorgid[CUNO],\n"
        "\tCSCD||'D' AS [CUCD],\n"
        "\tcase when ptyorgmailtoaddresscity = '' then UPPER(ptyorgmdrmcity) else UPPER(ptyorgmailtoaddresscity) end as [TOWN],\n"
        "\tcase when mdr_dept_name_full = '' and ptyorgmailtostreet1 = '' then UPPER(ptyorgmdrmstreet)\n"
        "\t     when mdr_dept_name_full <> '' then UPPER('dept:' || mdr_dept_name_full) \n"
        "\t     else ptyorgmailtostreet1 end as [CUA1],\n"
        "\tSUBSTR(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(UPPER(ptyorgpublicname), \"'\", ''), '*', ''), '\"', ''), ' & ', '&'), '& ', '&'), ' &', '&'), '&', ' & '),1,36) as [CUNM]\n"
        "\tFROM SyndigoOrg\n"
        "\tWHERE mdr_pid = ''\n"
        ") as temp\n"
        "WHERE 1=1 \n"
        "AND CUNM <> '' AND CUA1 <> ''\n"
        "AND NOT (LENGTH(CUNM) > 36 OR LENGTH(CUA1) > 36 OR LENGTH(TOWN) > 20)\n"
    ),
    "API_CRS610MI_ChgOrderInfo": (
        "SELECT 'CRS610MI'[minm],'ChgOrderInfo'[trnm],\n"
        "* FROM (\n"
        "\tSELECT \n"
        "\tptyorgid[CUNO],\n"
        "\tCASE WHEN ptyorgtype = 'Scholastic' THEN 'SCH' \n"
        "\t\tWHEN ptyorgtype = 'Commercial' THEN 'COM' \n"
        "\t\tWHEN ptyorgtype = 'College' THEN 'COL' \n"
        "\t\tWHEN ptyorgtype = 'Greek' THEN 'GRK' \n"
        "\tELSE '' END AS [CUCL],\n"
        "\tcase when ptyorglegacyaccountrecievablenumber = '' THEN 'MDM:'||mdr_pid else ptyorglegacyaccountrecievablenumber end as [OREF],\n"
        "\t'?'[SMCD],\n"
        "\t''[PYNO],\n"
        "\t''[INRC],\n"
        "\t''[DOGR],\n"
        "\t'0'[ADBO],\n"
        "\t'0'[AICD],\n"
        "\t'0'[BOP1],\n"
        "\tcase when ptyorgmailtoaddresscity = '' then UPPER(ptyorgmdrmcity) else UPPER(ptyorgmailtoaddresscity) end as [TOWN],\n"
        "\tcase when mdr_dept_name_full = '' and ptyorgmailtostreet1 = '' then UPPER(ptyorgmdrmstreet)\n"
        "\t     when mdr_dept_name_full <> '' then UPPER('dept:' || mdr_dept_name_full) \n"
        "\t     else ptyorgmailtostreet1 end as [CUA1],\n"
        "\tSUBSTR(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(UPPER(ptyorgpublicname), \"'\", ''), '*', ''), '\"', ''), ' & ', '&'), '& ', '&'), ' &', '&'), '&', ' & '),1,36) as [CUNM]\n"
        "\tFROM SyndigoOrg\n"
        "\tWHERE mdr_pid = ''\n"
        ") as temp\n"
        "WHERE 1=1 \n"
        "AND CUNM <> '' AND CUA1 <> ''\n"
        "AND NOT (LENGTH(CUNM) > 36 OR LENGTH(CUA1) > 36 OR LENGTH(TOWN) > 20)\n"
    ),
}

# (api_field, description, required, mapping_note)
FIELDS_610_ADD = [
    ("CUTM", "Customer type",            "Yes", '"Z00001" — hardcoded'),
    ("CONO", "Company",                  "Yes", '"300" — hardcoded'),
    ("DIVI", "Division",                 "No",  '"" — hardcoded'),
    ("LNCD", "Language code",            "Yes", '"GB" — hardcoded'),
    ("CUNO", "Customer number",          "Yes", "ptyorgid"),
    ("CUNM", "Name",                     "Yes", "UPPER(ptyorgpublicname) — cleaned, max 36"),
    ("CUA1", "Address line 1",           "Yes", "UPPER('dept:'||mdr_dept_name_full) if dept; else ptyorgmailtostreet1; else UPPER(ptyorgmdrmstreet)"),
    ("CUA2", "Address line 2",           "No",  "ptyorgmailtostreet2 (or ptyorgmdrmstreet when dept present)"),
    ("PONO", "Postal code",              "No",  "ptyorgmailtoaddresspostalcode or ptyorgmdrmzipcode"),
    ("PHNO", "Phone",                    "No",  "PHONE"),
    ("PHN2", "Phone 2",                  "No",  '"" — hardcoded'),
    ("TFNO", "Fax",                      "No",  "FAX"),
    ("CUTP", "Customer type code",       "Yes", '"0" — hardcoded'),
    ("YREF", "Your reference",           "No",  '"" — hardcoded'),
    ("YRE2", "Your reference 2",         "No",  '"" — hardcoded'),
    ("CSCD", "Country code",             "Yes", "CSCD — CA or US derived from postal/state"),
    ("ECAR", "State / province",         "No",  "ptyorgmailtoaddressstate or ptyorgmdrmstate"),
    ("STAT", "Status",                   "Yes", '"20" — hardcoded'),
    ("TOWN", "City",                     "No",  "UPPER(ptyorgmailtoaddresscity or ptyorgmdrmcity) — max 20"),
    ("EDES", "District / delivery addr", "No",  "CSCD + ptyorgmailtoaddressstate"),
]

FIELDS_610_CHG_FINANCIAL = [
    ("CUNO", "Customer number", "Yes", "ptyorgid"),
    ("CUCD", "Currency code",   "Yes", 'CSCD + "D" (e.g. CAD); rows with USD excluded'),
    ("TOWN", "City",            "No",  "UPPER(ptyorgmailtoaddresscity or ptyorgmdrmcity)"),
    ("CUA1", "Address line 1",  "No",  "dept: prefix if mdr_dept_name_full; else ptyorgmailtostreet1 or UPPER(ptyorgmdrmstreet)"),
    ("CUNM", "Name",            "Yes", "UPPER(ptyorgpublicname) — cleaned, max 36"),
]

FIELDS_610_CHG_ORDER_INFO = [
    ("CUNO", "Customer number",   "Yes", "ptyorgid"),
    ("CUCL", "Customer class",    "No",  "ptyorgtype → Scholastic=SCH, Commercial=COM, College=COL, Greek=GRK"),
    ("OREF", "Order reference",   "No",  "ptyorglegacyaccountrecievablenumber or MDM:mdr_pid"),
    ("SMCD", "Salesperson code",  "No",  '"?" — hardcoded'),
    ("PYNO", "Payer",             "No",  '"" — hardcoded'),
    ("INRC", "Invoice recipient", "No",  '"" — hardcoded'),
    ("DOGR", "Document group",    "No",  '"" — hardcoded'),
    ("ADBO", "Add-on charge",     "No",  '"0" — hardcoded'),
    ("AICD", "Auto invoicing",    "No",  '"0" — hardcoded'),
    ("BOP1", "BOP option 1",      "No",  '"0" — hardcoded'),
    ("TOWN", "City",              "No",  "UPPER(ptyorgmailtoaddresscity or ptyorgmdrmcity)"),
    ("CUA1", "Address line 1",    "No",  "dept: prefix if mdr_dept_name_full; else ptyorgmailtostreet1 or UPPER(ptyorgmdrmstreet)"),
    ("CUNM", "Name",              "Yes", "UPPER(ptyorgpublicname) — cleaned, max 36"),
]

# Colours (match MpGd_Build_Template.py)
FILL_BLUE  = PatternFill("solid", start_color="ADD8E6", end_color="ADD8E6")
FILL_GRAY  = PatternFill("solid", start_color="D3D3D3", end_color="D3D3D3")
FILL_WHITE = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
FILL_ALT   = PatternFill("solid", start_color="F5F5F5", end_color="F5F5F5")

FONT_DEFAULT = Font(name="Arial", size=10)
FONT_BOLD    = Font(name="Arial", size=10, bold=True)
FONT_ITALIC  = Font(name="Arial", size=10, italic=True)
FONT_LABEL   = Font(name="Arial", size=10, bold=True, color="444444")

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(
            os.path.join(SCRIPT_DIR, "HJ_Syndigo_Organization.log"),
            encoding="utf-8",
        ),
    ],
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers — column sanitization
# ---------------------------------------------------------------------------

def sanitize_column_name(name: str) -> str:
    name = name.strip()
    name = re.sub(r"[^A-Za-z0-9_]", "_", name)
    name = re.sub(r"_+", "_", name)
    name = name.strip("_")
    return name


def make_column_map(raw_columns: list[str]) -> dict[str, str]:
    seen: dict[str, int] = {}
    mapping: dict[str, str] = {}
    for raw in raw_columns:
        clean = sanitize_column_name(raw)
        if clean in seen:
            seen[clean] += 1
            clean = f"{clean}_{seen[clean]}"
        else:
            seen[clean] = 1
        mapping[raw] = clean
    return mapping


# ---------------------------------------------------------------------------
# Canadian postal code → province resolver
# ---------------------------------------------------------------------------

_CA_PREFIX_MAP: dict[str, str] = {
    "A": "NL", "B": "NS", "C": "PE", "E": "NB",
    "G": "QC", "H": "QC", "J": "QC",
    "K": "ON", "L": "ON", "M": "ON", "N": "ON", "P": "ON",
    "R": "MB", "S": "SK", "T": "AB", "V": "BC", "Y": "YT",
}

_CA_X_PREFIX_MAP: dict[str, str] = {
    "X0A": "NU", "X0B": "NU", "X0C": "NU",
    "X0E": "NT", "X0G": "NT", "X1A": "NT",
}

_CA_PROVINCES: frozenset[str] = frozenset({
    "AB", "BC", "MB", "NB", "NL", "NS", "NT", "NU", "ON", "PE", "QC", "SK", "YT"
})

_CA_POSTAL_RE = re.compile(r'^[A-Za-z]\d[A-Za-z]')
_US_POSTAL_RE = re.compile(r'^\d{5}')


def _ca_province_from_postal(postal: str) -> str:
    raw = postal.strip().upper().replace(" ", "").replace("-", "")
    if len(raw) < 3:
        return ""
    first = raw[0]
    if first == "X":
        return _CA_X_PREFIX_MAP.get(raw[:3], "NT")
    return _CA_PREFIX_MAP.get(first, "")


def fix_canadian_provinces(df: pd.DataFrame) -> pd.DataFrame:
    state_col  = "ptyorgmailtoaddressstate"
    postal_col = "ptyorgmailtoaddresspostalcode"

    if state_col not in df.columns or postal_col not in df.columns:
        return df

    mask = df[state_col].str.upper() == "CN"
    count = mask.sum()
    if count == 0:
        return df

    log.info(f"  Resolving {count:,} Canadian province code(s) from postal codes …")
    df = df.copy()
    resolved = 0

    for idx in df.index[mask]:
        postal   = df.at[idx, postal_col]
        province = _ca_province_from_postal(postal)
        if province:
            df.at[idx, state_col] = province
            resolved += 1
        else:
            log.warning(
                f"    ptyorgid={df.at[idx, 'ptyorgid']!r}: "
                f"could not resolve province from postal {postal!r} — left as 'CN'"
            )

    log.info(f"    → {resolved:,} resolved, {count - resolved:,} unresolved (kept as CN)")
    return df


def _detect_address_country(state: str, postal: str) -> str:
    cleaned_postal = postal.strip().upper().replace(" ", "").replace("-", "")
    if cleaned_postal:
        if _CA_POSTAL_RE.match(cleaned_postal):
            return "CA"
        if _US_POSTAL_RE.match(cleaned_postal):
            return "US"
    if state.strip().upper() in _CA_PROVINCES:
        return "CA"
    return "US"


def tag_cscd_country(df: pd.DataFrame) -> pd.DataFrame:
    state_col  = "ptyorgmailtoaddressstate"
    postal_col = "ptyorgmailtoaddresspostalcode"

    if state_col not in df.columns and postal_col not in df.columns:
        log.warning("  Neither state nor postal column found — skipping CSCD tag")
        return df

    df = df.copy()
    ca_count = us_count = 0
    countries = []

    for idx in df.index:
        state  = df.at[idx, state_col]  if state_col  in df.columns else ""
        postal = df.at[idx, postal_col] if postal_col in df.columns else ""
        country = _detect_address_country(state, postal)
        countries.append(country)
        if country == "CA":
            ca_count += 1
        else:
            us_count += 1

    df["CSCD"] = countries
    log.info(f"  Tagged CSCD → CA: {ca_count:,}  US: {us_count:,}")
    return df


def read_csv(path: str) -> pd.DataFrame:
    log.info(f"Reading: {path}")
    try:
        df = pd.read_csv(path, dtype=str, keep_default_na=False, encoding="utf-8")
    except UnicodeDecodeError:
        df = pd.read_csv(path, dtype=str, keep_default_na=False, encoding="latin-1")

    col_map = make_column_map(df.columns.tolist())
    df.rename(columns=col_map, inplace=True)
    df = df.apply(lambda col: col.str.strip() if col.dtype == object else col)
    log.info(f"  → {len(df):,} rows, {len(df.columns)} columns")
    return df


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def ensure_table(conn: sqlite3.Connection, df: pd.DataFrame, table: str) -> None:
    cursor = conn.cursor()
    existing_cols: set[str] = set()
    cursor.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)
    )
    if cursor.fetchone():
        cursor.execute(f"PRAGMA table_info({table})")
        existing_cols = {row[1] for row in cursor.fetchall()}

    if not existing_cols:
        col_defs = []
        for col in df.columns:
            if col == "ptyorgid":
                col_defs.append(f'"{col}" TEXT PRIMARY KEY')
            else:
                col_defs.append(f'"{col}" TEXT')
        col_defs.append('"_source_file" TEXT')
        col_defs.append('"_loaded_at"   TEXT')
        ddl = f'CREATE TABLE IF NOT EXISTS "{table}" (\n  ' + ",\n  ".join(col_defs) + "\n)"
        log.info(f"Creating table {table}")
        cursor.execute(ddl)
        conn.commit()
    else:
        df_cols = set(df.columns) | {"_source_file", "_loaded_at"}
        new_cols = df_cols - existing_cols
        for col in sorted(new_cols):
            log.info(f"  Adding new column: {col}")
            cursor.execute(f'ALTER TABLE "{table}" ADD COLUMN "{col}" TEXT')
        if new_cols:
            conn.commit()


def upsert_dataframe(
    conn: sqlite3.Connection,
    df: pd.DataFrame,
    table: str,
    source_file: str,
) -> tuple[int, int]:
    df = df.copy()
    df["_source_file"] = os.path.basename(source_file)
    df["_loaded_at"]   = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")

    cursor = conn.cursor()
    cursor.execute(f'SELECT COUNT(*) FROM "{table}"')
    before = cursor.fetchone()[0]

    cols         = [f'"{c}"' for c in df.columns]
    placeholders = ", ".join(["?"] * len(df.columns))
    sql = f'INSERT OR REPLACE INTO "{table}" ({", ".join(cols)}) VALUES ({placeholders})'

    rows = [tuple(row) for row in df.itertuples(index=False, name=None)]
    cursor.executemany(sql, rows)
    conn.commit()

    cursor.execute(f'SELECT COUNT(*) FROM "{table}"')
    after = cursor.fetchone()[0]

    net_new      = after - before
    replacements = len(rows) - net_new
    return net_new, replacements


# ---------------------------------------------------------------------------
# Row-level field helpers
# ---------------------------------------------------------------------------

def _col(row: pd.Series, name: str) -> str:
    v = row.get(name, "")
    return "" if v is None else str(v).strip()


def _clean_name(raw: str) -> str:
    """Match SQL: UPPER → strip ' * " → normalise & → SUBSTR 36."""
    s = raw.upper()
    for ch in ("'", "*", '"'):
        s = s.replace(ch, "")
    s = s.replace(" & ", "&").replace("& ", "&").replace(" &", "&").replace("&", " & ")
    return s[:36]


def _cua1(row: pd.Series) -> str:
    dept    = _col(row, "mdr_dept_name_full")
    street1 = _col(row, "ptyorgmailtostreet1")
    mdrm    = _col(row, "ptyorgmdrmstreet")
    if dept:
        return ("dept:" + dept).upper()
    if street1:
        return street1
    return mdrm.upper()


def _cua2(row: pd.Series) -> str:
    dept    = _col(row, "mdr_dept_name_full")
    street2 = _col(row, "ptyorgmailtostreet2")
    mdrm    = _col(row, "ptyorgmdrmstreet")
    return mdrm if dept else street2


def _pono(row: pd.Series) -> str:
    pono = _col(row, "ptyorgmailtoaddresspostalcode")
    return pono if pono else _col(row, "ptyorgmdrmzipcode")


def _ecar(row: pd.Series) -> str:
    state = _col(row, "ptyorgmailtoaddressstate")
    return state if state else _col(row, "ptyorgmdrmstate")


def _town(row: pd.Series) -> str:
    city = _col(row, "ptyorgmailtoaddresscity")
    mdrm = _col(row, "ptyorgmdrmcity")
    return (city if city else mdrm).upper()


# ---------------------------------------------------------------------------
# Row mappers
# ---------------------------------------------------------------------------

def map_row_610_add(row: pd.Series) -> dict:
    cscd = _col(row, "CSCD")
    ecar = _ecar(row)
    return {
        "CUTM": "Z00001",
        "CONO": "300",
        "DIVI": "",
        "LNCD": "GB",
        "CUNO": _col(row, "ptyorgid"),
        "CUNM": _clean_name(_col(row, "ptyorgpublicname")),
        "CUA1": _cua1(row),
        "CUA2": _cua2(row),
        "PONO": _pono(row),
        "PHNO": _col(row, "PHONE"),
        "PHN2": "",
        "TFNO": _col(row, "FAX"),
        "CUTP": "0",
        "YREF": "",
        "YRE2": "",
        "CSCD": cscd,
        "ECAR": ecar,
        "STAT": "20",
        "TOWN": _town(row),
        "EDES": cscd + ecar,
    }


def map_row_610_chg_financial(row: pd.Series) -> dict:
    cscd = _col(row, "CSCD")
    return {
        "CUNO": _col(row, "ptyorgid"),
        "CUCD": cscd + "D",
        "TOWN": _town(row),
        "CUA1": _cua1(row),
        "CUNM": _clean_name(_col(row, "ptyorgpublicname")),
    }


def map_row_610_chg_order_info(row: pd.Series) -> dict:
    cucl_map = {"Scholastic": "SCH", "Commercial": "COM", "College": "COL", "Greek": "GRK"}
    legacy   = _col(row, "ptyorglegacyaccountrecievablenumber")
    mdr_pid  = _col(row, "mdr_pid")
    return {
        "CUNO": _col(row, "ptyorgid"),
        "CUCL": cucl_map.get(_col(row, "ptyorgtype"), ""),
        "OREF": legacy if legacy else "MDM:" + mdr_pid,
        "SMCD": "?",
        "PYNO": "",
        "INRC": "",
        "DOGR": "",
        "ADBO": "0",
        "AICD": "0",
        "BOP1": "0",
        "TOWN": _town(row),
        "CUA1": _cua1(row),
        "CUNM": _clean_name(_col(row, "ptyorgpublicname")),
    }


# ---------------------------------------------------------------------------
# SQL file writer
# ---------------------------------------------------------------------------

def write_sql_files() -> None:
    for tab_name, sql in _SQL_FILES.items():
        path = os.path.join(SCRIPT_DIR, f"Customers_{tab_name}.sql")
        with open(path, "w", encoding="utf-8") as fh:
            fh.write(sql)
        log.info(f"  Wrote: {os.path.basename(path)}")


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def autofit(ws) -> None:
    for col_cells in ws.columns:
        max_len    = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")


def style_cell(cell, fill=None, font=None, h_align="left") -> None:
    cell.fill      = fill or FILL_WHITE
    cell.font      = font or FONT_DEFAULT
    cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)


def add_api_sheet(
    wb: Workbook,
    ws_name: str,
    fields: list,
    data_rows: list[dict],
    example_row: dict,
) -> None:
    ws = wb.create_sheet(title=ws_name)

    field_names  = [f[0] for f in fields]
    descriptions = [f[1] for f in fields]
    required     = [f[2] for f in fields]
    mappings     = [f[3] for f in fields]
    examples     = [example_row.get(fn, "") for fn in field_names]

    header_rows = [
        ("FIELD",    field_names,  FILL_BLUE,  FONT_BOLD,    "center"),
        ("Message",  descriptions, FILL_BLUE,  FONT_DEFAULT, "left"),
        ("Required", required,     FILL_BLUE,  FONT_DEFAULT, "center"),
        ("Mapping",  mappings,     FILL_GRAY,  FONT_BOLD,    "left"),
        ("Example",  examples,     FILL_WHITE, FONT_ITALIC,  "left"),
    ]

    for row_idx, (label, values, fill, font, align) in enumerate(header_rows, start=1):
        cell_a = ws.cell(row=row_idx, column=1, value=label)
        style_cell(cell_a, fill=fill, font=FONT_LABEL, h_align="right")
        for col_idx, val in enumerate(values, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            style_cell(cell, fill=fill, font=font, h_align=align)

    ws.freeze_panes = ws.cell(row=4, column=2)

    for i, record in enumerate(data_rows):
        row_num  = i + len(header_rows) + 1
        row_fill = FILL_ALT if i % 2 else FILL_WHITE
        ws.cell(row=row_num, column=1, value=i + 1).fill = row_fill
        for col_idx, fn in enumerate(field_names, start=2):
            cell = ws.cell(row=row_num, column=col_idx, value=record.get(fn, ""))
            style_cell(cell, fill=row_fill, font=FONT_DEFAULT)

    ws.cell(row=1, column=1).value = "#"
    autofit(ws)
    ws.column_dimensions["A"].width = 6


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------

def build_workbook(
    mapped_add:     list[dict],
    mapped_chg_fin: list[dict],
    mapped_chg_ord: list[dict],
) -> Workbook:
    wb = Workbook()

    ctrl      = wb.active
    ctrl.title = "Control"
    hdr_fill  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")

    for c, label in enumerate(["Worksheet", "Description", "Data"], 1):
        cell = ctrl.cell(row=1, column=c, value=label)
        cell.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ctrl_entries = [
        ("API_CRS610MI_Add",          "Customer master (Add)",          "x"),
        ("API_CRS610MI_ChgFinancial",  "Customer master (ChgFinancial)", "x"),
        ("API_CRS610MI_ChgOrderInfo",  "Customer master (ChgOrderInfo)", "x"),
    ]
    for entry in ctrl_entries:
        ctrl.append(list(entry))
    autofit(ctrl)

    add_api_sheet(
        wb, "API_CRS610MI_Add",
        FIELDS_610_ADD, mapped_add,
        mapped_add[0] if mapped_add else {},
    )
    add_api_sheet(
        wb, "API_CRS610MI_ChgFinancial",
        FIELDS_610_CHG_FINANCIAL, mapped_chg_fin,
        mapped_chg_fin[0] if mapped_chg_fin else {},
    )
    add_api_sheet(
        wb, "API_CRS610MI_ChgOrderInfo",
        FIELDS_610_CHG_ORDER_INFO, mapped_chg_ord,
        mapped_chg_ord[0] if mapped_chg_ord else {},
    )

    return wb


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------

def process(input_folder: str, output_folder: str, db_path: str) -> None:
    csv_files = sorted(glob.glob(os.path.join(input_folder, "*.csv")))
    if not csv_files:
        log.error(f"No CSV files found in: {input_folder}")
        sys.exit(1)

    log.info(f"Found {len(csv_files)} CSV file(s)")
    log.info(f"Target database : {db_path}")
    log.info(f"Target table    : {TABLE_NAME}")

    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")

    total_new = total_replaced = 0
    all_dfs: list[pd.DataFrame] = []

    for path in csv_files:
        try:
            df = read_csv(path)
            if df.empty:
                log.warning(f"  Skipping empty file: {path}")
                continue
            df = fix_canadian_provinces(df)
            df = tag_cscd_country(df)
            ensure_table(conn, df, TABLE_NAME)
            new, replaced = upsert_dataframe(conn, df, TABLE_NAME, path)
            total_new      += new
            total_replaced += replaced
            log.info(f"  ✓ {os.path.basename(path)}: {new:,} new rows, {replaced:,} updated")
            all_dfs.append(df)
        except Exception as exc:
            log.error(f"  ✗ Failed on {path}: {exc}", exc_info=True)

    conn.close()
    log.info("─" * 60)
    log.info(f"DB load done. Total new: {total_new:,}  |  Total updated: {total_replaced:,}")

    if not all_dfs:
        log.error("No data loaded — cannot build Excel")
        sys.exit(1)

    df_all = pd.concat(all_dfs, ignore_index=True)
    log.info(f"Total rows for Excel mapping: {len(df_all):,}")

    mapped_add: list[dict] = []
    mapped_chg_fin: list[dict] = []
    mapped_chg_ord: list[dict] = []

    for _, row in df_all.iterrows():
        mdr_pid = _col(row, "mdr_pid")

        rec = map_row_610_add(row)
        if (mdr_pid == ""
                and rec["CUNM"] != ""
                and rec["CUA1"] != ""
                and len(rec["CUA1"]) <= 36
                and len(rec["TOWN"]) <= 20):
            mapped_add.append(rec)

        rec = map_row_610_chg_financial(row)
        if (mdr_pid == ""
                and rec["CUNM"] != ""
                and rec["CUA1"] != ""
                and len(rec["CUA1"]) <= 36
                and len(rec["TOWN"]) <= 20
                and rec["CUCD"] != "USD"):
            mapped_chg_fin.append(rec)

        rec = map_row_610_chg_order_info(row)
        if (mdr_pid == ""
                and rec["CUNM"] != ""
                and rec["CUA1"] != ""
                and len(rec["CUA1"]) <= 36
                and len(rec["TOWN"]) <= 20):
            mapped_chg_ord.append(rec)

    log.info(f"  Mapped {len(mapped_add):,} records → CRS610MI.Add")
    log.info(f"  Mapped {len(mapped_chg_fin):,} records → CRS610MI.ChgFinancial")
    log.info(f"  Mapped {len(mapped_chg_ord):,} records → CRS610MI.ChgOrderInfo")

    wb = build_workbook(mapped_add, mapped_chg_fin, mapped_chg_ord)

    os.makedirs(output_folder, exist_ok=True)
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_folder, f"HJ_Customers_{timestamp}.xlsx")
    wb.save(output_path)
    log.info(f"  Saved: {output_path}")

    write_sql_files()
    log.info(f"  SQL files written to: {SCRIPT_DIR}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Load Syndigo Organization CSVs into SyndigoOrg and generate CRS610MI Excel"
    )
    parser.add_argument(
        "--input-folder",
        default=DEFAULT_INPUT_FOLDER,
        help=f"Folder containing *.csv files (default: {DEFAULT_INPUT_FOLDER})",
    )
    parser.add_argument(
        "--output-folder",
        default=DEFAULT_OUTPUT_FOLDER,
        help=f"Folder for output .xlsx files (default: {DEFAULT_OUTPUT_FOLDER})",
    )
    parser.add_argument(
        "--db",
        default=DEFAULT_DB_PATH,
        help=f"Path to doppio.db (default: {DEFAULT_DB_PATH})",
    )
    args = parser.parse_args()

    log.info("=" * 60)
    log.info("HJ_Syndigo_Organization  —  CRS610MI.Add + ChgFinancial + ChgOrderInfo")
    log.info("=" * 60)

    if not os.path.isdir(args.input_folder):
        log.error(f"Input folder not found: {args.input_folder}")
        sys.exit(1)

    process(args.input_folder, args.output_folder, args.db)
    log.info("Done.")


if __name__ == "__main__":
    main()
