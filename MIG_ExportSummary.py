#!/usr/bin/env python3
"""
MIG_ExportSummary.py
====================
Reads .log files from the input folder (or specified paths) and produces a
detailed Excel workbook summary for each file.

Output: Same filename as the log, with .xlsx extension, saved alongside the log.

Usage:
    python MIG_ExportSummary.py                     # scan current folder for *.log
    python MIG_ExportSummary.py path/to/file.log    # specific file(s)
    python MIG_ExportSummary.py /path/to/folder     # all *.log in folder
"""

import re
import sys
import os
import sqlite3
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


DB_PATH = Path.home() / 'sqlite' / 'doppio.db'


def _load_column_flags():
    """
    Returns (cono_tables, divi_tables, known_tables) — each a set of upper-cased
    table names drawn from m3TableCols in doppio.db.
    Falls back to empty sets if the DB is unavailable.
    """
    try:
        con = sqlite3.connect(DB_PATH)
        cur = con.cursor()
        cur.execute("SELECT DISTINCT tablename FROM m3TableCols WHERE ColumnName LIKE '%CONO'")
        cono_tables = {r[0].upper() for r in cur.fetchall()}
        cur.execute("SELECT DISTINCT tablename FROM m3TableCols WHERE ColumnName LIKE '%DIVI'")
        divi_tables = {r[0].upper() for r in cur.fetchall()}
        cur.execute("SELECT DISTINCT tablename FROM m3TableCols")
        known_tables = {r[0].upper() for r in cur.fetchall()}
        con.close()
        return cono_tables, divi_tables, known_tables
    except Exception:
        return set(), set(), set()


# ---------------------------------------------------------------------------
# Colour / style constants
# ---------------------------------------------------------------------------
CLR_HEADER_DARK  = "2E4057"   # dark blue-grey  – sheet column headers
CLR_HEADER_MED   = "4A7C9E"   # mid-blue        – section titles
CLR_HEADER_RED   = "C0392B"   # red             – error headers
CLR_WHITE        = "FFFFFF"
CLR_GREEN_FILL   = "C6EFCE"   # light green     – rows with data
CLR_AMBER_FILL   = "FFF3CD"   # amber           – warnings
CLR_RED_FILL     = "FFE0E0"   # light red       – errors
CLR_ALT_FILL     = "EAF0F6"   # pale blue       – alternating rows
CLR_TOTAL_FILL   = "D9E2EC"   # grey            – totals row


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

_TS_PAT = re.compile(r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d+)')

def _parse_ts(s):
    try:
        return datetime.strptime(s, '%Y-%m-%d %H:%M:%S,%f')
    except Exception:
        return None


def parse_log(log_path):
    """
    Parse an Export.log file and return a dict with all extracted data.

    Keys returned
    -------------
    meta          – dict of export configuration settings
    copy_events   – {table: {start, end, rows}}
    summary_rows  – list of {table, records, duration_s}
    errors        – list of {level, timestamp, message, line_no}
    skipped       – list of {table, reason}  (explicit SKIP entries)
    footer        – {total_rows, total_seconds, rows_per_sec, status,
                     total_time, start_time, end_time}
    raw_selection – the table-selection regex string (from header)
    """

    PAT_COPY_START = re.compile(r'INFO\s+\{(\w+)\} Copy \1\s*$')
    PAT_COPY_DONE  = re.compile(r'INFO\s+\{(\w+)\} \1: Copied (\d+) rows')
    PAT_SUMMARY    = re.compile(
        r'INFO\s+Table\s+:\s+(\w+) \((\d[\d,]*) records(?:,\s*(\d+)s)?\)'
    )
    PAT_EXECUTE    = re.compile(
        r'INFO\s+execute ([\d,]+) in (\d+)s\.\s+\(([\d,]+) row/s\)'
    )
    PAT_TOTAL_TIME = re.compile(
        r'INFO\s+Jcpydta (succeeded|failed|SUCCEEDED|FAILED),\s+total time\s+(\S+)'
    )
    PAT_WARN       = re.compile(r'\bWARN\b')
    PAT_ERROR      = re.compile(r'\bERROR\b')
    PAT_SKIP       = re.compile(r'\b(?:SKIP|skip|Skipping|skipping|Skipped|skipped)\b')
    PAT_SCHEMA     = re.compile(r'Schema\s+(\w+)')
    PAT_DS_NAME    = re.compile(r"getName\(\)=([^,\]]+)")
    PAT_SELECTION  = re.compile(r'table-selection\s+(.+)')

    data = {
        'meta': {},
        'copy_events': {},
        'summary_rows': [],
        'errors': [],
        'skipped': [],
        'footer': {},
        'raw_selection': '',
    }

    with open(log_path, 'r', encoding='utf-8', errors='replace') as fh:
        lines = fh.readlines()

    # ---- first-pass: header lines (before log timestamps start) ----
    meta = data['meta']
    for raw in lines[:50]:
        line = raw.rstrip()
        if line.startswith('Export data'):
            continue
        elif line.startswith('Component:'):
            meta['component'] = line.split(':', 1)[1].strip()
        elif line.startswith('All companies:'):
            meta['all_companies'] = line.split(':', 1)[1].strip()
        elif line.startswith('Company:'):
            raw_co = line.split(':', 1)[1].strip()
            meta['company'] = raw_co.zfill(3) if raw_co.isdigit() else raw_co
        elif line.startswith('Filter on division:'):
            meta['filter_division'] = line.split(':', 1)[1].strip()
        m = PAT_SCHEMA.search(line)
        if m:
            meta['schema'] = m.group(1)
        m = PAT_DS_NAME.search(line)
        if m:
            meta.setdefault('datasource', m.group(1))

    # ---- record start_time from first timestamp in file ----
    for raw in lines:
        m = _TS_PAT.match(raw)
        if m:
            data['footer']['start_time'] = _parse_ts(m.group(1))
            break

    # ---- second-pass: full log ----
    for lineno, raw in enumerate(lines, 1):
        line = raw.rstrip()

        ts_m = _TS_PAT.match(line)
        ts   = _parse_ts(ts_m.group(1)) if ts_m else None

        # table-selection regex (capture once)
        sel_m = PAT_SELECTION.search(line)
        if sel_m and not data['raw_selection']:
            data['raw_selection'] = sel_m.group(1).strip()

        # fetch / thread settings
        if 'fetch.size=' in line:
            meta['fetch_size'] = line.split('fetch.size=', 1)[1].strip()
        if 'noofthreads=' in line:
            meta['threads'] = line.split('noofthreads=', 1)[1].strip()
        if 'clearTable' in line:
            meta['clear_table'] = line.split('clearTable', 1)[1].strip()

        # WARN / ERROR
        if ts_m and PAT_ERROR.search(line):
            data['errors'].append({
                'level': 'ERROR', 'timestamp': ts,
                'message': line, 'line_no': lineno
            })
        elif ts_m and PAT_WARN.search(line):
            data['errors'].append({
                'level': 'WARN', 'timestamp': ts,
                'message': line, 'line_no': lineno
            })

        # Explicit SKIP
        if ts_m and PAT_SKIP.search(line) and 'Table' in line:
            tbl_m = re.search(r'\b([A-Z][A-Z0-9]{3,5})\b', line)
            data['skipped'].append({
                'table': tbl_m.group(1) if tbl_m else '(unknown)',
                'reason': line,
                'line_no': lineno
            })

        # Copy start
        cs_m = PAT_COPY_START.search(line)
        if cs_m:
            tbl = cs_m.group(1)
            ev  = data['copy_events'].setdefault(tbl, {})
            ev['start'] = ts
            continue

        # Copy done
        cd_m = PAT_COPY_DONE.search(line)
        if cd_m:
            tbl  = cd_m.group(1)
            rows = int(cd_m.group(2))
            ev   = data['copy_events'].setdefault(tbl, {})
            ev['end']  = ts
            ev['rows'] = rows
            continue

        # Summary table line
        su_m = PAT_SUMMARY.search(line)
        if su_m:
            data['summary_rows'].append({
                'table':      su_m.group(1),
                'records':    int(su_m.group(2).replace(',', '')),
                'duration_s': int(su_m.group(3)) if su_m.group(3) else None,
            })
            continue

        # Footer: execute line
        ex_m = PAT_EXECUTE.search(line)
        if ex_m:
            data['footer']['total_rows']    = int(ex_m.group(1).replace(',', ''))
            data['footer']['total_seconds'] = int(ex_m.group(2))
            data['footer']['rows_per_sec']  = int(ex_m.group(3).replace(',', ''))
            continue

        # Footer: final status
        ft_m = PAT_TOTAL_TIME.search(line)
        if ft_m:
            data['footer']['status']     = ft_m.group(1).lower()
            data['footer']['total_time'] = ft_m.group(2)
            data['footer']['end_time']   = ts
            continue

    return data


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr_cell(ws, row, col, value,
              bg=CLR_HEADER_DARK, fg=CLR_WHITE, size=10, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name='Arial', bold=True, color=fg, size=size, italic=italic)
    c.fill      = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border    = _thin_border()
    return c


def _data_cell(ws, row, col, value, bold=False, fg='000000',
               bg=None, number_fmt=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name='Arial', bold=bold, color=fg)
    c.alignment = Alignment(vertical='center', wrap_text=wrap)
    c.border    = _thin_border()
    if bg:
        c.fill = PatternFill('solid', start_color=bg)
    if number_fmt:
        c.number_format = number_fmt
    return c


def _section_title(ws, row, col_span, text,
                   bg=CLR_HEADER_MED, fg=CLR_WHITE):
    ws.cell(row=row, column=1, value=text).font = \
        Font(name='Arial', bold=True, color=fg, size=11)
    ws.cell(row=row, column=1).fill = PatternFill('solid', start_color=bg)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    if col_span > 1:
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=col_span
        )
    ws.row_dimensions[row].height = 20


def _auto_width(ws, min_w=10, max_w=55):
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=0
        )
        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max_w, max(min_w, max_len + 3))


def _freeze(ws, row=2, col=1):
    ws.freeze_panes = ws.cell(row=row, column=col)


def _row_bg(i):
    return CLR_ALT_FILL if i % 2 == 0 else None


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _sheet_overview(wb, log_path, data):
    ws = wb.create_sheet('Overview')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 44

    meta   = data['meta']
    footer = data['footer']
    rows   = data['summary_rows']
    errors = data['errors']

    total_tables   = len(rows)
    tables_w_data  = sum(1 for r in rows if r['records'] > 0)
    empty_tables   = total_tables - tables_w_data
    skipped_tables = len(data['skipped'])

    status = footer.get('status', 'unknown')
    status_color = '006400' if 'succeed' in status else 'CC0000'

    # ---- Title ----
    ws.row_dimensions[1].height = 28
    ws.merge_cells('A1:B1')
    c = ws['A1']
    c.value     = 'Infor M3 — Export Summary Report'
    c.font      = Font(name='Arial', bold=True, size=14, color=CLR_HEADER_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill      = PatternFill('solid', start_color='EAF0F6')

    ws.merge_cells('A2:B2')
    c2 = ws['A2']
    c2.value     = f'Source file: {Path(log_path).name}'
    c2.font      = Font(name='Arial', italic=True, size=9, color='555555')
    c2.alignment = Alignment(horizontal='center')

    row = 4

    # ---- Export Configuration ----
    _section_title(ws, row, 2, '  Export Configuration')
    row += 1

    cfg_items = [
        ('Component',         meta.get('component', 'N/A')),
        ('Source Schema',     meta.get('schema',    'N/A')),
        ('Data Source',       meta.get('datasource','N/A')),
        ('Fetch Size',        meta.get('fetch_size','N/A')),
        ('Copy Threads',      meta.get('threads',   'N/A')),
        ('Company Filter',    meta.get('company',   'N/A')),
        ('All Companies',     meta.get('all_companies','N/A')),
        ('Filter on Division',meta.get('filter_division','N/A')),
        ('Clear Table',       meta.get('clear_table','N/A')),
        ('Table Selection',   data.get('raw_selection','N/A')),
    ]
    for label, val in cfg_items:
        _data_cell(ws, row, 1, label, bold=True)
        c = _data_cell(ws, row, 2, val, wrap=(len(str(val)) > 40))
        if label == 'Table Selection':
            c.font = Font(name='Courier New', size=8, color='444444')
            ws.row_dimensions[row].height = 30
        row += 1

    row += 1

    # ---- Execution Results ----
    _section_title(ws, row, 2, '  Execution Results')
    row += 1

    start_ts = footer.get('start_time')
    end_ts   = footer.get('end_time')

    exec_items = [
        ('Status',                status.upper()),
        ('Start Time',            start_ts.strftime('%Y-%m-%d  %H:%M:%S') if start_ts else 'N/A'),
        ('End Time',              end_ts.strftime('%Y-%m-%d  %H:%M:%S')   if end_ts   else 'N/A'),
        ('Total Duration',        footer.get('total_time', 'N/A')),
        ('Total Records Exported',footer.get('total_rows', 0)),
        ('Throughput',            f"{footer.get('rows_per_sec', 0):,} rows / second"),
    ]
    for label, val in exec_items:
        _data_cell(ws, row, 1, label, bold=True)
        c = _data_cell(ws, row, 2, val)
        if label == 'Status':
            c.font = Font(name='Arial', bold=True, color=status_color)
        if label == 'Total Records Exported' and isinstance(val, int):
            c.number_format = '#,##0'
        row += 1

    row += 1

    # ---- Table Statistics ----
    _section_title(ws, row, 2, '  Table Statistics')
    row += 1

    stat_items = [
        ('Total Tables Processed',   total_tables),
        ('Tables with Data',         tables_w_data),
        ('Empty Tables (0 records)', empty_tables),
        ('Explicitly Skipped',       skipped_tables),
        ('Errors / Warnings',        len(errors)),
    ]
    for label, val in stat_items:
        _data_cell(ws, row, 1, label, bold=True)
        c = _data_cell(ws, row, 2, val)
        c.number_format = '#,##0'
        if label == 'Errors / Warnings' and val > 0:
            c.font = Font(name='Arial', bold=True, color='CC0000')
        elif label in ('Tables with Data',) and val > 0:
            c.font = Font(name='Arial', bold=True, color='006400')
        row += 1

    ws.freeze_panes = 'A1'


def _sheet_tables_with_data(wb, data):
    ws = wb.create_sheet('Tables With Data')
    ws.sheet_view.showGridLines = False

    copy_events = data['copy_events']
    with_data   = sorted(
        [r for r in data['summary_rows'] if r['records'] > 0],
        key=lambda x: x['records'], reverse=True
    )

    COLS = [
        ('Rank',            6),
        ('Table Name',     14),
        ('Records Copied', 16),
        ('Duration (s)',   13),
        ('Elapsed (ms)',   13),
        ('Copy Start',     16),
        ('Copy End',       16),
    ]
    for ci, (hdr, w) in enumerate(COLS, 1):
        _hdr_cell(ws, 1, ci, hdr)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22

    # Colour scale: top 10% get a darker green, rest light green
    total = len(with_data)
    top10 = max(1, int(total * 0.10))

    for i, t in enumerate(with_data, 1):
        ev    = copy_events.get(t['table'], {})
        st    = ev.get('start')
        en    = ev.get('end')
        el_ms = int((en - st).total_seconds() * 1000) if st and en else None
        bg    = ('A8D5B5' if i <= top10 else CLR_GREEN_FILL)

        row = i + 1
        _data_cell(ws, row, 1, i,          bg=bg, number_fmt='#,##0')
        _data_cell(ws, row, 2, t['table'], bg=bg, bold=True)
        _data_cell(ws, row, 3, t['records'],bg=bg, number_fmt='#,##0')
        _data_cell(ws, row, 4, t['duration_s'], bg=bg)
        _data_cell(ws, row, 5, el_ms,      bg=bg, number_fmt='#,##0')
        _data_cell(ws, row, 6, st.strftime('%H:%M:%S.%f')[:-3] if st else '', bg=bg)
        _data_cell(ws, row, 7, en.strftime('%H:%M:%S.%f')[:-3] if en else '', bg=bg)

    # Totals
    tr = total + 2
    _data_cell(ws, tr, 1, 'TOTAL', bold=True, bg=CLR_TOTAL_FILL)
    c = ws.cell(row=tr, column=3, value=f'=SUM(C2:C{tr-1})')
    c.font   = Font(name='Arial', bold=True)
    c.fill   = PatternFill('solid', start_color=CLR_TOTAL_FILL)
    c.number_format = '#,##0'
    c.border = _thin_border()

    _freeze(ws)


def _sheet_empty_tables(wb, data):
    ws = wb.create_sheet('Empty Tables')
    ws.sheet_view.showGridLines = False

    copy_events = data['copy_events']
    empty       = [r for r in data['summary_rows'] if r['records'] == 0]

    COLS = [
        ('#',          6),
        ('Table Name', 14),
        ('Records',    10),
        ('Elapsed (ms)', 13),
        ('Copy Start', 16),
        ('Copy End',   16),
    ]
    for ci, (hdr, w) in enumerate(COLS, 1):
        _hdr_cell(ws, 1, ci, hdr, bg='607D8B')
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22

    for i, t in enumerate(empty, 1):
        ev    = copy_events.get(t['table'], {})
        st    = ev.get('start')
        en    = ev.get('end')
        el_ms = int((en - st).total_seconds() * 1000) if st and en else None
        bg    = _row_bg(i)
        row   = i + 1

        _data_cell(ws, row, 1, i,           bg=bg, number_fmt='#,##0')
        _data_cell(ws, row, 2, t['table'],  bg=bg)
        _data_cell(ws, row, 3, 0,           bg=bg, number_fmt='#,##0')
        _data_cell(ws, row, 4, el_ms,       bg=bg, number_fmt='#,##0')
        _data_cell(ws, row, 5, st.strftime('%H:%M:%S.%f')[:-3] if st else '', bg=bg)
        _data_cell(ws, row, 6, en.strftime('%H:%M:%S.%f')[:-3] if en else '', bg=bg)

    # row count note
    note_row = len(empty) + 3
    ws.cell(row=note_row, column=1,
            value=f'Total empty tables: {len(empty):,}').font = \
        Font(name='Arial', italic=True, color='555555')

    _freeze(ws)


def _sheet_all_tables(wb, data, cono_tables, divi_tables, known_tables):
    ws = wb.create_sheet('All Tables')
    ws.sheet_view.showGridLines = False

    copy_events = data['copy_events']

    COLS = [
        ('#',                6),
        ('Table Name',      14),
        ('Records',         14),
        ('Has Data',        10),
        ('Has Company',     14),
        ('Has Division',    14),
        ('Duration (s)',    13),
        ('Elapsed (ms)',    13),
        ('Copy Start',      16),
        ('Copy End',        16),
    ]
    for ci, (hdr, w) in enumerate(COLS, 1):
        _hdr_cell(ws, 1, ci, hdr)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22

    for i, t in enumerate(data['summary_rows'], 1):
        ev       = copy_events.get(t['table'], {})
        st       = ev.get('start')
        en       = ev.get('end')
        el_ms    = int((en - st).total_seconds() * 1000) if st and en else None
        has_data = t['records'] > 0
        bg       = CLR_GREEN_FILL if has_data else _row_bg(i)
        row      = i + 1
        tbl_up   = t['table'].upper()

        if tbl_up in known_tables:
            cono_val = 'Yes' if tbl_up in cono_tables else 'No'
            divi_val = 'Yes' if tbl_up in divi_tables else 'No'
            cono_fg  = '006400' if tbl_up in cono_tables else 'CC0000'
            divi_fg  = '006400' if tbl_up in divi_tables else 'CC0000'
        else:
            cono_val = divi_val = '-'
            cono_fg  = divi_fg  = '888888'

        _data_cell(ws, row, 1, i,              bg=bg, number_fmt='#,##0')
        _data_cell(ws, row, 2, t['table'],     bg=bg, bold=has_data)
        _data_cell(ws, row, 3, t['records'],   bg=bg, number_fmt='#,##0')
        _data_cell(ws, row, 4, 'Yes' if has_data else 'No',
                   bg=bg, fg=('006400' if has_data else '888888'))
        _data_cell(ws, row, 5, cono_val,       bg=bg, fg=cono_fg)
        _data_cell(ws, row, 6, divi_val,       bg=bg, fg=divi_fg)
        _data_cell(ws, row, 7, t['duration_s'],bg=bg)
        _data_cell(ws, row, 8, el_ms,          bg=bg, number_fmt='#,##0')
        _data_cell(ws, row, 9, st.strftime('%H:%M:%S.%f')[:-3] if st else '', bg=bg)
        _data_cell(ws, row, 10, en.strftime('%H:%M:%S.%f')[:-3] if en else '', bg=bg)

    _freeze(ws)


def _sheet_errors(wb, data):
    ws = wb.create_sheet('Errors & Warnings')
    ws.sheet_view.showGridLines = False

    errors = data['errors']

    if not errors:
        ws.column_dimensions['A'].width = 50
        c = ws.cell(row=2, column=1,
                    value='✓  No errors or warnings detected in this log.')
        c.font = Font(name='Arial', color='006400', bold=True, size=11)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A2:E2')
        ws.row_dimensions[2].height = 30
        return

    COLS = [
        ('Line #',    8),
        ('Level',     9),
        ('Timestamp', 20),
        ('Message',   80),
    ]
    for ci, (hdr, w) in enumerate(COLS, 1):
        _hdr_cell(ws, 1, ci, hdr, bg=CLR_HEADER_RED)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22

    for i, e in enumerate(errors, 1):
        bg  = CLR_RED_FILL if e['level'] == 'ERROR' else CLR_AMBER_FILL
        row = i + 1
        _data_cell(ws, row, 1, e.get('line_no'),  bg=bg)
        _data_cell(ws, row, 2, e['level'],         bg=bg, bold=True,
                   fg=('CC0000' if e['level'] == 'ERROR' else 'B7800E'))
        ts = e.get('timestamp')
        _data_cell(ws, row, 3,
                   ts.strftime('%Y-%m-%d %H:%M:%S') if ts else '',
                   bg=bg)
        _data_cell(ws, row, 4, e['message'], bg=bg, wrap=True)
        ws.row_dimensions[row].height = 30

    _freeze(ws)


def _expand_exclusion_filter(raw_selection):
    """
    Parse the negative-lookahead exclusion regex from the log header and return:
      individual_tables : list of concrete 6-character table names
      pattern_rules     : list of (pattern_str, plain-English description)
    Only 6-character names are included in individual_tables.
    """
    individual_tables = []
    pattern_rules     = []

    if not raw_selection:
        return individual_tables, pattern_rules

    # ── Step 1: extract everything inside  (?!  ...  )  ────────────────────
    # Walk character by character to find the matching closing paren of (?!
    start = raw_selection.find('(?!')
    if start == -1:
        return individual_tables, pattern_rules

    depth, i = 0, start
    while i < len(raw_selection):
        if raw_selection[i] == '(':
            depth += 1
        elif raw_selection[i] == ')':
            depth -= 1
            if depth == 0:
                break
        i += 1

    # content between (?! and the matching )
    inner = raw_selection[start + 3 : i]

    # inner may itself be wrapped in one extra paren group — strip it
    if inner.startswith('(') and inner.endswith(')'):
        inner = inner[1:-1]

    # ── Step 2: split on top-level | respecting nested parens ───────────────
    def split_top(s):
        parts, depth, buf = [], 0, []
        for ch in s:
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
            if ch == '|' and depth == 0:
                parts.append(''.join(buf))
                buf = []
            else:
                buf.append(ch)
        if buf:
            parts.append(''.join(buf))
        return parts

    # ── Step 3: expand PREFIX(A|B|C) → [PREFIXA, PREFIXB, PREFIXC] ─────────
    def expand(token):
        m = re.match(r'^([A-Z0-9]+)\(([^)]+)\)$', token)
        if m:
            prefix = m.group(1)
            return [prefix + alt for alt in m.group(2).split('|')]
        return [token]

    # ── Step 4: classify each token ─────────────────────────────────────────
    for token in split_top(inner):
        token = token.strip()
        if not token:
            continue

        # Contains regex metacharacters → pattern rule
        if re.search(r'[\[\]{\\.*+?^$]', token):
            if re.search(r'O\[0-9\]\{6,6\}', token):
                desc = 'Tables named "O" + exactly 6 digits  (e.g. O000001)'
            else:
                desc = 'Pattern-based exclusion rule'
            pattern_rules.append((token, desc))
            continue

        # Expand and keep only valid 6-character table names
        for name in expand(token):
            if len(name) == 6 and re.match(r'^[A-Z][A-Z0-9]{5}$', name):
                individual_tables.append(name)

    return individual_tables, pattern_rules


def _sheet_skipped(wb, data):
    ws = wb.create_sheet('Skipped Tables')
    ws.sheet_view.showGridLines = False

    sel = data.get('raw_selection', '')

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 55

    individual_tables, pattern_rules = _expand_exclusion_filter(sel)

    row = 1

    # ── Raw filter regex ────────────────────────────────────────────────────
    _section_title(ws, row, 2, '  Table Exclusion Filter  (from log header)')
    row += 1

    if sel:
        c = ws.cell(row=row, column=1, value=sel)
        c.font      = Font(name='Courier New', size=8, color='333333')
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:B{row}')
        ws.row_dimensions[row].height = 42
        row += 1
    else:
        ws.cell(row=row, column=1, value='(not found in log)')
        row += 1

    row += 1

    # ── Excluded table names (concrete, 6-char) ─────────────────────────────
    _section_title(ws, row, 2, f'  Excluded Tables  ({len(individual_tables)} names)')
    row += 1

    _hdr_cell(ws, row, 1, 'Table Name', bg='607D8B')
    _hdr_cell(ws, row, 2, 'Type',       bg='607D8B')
    row += 1

    for i, name in enumerate(individual_tables):
        bg = _row_bg(i)
        _data_cell(ws, row, 1, name, bold=True, bg=bg)
        _data_cell(ws, row, 2, 'Exact table name (6 characters)', bg=bg)
        row += 1

    row += 1

    # ── Pattern-based exclusion rules ────────────────────────────────────────
    _section_title(ws, row, 2, f'  Pattern-Based Exclusion Rules  ({len(pattern_rules)} rules)')
    row += 1

    _hdr_cell(ws, row, 1, 'Pattern',     bg='607D8B')
    _hdr_cell(ws, row, 2, 'Description', bg='607D8B')
    row += 1

    for i, (pat, desc) in enumerate(pattern_rules):
        bg = _row_bg(i)
        c = ws.cell(row=row, column=1, value=pat)
        c.font      = Font(name='Courier New', size=9, color='333333')
        c.fill      = PatternFill('solid', start_color=bg) if bg else PatternFill()
        c.border    = _thin_border()
        _data_cell(ws, row, 2, desc, bg=bg)
        row += 1


# ---------------------------------------------------------------------------
# Workbook factory
# ---------------------------------------------------------------------------

def build_workbook(log_path, data):
    wb = Workbook()
    wb.remove(wb.active)               # remove default blank sheet

    cono_tables, divi_tables, known_tables = _load_column_flags()

    _sheet_overview(wb, log_path, data)
    _sheet_tables_with_data(wb, data)
    _sheet_empty_tables(wb, data)
    _sheet_all_tables(wb, data, cono_tables, divi_tables, known_tables)
    _sheet_errors(wb, data)
    _sheet_skipped(wb, data)

    return wb


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def process_log(log_path):
    log_path = Path(log_path).resolve()
    if not log_path.exists():
        print(f"  [ERROR] File not found: {log_path}")
        return False

    print(f"  Parsing  : {log_path.name}")
    data = parse_log(log_path)

    wb = build_workbook(log_path, data)

    meta      = data['meta']
    datasource = re.sub(r'[^\w\-]', '_', meta.get('datasource', 'UNKNOWN'))
    company    = meta.get('company', '000')
    footer     = data['footer']
    start_ts   = footer.get('start_time')
    dt_tag     = start_ts.strftime('%Y%m%d_%H%M%S') if start_ts else 'nodate'

    out_dir = DEFAULT_OUTPUT_FOLDER
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{datasource}_{company}_{dt_tag}.xlsx"
    wb.save(str(out_path))

    footer = data['footer']
    tables = data['summary_rows']
    print(f"  Status   : {footer.get('status','?').upper()}")
    print(f"  Duration : {footer.get('total_time','?')}")
    print(f"  Records  : {footer.get('total_rows', 0):,}")
    print(f"  Tables   : {len(tables):,}  ({sum(1 for t in tables if t['records'] > 0):,} with data)")
    print(f"  Errors   : {len(data['errors'])}")
    print(f"  Saved    : {out_path.name}")
    print()
    return True


DEFAULT_INPUT_FOLDER  = Path.home() / 'Doppio' / 'input'
DEFAULT_OUTPUT_FOLDER = Path.home() / 'Doppio' / 'output'


def main():
    targets = sys.argv[1:] if len(sys.argv) > 1 else [str(DEFAULT_INPUT_FOLDER)]

    log_files = []
    for t in targets:
        p = Path(t)
        if p.is_dir():
            log_files.extend(sorted(p.glob('*.log')))
        elif p.suffix.lower() == '.log':
            log_files.append(p)
        else:
            print(f"  [SKIP] Not a .log file or directory: {t}")

    if not log_files:
        print("No .log files found.")
        print("Usage: python MIG_ExportSummary.py [file.log | folder]")
        return

    print(f"\nMIG_ExportSummary — processing {len(log_files)} file(s)\n{'='*55}")
    ok = 0
    for lf in log_files:
        ok += process_log(lf)
    print(f"Done. {ok}/{len(log_files)} file(s) processed successfully.")


if __name__ == '__main__':
    main()
