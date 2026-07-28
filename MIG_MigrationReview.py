# MIG_MigrationReview.py
#
# Monitors the [input] folder for Excel workbooks that contain a [DTA_FIN_MVX]
# sheet, queries live M3 record counts via EXPORTMI, and produces a
# [migration review] copy with five new columns:
#   J  MT actual               – live count returned by the API
#   K  MT count vs actual      – =J{r}-F{r}
#   L  Precent Diff vs Actual  – =IFERROR(ABS(F-J)/((F+J)/2),0)  [0% format]
#   M  Table Description       – looked up from m3tables in local SQLite DB
#   N  Maintained By           – looked up from m3tables in local SQLite DB
#
# The original workbook is moved to [processed]; the reviewed copy lands in
# [migration review] with a _REVIEWED suffix.

import shutil
import sqlite3
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from tqdm import tqdm

from config import BASE_DIR, get_sqlite_db_path
from InforMI import CONFIG, get_ion_token, select_ionapi_file, prompt_for_company_division

SQLITE_DB_PATH = get_sqlite_db_path()

# =============================================================================
# Folder layout  (all siblings of this script)
# =============================================================================
INPUT_DIR    = BASE_DIR / "input"
PROCESSED_DIR = BASE_DIR / "processed"
OUTPUT_DIR   = BASE_DIR / "migration review"

REQUIRED_SHEET = "DTA_FIN_MVX"

# Column positions (1-based) for the five new columns
COL_J = 10   # MT actual
COL_K = 11   # MT count vs actual
COL_L = 12   # Precent Diff vs Actual
COL_M = 13   # Table Description
COL_N = 14   # Maintained By

# =============================================================================
# Workbook helpers
# =============================================================================

def _find_header_row(ws) -> int | None:
    """Return the 1-based row number of the row whose col A = 'Library'
    and col B = 'Filename', or None if not found."""
    for row in ws.iter_rows():
        if (str(row[0].value or "").strip().lower() == "library" and
                str(row[1].value or "").strip().lower() == "filename"):
            return row[0].row
    return None


def _collect_data_rows(ws, header_row: int) -> list[tuple[int, str]]:
    """Return a list of (excel_row_number, filename) for every data row
    that has a non-blank value in column B (Filename)."""
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        filename = str(row[1].value or "").strip()   # col B
        if filename:
            rows.append((row[0].row, filename))
    return rows


# =============================================================================
# SQLite lookup – m3tables (Table Description / Maintained By)
# =============================================================================

def _lookup_table_info(filenames: list[str]) -> dict[str, tuple[str, str]]:
    """Return {filename: (tableDescription, tableMaintainedBy)} for every
    filename in *filenames*, using a single bulk query against m3tables."""
    if not filenames:
        return {}

    placeholders = ",".join("?" * len(filenames))
    sql = (
        f"SELECT tablename, tableDescription, tableMaintainedBy "
        f"FROM m3tables WHERE tablename IN ({placeholders})"
    )

    result: dict[str, tuple[str, str]] = {}
    try:
        with sqlite3.connect(SQLITE_DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            for row in conn.execute(sql, filenames):
                result[row["tablename"]] = (
                    row["tableDescription"] or "",
                    row["tableMaintainedBy"] or "",
                )
    except Exception as exc:
        tqdm.write(f"  ⚠️   SQLite lookup failed: {exc}")

    return result


# =============================================================================
# M3 API – EXPORTMI Select  (count(#) from <filename>)
# =============================================================================

def _get_live_count(filename: str, session: requests.Session, max_retries: int = 3) -> int | str:
    """Call EXPORTMI/Select and return the integer count for *filename*.
    Returns empty string on unrecoverable error."""
    payload = {
        "program": "EXPORTMI",
        "transactions": [{
            "transaction": "Select",
            "record": {"QERY": f"count(#) from {filename}"},
            "selectedColumns": ["QERY", "REPL"]
        }]
    }

    for attempt in range(1, max_retries + 1):
        try:
            headers = {
                "Authorization": f"Bearer {CONFIG['access_token']}",
                "Content-Type": "application/json; charset=UTF-8",
                "accept": "application/json; charset=UTF-8",
            }
            resp = session.post(
                CONFIG["api_url"],
                json=payload,
                headers=headers,
                timeout=300
            )

            if resp.status_code == 401:
                get_ion_token()
                continue

            resp.raise_for_status()
            data = resp.json()
            records = data.get("results", [{}])[0].get("records", [])
            if records:
                return int(records[0].get("REPL", 0))
            return 0

        except Exception as exc:
            if attempt == max_retries:
                tqdm.write(f"    ⚠️  {filename}: {exc}")
                return ""
            get_ion_token()
            time.sleep(2)

    return ""


def _fetch_counts_parallel(data_rows: list[tuple[int, str]],
                            max_workers: int = 8) -> dict[int, int | str]:
    """Fetch live counts for all rows in parallel. Returns {row_num: count}."""
    results: dict[int, int | str] = {}

    with requests.Session() as session:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_row = {
                executor.submit(_get_live_count, filename, session): row_num
                for row_num, filename in data_rows
            }
            with tqdm(total=len(data_rows), desc="  Querying M3", unit="table", leave=False) as pbar:
                for future in as_completed(future_to_row):
                    row_num = future_to_row[future]
                    results[row_num] = future.result()
                    pbar.update(1)

    return results


# =============================================================================
# Sheet formatting helpers
# =============================================================================

_SUMMARY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_SUMMARY_FONT = Font(bold=True, color="000000")


def _write_summary(ws, header_row: int, last_data_row: int) -> None:
    """Write the four-row summary block at K5:L8.

    Counts are driven by Remarks (col I) and MT count vs actual (col K):
      matching table counts – blank remarks AND K = 0
      protected             – non-blank remarks (any K)
      ok to be different    – blank remarks AND K > 0
      questionable          – blank remarks AND K < 0
    """
    data_start = header_row + 1
    i_rng = f"I{data_start}:I{last_data_row}"   # Remarks
    k_rng = f"K{data_start}:K{last_data_row}"   # MT count vs actual

    summary = [
        (5, "matching table counts:", f'=COUNTIFS({i_rng},"",{k_rng},0)'),
        (6, "protected:",              f'=COUNTIF({i_rng},"<>")'),
        (7, "ok to be different:",     f'=COUNTIFS({i_rng},"",{k_rng},">0")'),
        (8, "questionable:",           f'=COUNTIFS({i_rng},"",{k_rng},"<0")'),
    ]

    for row_num, label, formula in summary:
        lbl = ws.cell(row=row_num, column=11)   # col K
        val = ws.cell(row=row_num, column=12)   # col L

        lbl.value = label
        lbl.font  = _SUMMARY_FONT
        lbl.fill  = _SUMMARY_FILL

        val.value = formula
        val.font  = _SUMMARY_FONT
        val.fill  = _SUMMARY_FILL


def _reset_autofilter(ws, header_row: int, last_data_row: int, last_col: int) -> None:
    """Clear the existing auto-filter and re-apply it across all columns
    from A{header_row} to the rightmost column / last data row."""
    ws.auto_filter.ref = None
    end_col_letter = get_column_letter(last_col)
    ws.auto_filter.ref = f"A{header_row}:{end_col_letter}{last_data_row}"


def _autofit_columns(ws, col_indices: list[int], header_row: int) -> None:
    """Set each column's width to fit its widest cell value (header + data).
    Caps width at 60 characters to avoid extremely wide columns."""
    for col_idx in col_indices:
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.row < header_row:
                continue          # skip rows above the header
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


# =============================================================================
# Core workbook processor
# =============================================================================

def process_workbook(file_path: Path) -> bool:
    """
    Process a single workbook.
    Returns True on success, False if the file should be skipped.
    """
    tqdm.write(f"\n📖  {file_path.name}")

    # --- load and validate ---
    try:
        wb = load_workbook(file_path)
    except Exception as exc:
        tqdm.write(f"  ❌  Could not open workbook: {exc}")
        return False

    if REQUIRED_SHEET not in wb.sheetnames:
        tqdm.write(f"  ⚠️   Sheet '{REQUIRED_SHEET}' not found – skipping.")
        return False

    ws = wb[REQUIRED_SHEET]
    header_row = _find_header_row(ws)

    if header_row is None:
        tqdm.write(f"  ⚠️   Header row (Library / Filename) not found in '{REQUIRED_SHEET}' – skipping.")
        return False

    # --- write / overwrite the five column headers ---
    ws.cell(row=header_row, column=COL_J).value = "MT actual"
    ws.cell(row=header_row, column=COL_K).value = "MT count vs actual"
    ws.cell(row=header_row, column=COL_L).value = "Precent Diff vs Actual"
    ws.cell(row=header_row, column=COL_M).value = "Table Description"
    ws.cell(row=header_row, column=COL_N).value = "Maintained By"

    # --- collect data rows ---
    data_rows = _collect_data_rows(ws, header_row)
    if not data_rows:
        tqdm.write("  ⚠️   No data rows found – skipping.")
        return False

    tqdm.write(f"  📊  {len(data_rows)} tables to count...")

    # --- query M3 in parallel + SQLite in one shot ---
    counts = _fetch_counts_parallel(data_rows)
    unique_filenames = list({fn for _, fn in data_rows})
    table_info = _lookup_table_info(unique_filenames)
    tqdm.write(f"  🗄   SQLite matched {len(table_info)}/{len(unique_filenames)} table(s).")

    # --- write results + formulas back into the sheet ---
    for row_num, filename in data_rows:
        count = counts.get(row_num, "")
        description, maintained_by = table_info.get(filename, ("", ""))

        j_cell = ws.cell(row=row_num, column=COL_J)
        k_cell = ws.cell(row=row_num, column=COL_K)
        l_cell = ws.cell(row=row_num, column=COL_L)
        m_cell = ws.cell(row=row_num, column=COL_M)
        n_cell = ws.cell(row=row_num, column=COL_N)

        j_cell.value = count
        k_cell.value = f"=J{row_num}-F{row_num}"
        l_cell.value = f"=IFERROR(ABS(F{row_num} - J{row_num}) / ((F{row_num} + J{row_num}) / 2),0)"
        l_cell.number_format = "0%"
        m_cell.value = description
        n_cell.value = maintained_by

    # --- summary block at K5:L8 ---
    last_data_row = data_rows[-1][0] if data_rows else header_row
    _write_summary(ws, header_row, last_data_row)

    # --- reset auto-filter to cover all columns including the new ones ---
    _reset_autofilter(ws, header_row, last_data_row, COL_N)

    # --- autofit the five new columns ---
    _autofit_columns(ws, [COL_J, COL_K, COL_L, COL_M, COL_N], header_row)

    # --- save reviewed copy ---
    output_name = file_path.stem + "_REVIEWED" + file_path.suffix
    output_path = OUTPUT_DIR / output_name
    try:
        wb.save(output_path)
    except Exception as exc:
        tqdm.write(f"  ❌  Could not save reviewed workbook: {exc}")
        return False

    # --- move original to processed ---
    try:
        shutil.move(str(file_path), PROCESSED_DIR / file_path.name)
    except Exception as exc:
        tqdm.write(f"  ⚠️   Could not move original to processed: {exc}")

    tqdm.write(f"  ✅  Saved → {output_path.name}")
    return True


# =============================================================================
# Entry point
# =============================================================================

def main() -> None:
    for d in (INPUT_DIR, PROCESSED_DIR, OUTPUT_DIR):
        d.mkdir(parents=True, exist_ok=True)

    # --- authenticate with M3 ---
    ionapi_dir = Path(__file__).parent / "ionapi"
    CONFIG["tenant"] = select_ionapi_file(ionapi_dir)
    prompt_for_company_division()
    get_ion_token()

    # --- find and process all workbooks in input/ ---
    files = sorted(
        [f for f in INPUT_DIR.glob("*.xlsx") if not f.name.startswith("~$")],
        key=lambda f: f.name.lower()
    )

    if not files:
        print(f"📂  No Excel files found in {INPUT_DIR}")
        return

    print(f"\n🔍  Found {len(files)} workbook(s) in {INPUT_DIR}\n")
    for file_path in files:
        process_workbook(file_path)

    print("\n✅  Done.")


if __name__ == "__main__":
    main()
