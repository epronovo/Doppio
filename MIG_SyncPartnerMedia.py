# MIG_SyncPartnerMedia.py
"""
MIG_SyncPartnerMedia.py
-----------------------
Syncs partner media data (CRS949) from a SOURCE tenant to a
DESTINATION tenant via CRS949MI.LstPartnerMedia, CRS949MI.DltPartnerMedia,
and CRS949MI.AddPartnerEmail.

Steps
-----
1.  Prompt for SOURCE tenant → CRS949MI.LstPartnerMedia (SOURCE DIVI) → records in memory.
2.  Prompt for DEST   tenant → CRS949MI.LstPartnerMedia (DEST   DIVI) → records in memory.
3.  Diff: find records where any value field differs (matched on key fields),
    and records missing from DEST entirely.  Display first 10 and pause for confirmation.
4.  Re-authenticate to DEST and:
      a. CRS949MI.DltPartnerMedia + CRS949MI.AddPartnerEmail – update (delete & re-add) records that exist but differ.
      b. CRS949MI.AddPartnerEmail – add records that are missing from DEST.
5.  Write a summary .xlsx of all API calls made.

Usage:
    python MIG_SyncPartnerMedia.py

Notes:
    - Batch size is fixed at 100 (no prompt).
    - SOURCE and DEST tenants are always prompted (no skip).
"""

import datetime
import os
import requests
from pathlib import Path
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from InforMI import (
    CONFIG,
    get_ion_token,
    post_to_m3,
)
from UserDefaults import load_user_defaults, save_user_defaults

# =============================================================================
# Constants
# =============================================================================

# All columns returned by CRS949MI.LstPartnerMedia
PARTNER_MEDIA_COLS = [
    "DIVI", "DONR", "DOVA", "MEPF", "PRF1", "PRF2", "MEDC", "SEQN",
    "SIID", "MVIF", "METY", "FMTP", "COPY", "OUTP", "UDTA", "1UDT",
    "TFNO", "TFT1", "1TFT", "FLRN", "PAFD", "FSUX", "GNNM", "DEV",
    "FOVR", "BINX", "TOMA", "FRMA", "CCMA", "FAXT", "SUBJ", "NOTE",
    "CPPL", "LSID", "LSAD", "TEME", "E065", "SRD1", "SRD2", "RRD1",
    "RRD2", "RRD3", "TRAY", "LAYC", "MARI", "ARCH", "CNID", "EVPR",
    "BONM", "BOVB", "FIET", "EMBT", "EMGR", "FILM", "FNAM", "GRCO", "PRTP",
]

# Composite key used to match SOURCE vs DEST rows
KEY_FIELDS = ["DIVI", "DONR", "DOVA", "MEPF", "PRF1", "PRF2", "MEDC", "SEQN"]

# Fields compared for differences (all non-key columns)
VALUE_FIELDS = [f for f in PARTNER_MEDIA_COLS if f not in KEY_FIELDS]

# Fixed batch size — no prompt
BATCH_SIZE = 100

# Maps MEDC value → the CRS949MI Add transaction to use.
# Records whose MEDC is not in this map are skipped entirely (not deleted, not added).
MEDC_TRANSACTION: dict[str, str] = {
    "MAIL":     "AddPartnerEmail",
    "MBMEVENT": "AddPartnerMBM",
}


def resolve_add_transaction(rec: dict) -> str | None:
    """Return the CRS949MI Add transaction for this record's MEDC, or None to skip."""
    return MEDC_TRANSACTION.get(rec.get("MEDC", "").strip())


# =============================================================================
# Tenant helpers
# =============================================================================

def snapshot_config() -> dict:
    return dict(CONFIG)


def restore_config(snapshot: dict) -> None:
    CONFIG.clear()
    CONFIG.update(snapshot)


def _select_ionapi_forced(ionapi_dir: Path, label: str) -> Path:
    """
    Always shows the ionapi file list and requires a selection — bypasses
    the 1-hour cache in InforMI.select_ionapi_file so SOURCE and DEST can
    use different files in the same run.
    """
    files = sorted(
        [f for f in os.listdir(ionapi_dir) if f.endswith(".ionapi")],
        key=str.lower,
    )
    if not files:
        raise FileNotFoundError(f"No .ionapi files found in: {ionapi_dir}")

    print(f"\n  🔹  Select the {label} ION API file:")
    for i, f in enumerate(files, start=1):
        print(f"    {i}. {f}")

    while True:
        try:
            choice = int(input(f"\n  Enter choice (1-{len(files)}): "))
            if 1 <= choice <= len(files):
                return ionapi_dir / files[choice - 1]
            print("  ❌  Invalid choice. Try again.")
        except ValueError:
            print("  ❌  Please enter a number.")


def _prompt_company_division_forced(label: str) -> None:
    """
    Always prompts for company and division — bypasses the 1-hour cache in
    InforMI.prompt_for_company_division so each tenant is configured fresh.
    """
    from UserDefaults import load_user_defaults, save_user_defaults
    defaults = load_user_defaults()

    default_company  = defaults.get("company",  "100")
    default_division = defaults.get("division", "500")

    print(f"\n  Configure {label} company / division:")
    company  = input(f"    Company  (default: {default_company}):  ").strip()
    division = input(f"    Division (default: {default_division}): ").strip()

    CONFIG["company"]  = company  if company  else default_company
    CONFIG["division"] = division if division else default_division

    # Persist as new defaults (but do NOT update last_prompt_time so the
    # next call is also forced to prompt)
    defaults["company"]  = CONFIG["company"]
    defaults["division"] = CONFIG["division"]
    save_user_defaults(defaults)


def setup_tenant(ionapi_dir: Path, label: str) -> dict:
    """
    Always prompts for ionapi file and company/division — no 1-hour cache,
    no skip.  Ensures SOURCE and DEST are always configured independently.
    """
    print(f"\n{'─' * 60}")
    print(f"  📡  Configure {label} tenant")
    print(f"{'─' * 60}")

    CONFIG["tenant"] = _select_ionapi_forced(ionapi_dir, label)
    _prompt_company_division_forced(label)
    get_ion_token()

    snap = snapshot_config()
    print(f"\n  ✅  {label} ready: {Path(CONFIG['tenant']).name}  "
          f"CONO={CONFIG.get('company', '')}  DIVI={CONFIG.get('division', '')}")
    return snap


# =============================================================================
# CRS949MI.LstPartnerMedia  →  list of dicts
# =============================================================================

def fetch_partner_media(session: requests.Session, divi: str, label: str) -> list[dict]:
    """
    Calls CRS949MI.LstPartnerMedia for the given division and returns a list of
    dicts, one per partner-media row, keyed by the column names in
    PARTNER_MEDIA_COLS.  Handles 401 token refresh automatically.
    """
    list_url = CONFIG["api_url"]

    payload = {
        "program": "CRS949MI",
        "transactions": [{
            "transaction": "LstPartnerMedia",
            "record": {"DIVI": divi},
            "selectedColumns": PARTNER_MEDIA_COLS,
        }],
    }

    headers = {
        "Authorization": f"Bearer {CONFIG['access_token']}",
        "Content-Type": "application/json",
    }

    response = session.post(list_url, json=payload, headers=headers)
    if response.status_code == 401:
        get_ion_token()
        headers["Authorization"] = f"Bearer {CONFIG['access_token']}"
        response = session.post(list_url, json=payload, headers=headers)

    response.raise_for_status()
    data = response.json()

    rows: list[dict] = []
    for result in data.get("results", []):
        for record in result.get("records", []):
            row = {col: record.get(col, "").strip() for col in PARTNER_MEDIA_COLS}
            rows.append(row)

    if not rows:
        print(f"  ⚠️   No partner-media records returned ({label}).")
    else:
        print(f"  ✅  {len(rows)} partner-media records retrieved ({label}).")

    return rows


# =============================================================================
# Diff logic
# =============================================================================

def _row_key(row: dict) -> tuple:
    return tuple(row.get(f, "") for f in KEY_FIELDS)


def diff_partner_media(
    source_rows: list[dict],
    dest_rows: list[dict],
) -> tuple[list[dict], list[dict]]:
    """
    Compare SOURCE and DEST partner-media lists.

    Returns:
        to_update  – rows whose key exists in DEST but whose value fields differ
        to_add     – rows whose key is entirely absent from DEST
    """
    dest_index: dict[tuple, dict] = {_row_key(r): r for r in dest_rows}

    to_update: list[dict] = []
    to_add:    list[dict] = []

    for src in source_rows:
        key = _row_key(src)
        dest = dest_index.get(key)
        if dest is None:
            to_add.append(src)
        else:
            if any(src.get(f, "") != dest.get(f, "") for f in VALUE_FIELDS):
                to_update.append(src)

    return to_update, to_add


# =============================================================================
# Build API records
# =============================================================================

def build_dlt_record(dest_row: dict, dest_divi: str) -> dict:
    """Build the record for CRS949MI.DltPartnerMedia using DEST values + DEST DIVI."""
    rec = {col: dest_row.get(col, "") for col in PARTNER_MEDIA_COLS}
    rec["DIVI"] = dest_divi
    return {k: v for k, v in rec.items() if v != ""}


def build_add_record(src_row: dict, dest_divi: str) -> dict:
    """Build the record for CRS949MI.Add* using SOURCE values + DEST DIVI."""
    rec = {col: src_row.get(col, "") for col in PARTNER_MEDIA_COLS}
    rec["DIVI"] = dest_divi
    return {k: v for k, v in rec.items() if v != ""}


# =============================================================================
# CRS949MI.DltPartnerMedia on DEST (batched)
# =============================================================================

def run_dlt_batched(
    records: list[dict],
    session: requests.Session,
    batch_size: int,
) -> tuple[list[dict], list[tuple[dict, str]]]:
    """
    Deletes existing DEST records (before re-adding updated SOURCE values).

    Returns:
        successes : list of record dicts that were accepted
        failures  : list of (record dict, error message) for rejected records
    """
    successes: list[dict] = []
    failures:  list[tuple[dict, str]] = []

    for i in tqdm(
        range(0, len(records), batch_size),
        desc="CRS949MI.DltPartnerMedia",
        unit="batch",
        leave=False,
    ):
        batch = records[i : i + batch_size]
        payload = {
            "program": "CRS949MI",
            "transactions": [
                {
                    "transaction": "DltPartnerMedia",
                    "record": rec,
                    "selectedColumns": PARTNER_MEDIA_COLS,
                }
                for rec in batch
            ],
        }
        try:
            result = post_to_m3(payload, session)
            api_results = result.get("results", [])
            for j, res in enumerate(api_results):
                err = res.get("errorMessage", "").strip() if isinstance(res, dict) else ""
                rec = batch[j] if j < len(batch) else {}
                if err:
                    failures.append((rec, err))
                else:
                    successes.append(rec)
            for k in range(len(api_results), len(batch)):
                failures.append((batch[k], "No result returned"))
        except Exception as exc:
            for rec in batch:
                failures.append((rec, str(exc)))

    return successes, failures


# =============================================================================
# CRS949MI.Add* on DEST (batched, MEDC-routed)
# =============================================================================

def run_add_batched(
    records: list[dict],
    session: requests.Session,
    batch_size: int,
    desc: str = "add",
) -> tuple[list[dict], list[tuple[dict, str]], list[dict]]:
    """
    Adds records to DEST, routing each record to the correct transaction based
    on its MEDC value:
        MAIL     → CRS949MI.AddPartnerEmail
        MBMEVENT → CRS949MI.AddPartnerMBM
        (other)  → skipped (not sent to the API)

    Returns:
        successes : list of accepted record dicts
        failures  : list of (record dict, error message)
        skipped   : list of record dicts whose MEDC is not in MEDC_TRANSACTION
    """
    # Partition records by resolved transaction
    groups:  dict[str, list[dict]] = {}
    skipped: list[dict] = []

    for rec in records:
        txn = resolve_add_transaction(rec)
        if txn is None:
            skipped.append(rec)
        else:
            groups.setdefault(txn, []).append(rec)

    if skipped:
        unrecognised = sorted({r.get("MEDC", "") for r in skipped})
        print(f"  ⚠️   {len(skipped)} record(s) skipped — MEDC not in MEDC_TRANSACTION: {unrecognised}")

    successes: list[dict] = []
    failures:  list[tuple[dict, str]] = []

    for txn, grp in groups.items():
        for i in tqdm(
            range(0, len(grp), batch_size),
            desc=f"CRS949MI.{txn} [{desc}]",
            unit="batch",
            leave=False,
        ):
            batch = grp[i : i + batch_size]
            payload = {
                "program": "CRS949MI",
                "transactions": [
                    {
                        "transaction": txn,
                        "record": rec,
                        "selectedColumns": PARTNER_MEDIA_COLS,
                    }
                    for rec in batch
                ],
            }
            try:
                result = post_to_m3(payload, session)
                api_results = result.get("results", [])
                for j, res in enumerate(api_results):
                    err = res.get("errorMessage", "").strip() if isinstance(res, dict) else ""
                    rec = batch[j] if j < len(batch) else {}
                    if err:
                        failures.append((rec, err))
                    else:
                        successes.append(rec)
                for k in range(len(api_results), len(batch)):
                    failures.append((batch[k], "No result returned"))
            except Exception as exc:
                for rec in batch:
                    failures.append((rec, str(exc)))

    return successes, failures, skipped


# =============================================================================
# Console reporting
# =============================================================================

def print_summary(
    label: str,
    successes: list[dict],
    failures: list[tuple[dict, str]],
) -> None:
    from collections import Counter
    total = len(successes) + len(failures)
    print(f"\n{'═' * 60}")
    print(f"  {label}")
    print(f"  ✅ Succeeded : {len(successes)}")
    print(f"  ❌ Failed    : {len(failures)}")
    print(f"  📋 Total     : {total}")
    print(f"{'═' * 60}")

    msg_counts: Counter = Counter()
    msg_counts["OK"] = len(successes)
    for _, err in failures:
        msg_counts[err] += 1
    print()
    for msg, count in msg_counts.most_common():
        print(f"  {msg} = {count}")


# =============================================================================
# XLSX summary (always written, not just on errors)
# =============================================================================

def export_summary_xlsx(
    source_count: int,
    dest_count: int,
    dlt_successes:     list[dict],
    dlt_failures:      list[tuple[dict, str]],
    upd_add_successes: list[dict],
    upd_add_failures:  list[tuple[dict, str]],
    upd_skipped:       list[dict],
    add_successes:     list[dict],
    add_failures:      list[tuple[dict, str]],
    add_skipped:       list[dict],
    out_dir: Path,
) -> Path:
    """
    Always writes an xlsx summarising all API calls:
      Sheet 1 – Summary            : row per step with success / failed / skipped counts
      Sheet 2 – DltPartnerMedia    : detail rows for every Dlt result
      Sheet 3 – Add (Update)       : detail rows for re-adds after Dlt (all MEDC-routed txns)
      Sheet 4 – Add (New)          : detail rows for net-new adds
      Sheet 5 – Skipped            : records excluded due to unsupported MEDC (if any)
    """
    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"MIG_SyncPartnerMedia_Summary_{ts}.xlsx"

    wb = Workbook()

    hdr_font   = Font(name="Arial", bold=True, color="FFFFFF")
    hdr_fill   = PatternFill("solid", start_color="2F5496")
    ok_fill    = PatternFill("solid", start_color="E2EFDA")
    err_fill   = PatternFill("solid", start_color="FCE4D6")
    skip_fill  = PatternFill("solid", start_color="FFF2CC")
    info_fill  = PatternFill("solid", start_color="DDEBF7")
    ctr        = Alignment(horizontal="center")
    arial      = Font(name="Arial")

    # ── helpers ────────────────────────────────────────────────────────────

    def _hdr(ws, col: int, row: int, val: str):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = ctr
        return c

    def _cell(ws, col: int, row: int, val, fill=None, align=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = arial
        if fill:  c.fill      = fill
        if align: c.alignment = align
        return c

    def _write_detail_sheet(
        ws,
        ok_list:  list[dict],
        failures: list[tuple[dict, str]],
        cols:     list[str],
        extra_col: str | None = None,   # e.g. "TRANSACTION"
        extra_fn=None,                  # callable(rec) → value for extra_col
    ):
        """Write a detail sheet combining successes + failures."""
        full_cols = cols + ([extra_col] if extra_col else []) + ["STATUS"]

        for ci, f in enumerate(full_cols, start=1):
            _hdr(ws, ci, 1, f)

        all_rows = [(rec, "OK")  for rec in ok_list] + \
                   [(rec, err)   for rec, err in failures]

        for ri, (rec, status) in enumerate(all_rows, start=2):
            fill = ok_fill if status == "OK" else err_fill
            for ci, f in enumerate(full_cols, start=1):
                if f == "STATUS":
                    val = status
                elif f == extra_col and extra_fn:
                    val = extra_fn(rec)
                else:
                    val = rec.get(f, "")
                _cell(ws, ci, ri, val, fill=fill)

        for ci, f in enumerate(full_cols, start=1):
            sample = [
                (status if f == "STATUS"
                 else extra_fn(rec) if (f == extra_col and extra_fn)
                 else rec.get(f, ""))
                for rec, status in all_rows
            ]
            max_w = max(len(str(v)) for v in [f] + sample) if all_rows else len(f)
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_w + 2, 60)

    def _write_skipped_sheet(ws, skipped: list[dict]):
        full_cols = PARTNER_MEDIA_COLS + ["REASON"]
        for ci, f in enumerate(full_cols, start=1):
            _hdr(ws, ci, 1, f)
        for ri, rec in enumerate(skipped, start=2):
            for ci, f in enumerate(full_cols, start=1):
                val = f"MEDC '{rec.get('MEDC','')}' not in MEDC_TRANSACTION" if f == "REASON" else rec.get(f, "")
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = arial
                c.fill = skip_fill
        for ci, f in enumerate(full_cols, start=1):
            sample = [
                (f"MEDC '{r.get('MEDC','')}' not in MEDC_TRANSACTION" if f == "REASON" else r.get(f, ""))
                for r in skipped
            ]
            max_w = max(len(str(v)) for v in [f] + sample) if skipped else len(f)
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_w + 2, 60)

    # ── Sheet 1: Summary ───────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    for ci, heading in enumerate(
        ["Step", "Transaction(s)", "Success", "Failed", "Skipped", "Total"], start=1
    ):
        _hdr(ws_sum, ci, 1, heading)

    all_skipped = len(upd_skipped) + len(add_skipped)
    steps = [
        ("1 – Extract SOURCE",   "LstPartnerMedia",              source_count,          0,                   0),
        ("2 – Extract DEST",     "LstPartnerMedia",              dest_count,             0,                   0),
        ("3a – Delete (update)", "DltPartnerMedia",              len(dlt_successes),     len(dlt_failures),   0),
        ("3b – Re-add (update)", "AddPartnerEmail / AddPartnerMBM", len(upd_add_successes), len(upd_add_failures), len(upd_skipped)),
        ("4  – Add (new)",       "AddPartnerEmail / AddPartnerMBM", len(add_successes),     len(add_failures),    len(add_skipped)),
    ]

    for ri, (step_label, txn, ok, fail, skip) in enumerate(steps, start=2):
        fill = err_fill if fail > 0 else (skip_fill if skip > 0 else info_fill)
        _cell(ws_sum, 1, ri, step_label,    fill=fill)
        _cell(ws_sum, 2, ri, txn,           fill=fill)
        _cell(ws_sum, 3, ri, ok,            fill=fill, align=ctr)
        _cell(ws_sum, 4, ri, fail,          fill=fill, align=ctr)
        _cell(ws_sum, 5, ri, skip,          fill=fill, align=ctr)
        _cell(ws_sum, 6, ri, ok+fail+skip,  fill=fill, align=ctr)

    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 38
    ws_sum.column_dimensions["C"].width = 12
    ws_sum.column_dimensions["D"].width = 12
    ws_sum.column_dimensions["E"].width = 12
    ws_sum.column_dimensions["F"].width = 12

    # ── Sheet 2: DltPartnerMedia detail ────────────────────────────────────
    if dlt_successes or dlt_failures:
        ws_dlt = wb.create_sheet("DltPartnerMedia")
        _write_detail_sheet(ws_dlt, dlt_successes, dlt_failures, PARTNER_MEDIA_COLS)

    # ── Sheet 3: Add (Update) detail ──────────────────────────────────────
    if upd_add_successes or upd_add_failures:
        ws_upd = wb.create_sheet("Add (Update)")
        _write_detail_sheet(
            ws_upd, upd_add_successes, upd_add_failures, PARTNER_MEDIA_COLS,
            extra_col="TRANSACTION",
            extra_fn=lambda r: resolve_add_transaction(r) or "—",
        )

    # ── Sheet 4: Add (New) detail ──────────────────────────────────────────
    if add_successes or add_failures:
        ws_add = wb.create_sheet("Add (New)")
        _write_detail_sheet(
            ws_add, add_successes, add_failures, PARTNER_MEDIA_COLS,
            extra_col="TRANSACTION",
            extra_fn=lambda r: resolve_add_transaction(r) or "—",
        )

    # ── Sheet 5: Skipped detail (unsupported MEDC) ─────────────────────────
    all_skipped_recs = upd_skipped + add_skipped
    if all_skipped_recs:
        ws_skip = wb.create_sheet("Skipped")
        _write_skipped_sheet(ws_skip, all_skipped_recs)

    wb.save(out_path)
    return out_path


# =============================================================================
# Main
# =============================================================================

def sync_partner_media() -> None:
    ionapi_dir = Path(__file__).parent / "ionapi"

    # Batch size fixed at 100 — no prompt
    batch_size = BATCH_SIZE
    print(f"  ℹ️   Batch size fixed at {batch_size}.")

    # ------------------------------------------------------------------ #
    # Step 1 – SOURCE: CRS949MI.LstPartnerMedia                          #
    # ------------------------------------------------------------------ #
    source_snap = setup_tenant(ionapi_dir, "SOURCE")
    source_divi = CONFIG.get("division", "")

    print(f"\n🔍  Step 1 – CRS949MI.LstPartnerMedia (SOURCE DIVI={source_divi}) …")
    with requests.Session() as session:
        source_rows = fetch_partner_media(session, source_divi, label="SOURCE")

    if not source_rows:
        print("⚠️   No partner media records found in SOURCE. Exiting.")
        return

    # ------------------------------------------------------------------ #
    # Step 2 – DEST: CRS949MI.LstPartnerMedia                            #
    # ------------------------------------------------------------------ #
    dest_snap = setup_tenant(ionapi_dir, "DEST")
    dest_divi = CONFIG.get("division", "")

    print(f"\n🔍  Step 2 – CRS949MI.LstPartnerMedia (DEST DIVI={dest_divi}) …")
    with requests.Session() as session:
        dest_rows = fetch_partner_media(session, dest_divi, label="DEST")

    # ------------------------------------------------------------------ #
    # Step 3 – Diff                                                        #
    # ------------------------------------------------------------------ #
    print("\n🔗  Step 3 – Comparing SOURCE vs DEST …")
    to_update, to_add = diff_partner_media(source_rows, dest_rows)

    if not to_update and not to_add:
        print("\n✅  DEST partner media all match SOURCE. Nothing to do.")
        # Still write summary xlsx
        xlsx_path = export_summary_xlsx(
            len(source_rows), len(dest_rows), [], [], [], [], [], [], [], [],
            Path(__file__).parent,
        )
        print(f"\n  📄  Summary saved to: {xlsx_path.name}")
        return

    # ── Filter by supported MEDC ───────────────────────────────────────────
    # Records whose MEDC is not in MEDC_TRANSACTION are skipped entirely:
    # they are neither deleted nor added in DEST.
    dest_index: dict[tuple, dict] = {_row_key(r): r for r in dest_rows}

    to_update_ok   = [r for r in to_update if resolve_add_transaction(r) is not None]
    to_update_skip = [r for r in to_update if resolve_add_transaction(r) is None]
    to_add_ok      = [r for r in to_add    if resolve_add_transaction(r) is not None]
    to_add_skip    = [r for r in to_add    if resolve_add_transaction(r) is None]

    if to_update_skip or to_add_skip:
        skip_medcs = sorted({r.get("MEDC", "") for r in to_update_skip + to_add_skip})
        print(f"  ⚠️   {len(to_update_skip) + len(to_add_skip)} record(s) will be skipped "
              f"(unsupported MEDC: {skip_medcs})")

    dlt_records     = [build_dlt_record(dest_index[_row_key(r)], dest_divi) for r in to_update_ok]
    upd_add_records = [build_add_record(r, dest_divi) for r in to_update_ok]
    add_records     = [build_add_record(r, dest_divi) for r in to_add_ok]

    all_diff = to_update + to_add   # includes skipped — for display count

    # ── Preview (first 10) ─────────────────────────────────────────────────
    dest_index_preview: dict[tuple, dict] = {_row_key(r): r for r in dest_rows}

    def _changes_summary(src_row: dict) -> str:
        dest_row = dest_index_preview.get(_row_key(src_row), {})
        parts = []
        for f in VALUE_FIELDS:
            sv, dv = src_row.get(f, ""), dest_row.get(f, "")
            if sv != dv:
                parts.append(f'{f}: "{dv}"→"{sv}"')
        return " | ".join(parts)

    display_cols = KEY_FIELDS + ["ACTION", "TXN", "CHANGES"]

    def _annotate(row: dict, base_action: str, changes: str) -> dict:
        txn = resolve_add_transaction(row)
        return {
            **row,
            "ACTION":  "SKIP" if txn is None else base_action,
            "TXN":     "—"    if txn is None else txn,
            "CHANGES": changes,
        }

    preview_annotated = (
        [_annotate(row, "UPDATE", _changes_summary(row)) for row in to_update]
        + [_annotate(row, "ADD",  "")                    for row in to_add]
    )
    preview_rows = preview_annotated[:10]

    col_widths = {
        f: max(len(f), max((len(r.get(f, "")) for r in preview_rows), default=0))
        for f in display_cols
    }
    header_line = "  " + "  ".join(f.ljust(col_widths[f]) for f in display_cols)
    sep_line    = "  " + "  ".join("─" * col_widths[f] for f in display_cols)

    print(f"\n{'═' * max(len(header_line), 60)}")
    print(f"  📋  Plan: {len(to_update_ok)} update(s)  +  {len(to_add_ok)} add(s)"
          f"  +  {len(to_update_skip) + len(to_add_skip)} skip(s)  →  DEST")
    print(f"{'═' * max(len(header_line), 60)}")
    print(header_line)
    print(sep_line)
    for rec in preview_rows:
        print("  " + "  ".join(rec.get(f, "").ljust(col_widths[f]) for f in display_cols))
    if len(all_diff) > 10:
        print(f"  … and {len(all_diff) - 10} more record(s)")
    print(f"{'═' * max(len(header_line), 60)}")

    confirm = input("  Proceed with updating DEST? [y/N]: ").strip().lower()
    if confirm != "y":
        print("⛔  Aborted — no changes made to DEST.")
        return
    print()

    # ------------------------------------------------------------------ #
    # Step 4a – DEST: CRS949MI.DltPartnerMedia (records to update)       #
    # ------------------------------------------------------------------ #
    restore_config(dest_snap)
    get_ion_token()

    dlt_successes:     list[dict] = []
    dlt_failures:      list[tuple[dict, str]] = []
    upd_add_successes: list[dict] = []
    upd_add_failures:  list[tuple[dict, str]] = []
    upd_add_skipped:   list[dict] = []
    add_successes:     list[dict] = []
    add_failures:      list[tuple[dict, str]] = []
    add_skipped:       list[dict] = []

    with requests.Session() as session:

        # Step 4a: Delete existing DEST records that differ (MEDC-supported only)
        if dlt_records:
            print(f"▶   Step 4a – CRS949MI.DltPartnerMedia ({len(dlt_records)} records) …")
            dlt_successes, dlt_failures = run_dlt_batched(dlt_records, session, batch_size)
        else:
            print("  ℹ️   No records to delete (DltPartnerMedia skipped).")

        # ------------------------------------------------------------------ #
        # Step 4b – DEST: Add re-add updated records (MEDC-routed)           #
        # ------------------------------------------------------------------ #
        # Only re-add the ones whose Dlt succeeded
        if dlt_successes and upd_add_records:
            dlt_success_keys = {_row_key(r) for r in dlt_successes}
            # upd_add_records and to_update_ok are aligned (same index)
            records_to_readd = [
                rec for rec, src in zip(upd_add_records, to_update_ok)
                if _row_key(build_dlt_record(dest_index[_row_key(src)], dest_divi)) in dlt_success_keys
            ]
            if records_to_readd:
                print(f"▶   Step 4b – Add (update) ({len(records_to_readd)} records, MEDC-routed) …")
                upd_add_successes, upd_add_failures, upd_add_skipped = run_add_batched(
                    records_to_readd, session, batch_size, desc="update",
                )
            else:
                print("  ℹ️   No update re-adds to perform.")
        elif upd_add_records and not dlt_records:
            print("  ℹ️   No Dlt records processed; skipping update re-add.")

        if dlt_failures:
            print(f"  ⚠️   {len(dlt_failures)} Dlt failure(s) — corresponding Add skipped for those records.")

        # ------------------------------------------------------------------ #
        # Step 4c – DEST: Add net-new records (MEDC-routed)                  #
        # ------------------------------------------------------------------ #
        if add_records:
            print(f"▶   Step 4c – Add (new) ({len(add_records)} records, MEDC-routed) …")
            add_successes, add_failures, add_skipped = run_add_batched(
                add_records, session, batch_size, desc="new",
            )
        else:
            print("  ℹ️   No records to add (Add skipped).")

    # ------------------------------------------------------------------ #
    # Results                                                              #
    # ------------------------------------------------------------------ #
    if dlt_records:
        print_summary("DltPartnerMedia [DEST]", dlt_successes, dlt_failures)
    if upd_add_records:
        print_summary("Add / update [DEST]", upd_add_successes, upd_add_failures)
        if upd_add_skipped:
            print(f"  ⏭️   Skipped (unsupported MEDC): {len(upd_add_skipped)}")
    if add_records:
        print_summary("Add / new [DEST]", add_successes, add_failures)
        if add_skipped:
            print(f"  ⏭️   Skipped (unsupported MEDC): {len(add_skipped)}")

    xlsx_path = export_summary_xlsx(
        len(source_rows),
        len(dest_rows),
        dlt_successes,     dlt_failures,
        upd_add_successes, upd_add_failures, upd_add_skipped,
        add_successes,     add_failures,     add_skipped,
        Path(__file__).parent,
    )
    print(f"\n  📄  Summary saved to: {xlsx_path.name}")


# =============================================================================
# Entry Point
# =============================================================================

if __name__ == "__main__":
    sync_partner_media()
