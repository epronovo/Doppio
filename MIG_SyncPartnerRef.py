# MIG_SyncPartnerRef.py
"""
MIG_SyncPartnerRef.py
---------------------
Syncs partner reference data (CRS945) from a SOURCE tenant to a
DESTINATION tenant via CRS945MI.UpdPartnerRef and CRS945MI.AddPartnerRef.

Steps
-----
1.  Ask for SOURCE tenant → CRS945MI.LstPartnerRef (SOURCE DIVI) → records in memory.
2.  Ask for DEST   tenant → CRS945MI.LstPartnerRef (DEST   DIVI) → records in memory.
3.  Diff: find records where PRF1/PRF2/TX15/TX40 differ (matched on DONR+DOVA+MEPF),
    and records missing from DEST entirely.  Display first 10 and pause for confirmation.
4.  Re-authenticate to DEST and:
      a. CRS945MI.UpdPartnerRef – update records that exist but differ.
      b. CRS945MI.AddPartnerRef – add records that are missing from DEST.
    Summarise results.

Usage:
    python MIG_SyncPartnerRef.py
"""

import datetime
import os
import requests
from pathlib import Path
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from InforMI import (
    CONFIG,
    get_ion_token,
    post_to_m3,
)
from UserDefaults import load_user_defaults, save_user_defaults

# =============================================================================
# Constants
# =============================================================================

# Fields returned by CRS945MI.LstPartnerRef and sent to Upd/Add
PARTNER_REF_COLS = ["DIVI", "DONR", "DOVA", "MEPF", "PRF1", "PRF2", "TX15", "TX40"]

# Composite key used to match SOURCE vs DEST rows
KEY_FIELDS   = ["DIVI", "DONR", "DOVA", "MEPF", "PRF1", "PRF2"]
# Fields compared for differences (non-key, non-DIVI)
VALUE_FIELDS = ["TX15", "TX40"]

DEFAULT_BATCH_SIZE = 100

# =============================================================================
# Tenant helpers  (mirrors MIG_SyncPanelViews pattern)
# =============================================================================

def snapshot_config() -> dict:
    return dict(CONFIG)


def restore_config(snapshot: dict) -> None:
    CONFIG.clear()
    CONFIG.update(snapshot)


def _select_ionapi_forced(ionapi_dir: Path, label: str) -> Path:
    """
    Always shows the ionapi file list and requires a selection — bypasses
    the 1-hour cache in InforMI.select_ionapi_file so SOURCE and DEST can
    use different files in the same run.
    """
    files = sorted(
        [f for f in os.listdir(ionapi_dir) if f.endswith(".ionapi")],
        key=str.lower,
    )
    if not files:
        raise FileNotFoundError(f"No .ionapi files found in: {ionapi_dir}")

    print(f"\n  🔹  Select the {label} ION API file:")
    for i, f in enumerate(files, start=1):
        print(f"    {i}. {f}")

    while True:
        try:
            choice = int(input(f"\n  Enter choice (1-{len(files)}): "))
            if 1 <= choice <= len(files):
                return ionapi_dir / files[choice - 1]
            print("  ❌  Invalid choice. Try again.")
        except ValueError:
            print("  ❌  Please enter a number.")


def _prompt_company_division_forced(label: str) -> None:
    """
    Always prompts for company and division — bypasses the 1-hour cache in
    InforMI.prompt_for_company_division so each tenant is configured fresh.
    """
    defaults = load_user_defaults()

    default_company  = defaults.get("company",  "100")
    default_division = defaults.get("division", "500")

    print(f"\n  Configure {label} company / division:")
    company  = input(f"    Company  (default: {default_company}):  ").strip()
    division = input(f"    Division (default: {default_division}): ").strip()

    CONFIG["company"]  = company  if company  else default_company
    CONFIG["division"] = division if division else default_division

    # Persist as new defaults (but do NOT update last_prompt_time so the
    # next call is also forced to prompt)
    defaults["company"]  = CONFIG["company"]
    defaults["division"] = CONFIG["division"]
    save_user_defaults(defaults)


def setup_tenant(ionapi_dir: Path, label: str) -> dict:
    """
    Always prompts for ionapi file and company/division — no 1-hour cache,
    no skip.  Ensures SOURCE and DEST are always configured independently.
    """
    print(f"\n{'─' * 60}")
    print(f"  📡  Configure {label} tenant")
    print(f"{'─' * 60}")

    CONFIG["tenant"] = _select_ionapi_forced(ionapi_dir, label)
    _prompt_company_division_forced(label)
    get_ion_token()

    snap = snapshot_config()
    print(f"\n  ✅  {label} ready: {Path(CONFIG['tenant']).name}  "
          f"CONO={CONFIG.get('company', '')}  DIVI={CONFIG.get('division', '')}")
    return snap


# =============================================================================
# CRS945MI.LstPartnerRef  →  list of dicts
# =============================================================================

def fetch_partner_refs(session: requests.Session, divi: str, label: str) -> list[dict]:
    """
    Calls CRS945MI.LstPartnerRef for the given division and returns a list of
    dicts, one per partner-reference row, keyed by the column names in
    PARTNER_REF_COLS.  Handles 401 token refresh automatically.
    """
    list_url = CONFIG["api_url"]

    payload = {
        "program": "CRS945MI",
        "transactions": [{
            "transaction": "LstPartnerRef",
            "record": {"DIVI": divi},
            "selectedColumns": PARTNER_REF_COLS,
        }],
    }

    headers = {
        "Authorization": f"Bearer {CONFIG['access_token']}",
        "Content-Type": "application/json",
    }

    response = session.post(list_url, json=payload, headers=headers)
    if response.status_code == 401:
        get_ion_token()
        headers["Authorization"] = f"Bearer {CONFIG['access_token']}"
        response = session.post(list_url, json=payload, headers=headers)

    response.raise_for_status()
    data = response.json()

    rows: list[dict] = []
    for result in data.get("results", []):
        for record in result.get("records", []):
            row = {col: record.get(col, "").strip() for col in PARTNER_REF_COLS}
            rows.append(row)

    if not rows:
        print(f"  ⚠️   No partner-reference records returned ({label}).")
    else:
        print(f"  ✅  {len(rows)} partner-reference records retrieved ({label}).")

    return rows


# =============================================================================
# Diff logic
# =============================================================================

def _row_key(row: dict) -> tuple:
    return tuple(row.get(f, "") for f in KEY_FIELDS)


def diff_partner_refs(
    source_rows: list[dict],
    dest_rows: list[dict],
) -> tuple[list[dict], list[dict]]:
    """
    Compare SOURCE and DEST partner-ref lists.

    Returns:
        to_update  – rows whose key exists in DEST but whose value fields differ
        to_add     – rows whose key is entirely absent from DEST
    """
    dest_index: dict[tuple, dict] = {_row_key(r): r for r in dest_rows}

    to_update: list[dict] = []
    to_add:    list[dict] = []

    for src in source_rows:
        key = _row_key(src)
        dest = dest_index.get(key)
        if dest is None:
            to_add.append(src)
        else:
            if any(src.get(f, "") != dest.get(f, "") for f in VALUE_FIELDS):
                to_update.append(src)

    return to_update, to_add


# =============================================================================
# Build API records
# =============================================================================

def build_upd_record(src_row: dict, dest_divi: str) -> dict:
    """Build the record for CRS945MI.UpdPartnerRef using SOURCE values + DEST DIVI."""
    rec = {col: src_row.get(col, "") for col in PARTNER_REF_COLS}
    rec["DIVI"] = dest_divi
    # Drop blank fields
    return {k: v for k, v in rec.items() if v != ""}


def build_add_record(src_row: dict, dest_divi: str) -> dict:
    """Build the record for CRS945MI.AddPartnerRef using SOURCE values + DEST DIVI."""
    rec = {col: src_row.get(col, "") for col in PARTNER_REF_COLS}
    rec["DIVI"] = dest_divi
    # Drop blank fields
    return {k: v for k, v in rec.items() if v != ""}


# =============================================================================
# CRS945MI.UpdPartnerRef on DEST (batched)
# =============================================================================

def run_upd_batched(
    records: list[dict],
    session: requests.Session,
    batch_size: int,
) -> tuple[list[dict], list[tuple[dict, str]]]:
    """
    Returns:
        successes : list of record dicts that were accepted
        failures  : list of (record dict, error message) for rejected records
    """
    successes: list[dict] = []
    failures: list[tuple[dict, str]] = []

    for i in tqdm(
        range(0, len(records), batch_size),
        desc="CRS945MI.UpdPartnerRef",
        unit="batch",
        leave=False,
    ):
        batch = records[i : i + batch_size]
        payload = {
            "program": "CRS945MI",
            "transactions": [
                {
                    "transaction": "UpdPartnerRef",
                    "record": rec,
                    "selectedColumns": PARTNER_REF_COLS,
                }
                for rec in batch
            ],
        }
        try:
            result = post_to_m3(payload, session)
            api_results = result.get("results", [])
            for j, res in enumerate(api_results):
                err = res.get("errorMessage", "").strip() if isinstance(res, dict) else ""
                rec = batch[j] if j < len(batch) else {}
                if err:
                    failures.append((rec, err))
                else:
                    successes.append(rec)
            for k in range(len(api_results), len(batch)):
                failures.append((batch[k], "No result returned"))
        except Exception as exc:
            for rec in batch:
                failures.append((rec, str(exc)))

    return successes, failures


# =============================================================================
# CRS945MI.AddPartnerRef on DEST (batched)
# =============================================================================

def run_add_batched(
    records: list[dict],
    session: requests.Session,
    batch_size: int,
) -> tuple[list[dict], list[tuple[dict, str]]]:
    """
    Returns:
        successes : list of record dicts that were accepted
        failures  : list of (record dict, error message) for rejected records
    """
    successes: list[dict] = []
    failures: list[tuple[dict, str]] = []

    for i in tqdm(
        range(0, len(records), batch_size),
        desc="CRS945MI.AddPartnerRef",
        unit="batch",
        leave=False,
    ):
        batch = records[i : i + batch_size]
        payload = {
            "program": "CRS945MI",
            "transactions": [
                {
                    "transaction": "AddPartnerRef",
                    "record": rec,
                    "selectedColumns": PARTNER_REF_COLS,
                }
                for rec in batch
            ],
        }
        try:
            result = post_to_m3(payload, session)
            api_results = result.get("results", [])
            for j, res in enumerate(api_results):
                err = res.get("errorMessage", "").strip() if isinstance(res, dict) else ""
                rec = batch[j] if j < len(batch) else {}
                if err:
                    failures.append((rec, err))
                else:
                    successes.append(rec)
            for k in range(len(api_results), len(batch)):
                failures.append((batch[k], "No result returned"))
        except Exception as exc:
            for rec in batch:
                failures.append((rec, str(exc)))

    return successes, failures


# =============================================================================
# Reporting
# =============================================================================

def print_summary(
    label: str,
    successes: list[dict],
    failures: list[tuple[dict, str]],
) -> None:
    from collections import Counter
    total = len(successes) + len(failures)
    print(f"\n{'═' * 60}")
    print(f"  {label}")
    print(f"  ✅ Succeeded : {len(successes)}")
    print(f"  ❌ Failed    : {len(failures)}")
    print(f"  📋 Total     : {total}")
    print(f"{'═' * 60}")

    msg_counts: Counter = Counter()
    msg_counts["OK"] = len(successes)
    for _, err in failures:
        msg_counts[err] += 1
    print()
    for msg, count in msg_counts.most_common():
        print(f"  {msg} = {count}")


def export_errors_xlsx(
    upd_successes: list[dict],
    upd_failures: list[tuple[dict, str]],
    add_successes: list[dict],
    add_failures: list[tuple[dict, str]],
    out_dir: Path,
) -> Path:
    """
    Writes three sheets to an xlsx file:
      • Summary          – error-message totals for both steps
      • UpdPartnerRef Err – one row per failed UpdPartnerRef record
      • AddPartnerRef Err – one row per failed AddPartnerRef record
    Returns the path of the written file.
    """
    from collections import Counter

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"MIG_SyncPartnerRef_Errors_{ts}.xlsx"

    wb = Workbook()

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", start_color="2F5496")
    ok_fill  = PatternFill("solid", start_color="E2EFDA")
    err_fill = PatternFill("solid", start_color="FCE4D6")
    ctr      = Alignment(horizontal="center")

    def _write_error_sheet(ws, fields, failures):
        for col, f in enumerate(fields, start=1):
            c = ws.cell(row=1, column=col, value=f)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = ctr
        for row_i, (rec, err) in enumerate(failures, start=2):
            for col, f in enumerate(fields, start=1):
                val = err if f == "ERROR" else rec.get(f, "")
                c = ws.cell(row=row_i, column=col, value=val)
                c.font = Font(name="Arial")
        for col, f in enumerate(fields, start=1):
            max_len = max(
                len(f),
                *(len(str(rec.get(f, "") if f != "ERROR" else err))
                  for rec, err in failures),
            )
            ws.column_dimensions[
                ws.cell(row=1, column=col).column_letter
            ].width = min(max_len + 2, 60)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    for col, heading in enumerate(["Step", "Message", "Count"], start=1):
        c = ws_sum.cell(row=1, column=col, value=heading)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = ctr

    row_i = 2
    for step_label, successes, failures in [
        ("UpdPartnerRef", upd_successes, upd_failures),
        ("AddPartnerRef", add_successes, add_failures),
    ]:
        counts: Counter = Counter()
        counts["OK"] = len(successes)
        for _, err in failures:
            counts[err] += 1
        for msg, count in counts.most_common():
            fill = ok_fill if msg == "OK" else err_fill
            for col, val in enumerate([step_label, msg, count], start=1):
                c = ws_sum.cell(row=row_i, column=col, value=val)
                c.fill = fill
                c.font = Font(name="Arial")
                if col == 3:
                    c.alignment = ctr
            row_i += 1

    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 80
    ws_sum.column_dimensions["C"].width = 10

    # ── Sheet 2: UpdPartnerRef errors ──────────────────────────────────────
    if upd_failures:
        ws_upd = wb.create_sheet("UpdPartnerRef Err")
        _write_error_sheet(ws_upd, PARTNER_REF_COLS + ["ERROR"], upd_failures)

    # ── Sheet 3: AddPartnerRef errors ──────────────────────────────────────
    if add_failures:
        ws_add = wb.create_sheet("AddPartnerRef Err")
        _write_error_sheet(ws_add, PARTNER_REF_COLS + ["ERROR"], add_failures)

    wb.save(out_path)
    return out_path


# =============================================================================
# Main
# =============================================================================

def sync_partner_refs() -> None:
    ionapi_dir = Path(__file__).parent / "ionapi"

    batch_size = DEFAULT_BATCH_SIZE

    # ------------------------------------------------------------------ #
    # Step 1 – SOURCE: CRS945MI.LstPartnerRef                            #
    # ------------------------------------------------------------------ #
    save_user_defaults({"sync_batch_size": batch_size})
    source_snap = setup_tenant(ionapi_dir, "SOURCE")
    source_divi = CONFIG.get("division", "")

    print(f"\n🔍  Step 1 – CRS945MI.LstPartnerRef (SOURCE DIVI={source_divi}) …")
    with requests.Session() as session:
        source_rows = fetch_partner_refs(session, source_divi, label="SOURCE")

    if not source_rows:
        print("⚠️   No partner references found in SOURCE. Exiting.")
        return

    # ------------------------------------------------------------------ #
    # Step 2 – DEST: CRS945MI.LstPartnerRef                              #
    # ------------------------------------------------------------------ #
    save_user_defaults({"sync_batch_size": batch_size})
    dest_snap = setup_tenant(ionapi_dir, "DEST")
    dest_divi = CONFIG.get("division", "")

    print(f"\n🔍  Step 2 – CRS945MI.LstPartnerRef (DEST DIVI={dest_divi}) …")
    with requests.Session() as session:
        dest_rows = fetch_partner_refs(session, dest_divi, label="DEST")

    # ------------------------------------------------------------------ #
    # Step 3 – Diff                                                        #
    # ------------------------------------------------------------------ #
    print("\n🔗  Step 3 – Comparing SOURCE vs DEST …")
    to_update, to_add = diff_partner_refs(source_rows, dest_rows)

    if not to_update and not to_add:
        print("\n✅  DEST partner references all match SOURCE. Nothing to do.")
        return

    # Build API records (with DEST DIVI)
    upd_records = [build_upd_record(row, dest_divi) for row in to_update]
    add_records = [build_add_record(row, dest_divi) for row in to_add]

    all_diff = to_update + to_add

    # Annotate for preview — include CHANGES column for UPDATE rows
    dest_index_preview: dict[tuple, dict] = {_row_key(r): r for r in dest_rows}

    def _changes_summary(src_row: dict) -> str:
        dest_row = dest_index_preview.get(_row_key(src_row), {})
        parts = []
        for f in VALUE_FIELDS:
            sv, dv = src_row.get(f, ""), dest_row.get(f, "")
            if sv != dv:
                parts.append(f'{f}: "{dv}" → "{sv}"')
        return " | ".join(parts)

    display_cols = PARTNER_REF_COLS + ["ACTION", "CHANGES"]

    preview_annotated = [
        {**row, "ACTION": "UPDATE", "CHANGES": _changes_summary(row)} for row in to_update
    ] + [
        {**row, "ACTION": "ADD", "CHANGES": ""} for row in to_add
    ]
    preview_rows = preview_annotated[:10]

    col_widths = {
        f: max(len(f), max((len(r.get(f, "")) for r in preview_rows), default=0))
        for f in display_cols
    }
    header_line = "  " + "  ".join(f.ljust(col_widths[f]) for f in display_cols)
    sep_line    = "  " + "  ".join("─" * col_widths[f] for f in display_cols)

    print(f"\n{'═' * len(header_line)}")
    print(f"  📋  Plan: {len(to_update)} update(s)  +  {len(to_add)} add(s)  →  DEST")
    print(f"{'═' * len(header_line)}")
    print(header_line)
    print(sep_line)
    for rec in preview_rows:
        print("  " + "  ".join(rec.get(f, "").ljust(col_widths[f]) for f in display_cols))
    if len(all_diff) > 10:
        print(f"  … and {len(all_diff) - 10} more record(s)")
    print(f"{'═' * len(header_line)}")

    confirm = input("  Proceed with updating DEST? [y/N]: ").strip().lower()
    if confirm != "y":
        print("⛔  Aborted — no changes made to DEST.")
        return
    print()

    # ------------------------------------------------------------------ #
    # Step 4a – DEST: CRS945MI.UpdPartnerRef                             #
    # ------------------------------------------------------------------ #
    restore_config(dest_snap)
    get_ion_token()

    upd_successes: list[dict] = []
    upd_failures: list[tuple[dict, str]] = []

    with requests.Session() as session:
        if upd_records:
            print(f"▶   Step 4a – CRS945MI.UpdPartnerRef ({len(upd_records)} records) …")
            upd_successes, upd_failures = run_upd_batched(upd_records, session, batch_size)
        else:
            print("  ℹ️   No records to update (UpdPartnerRef skipped).")

        # ------------------------------------------------------------------ #
        # Step 4b – DEST: CRS945MI.AddPartnerRef                             #
        # ------------------------------------------------------------------ #
        if add_records:
            print(f"▶   Step 4b – CRS945MI.AddPartnerRef ({len(add_records)} records) …")
            add_successes, add_failures = run_add_batched(add_records, session, batch_size)
        else:
            print("  ℹ️   No records to add (AddPartnerRef skipped).")
            add_successes, add_failures = [], []

    # ------------------------------------------------------------------ #
    # Results                                                              #
    # ------------------------------------------------------------------ #
    if upd_records:
        print_summary("UpdPartnerRef [DEST]", upd_successes, upd_failures)
    if add_records:
        print_summary("AddPartnerRef [DEST]", add_successes, add_failures)

    if upd_failures or add_failures:
        xlsx_path = export_errors_xlsx(
            upd_successes, upd_failures,
            add_successes, add_failures,
            Path(__file__).parent,
        )
        print(f"\n  📄  Error report saved to: {xlsx_path.name}")


# =============================================================================
# Entry Point
# =============================================================================

if __name__ == "__main__":
    sync_partner_refs()
