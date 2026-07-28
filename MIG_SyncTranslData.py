# MIG_SyncTranslData.py
"""
MIG_SyncTranslData.py
---------------------
Extracts translation data (MBMTRN / MBMTRD) from a SOURCE tenant and writes
them to an EVS100-compatible Excel file for import into a DESTINATION tenant.

Steps
-----
1.  Ask for SOURCE tenant → EXPORTMI.Select (MBMTRN) → header rows.
2.                        → EXPORTMI.Select (MBMTRD) → detail rows.
3.  Join header + details on IDTR; display plan.
4.  Export to an EVS100-format Excel file
    (evs100/ToProcess/API_CRS881MI_<timestamp>.xlsx) with:
      • Control sheet listing both data sheets
      • Sheet "API_CRS881MI_AddTranslation"  – one row per unique header
      • Sheet "API_CRS881MI_AddTranslData"   – one row per detail (header + detail fields)
5.  Print summary, then optionally:
      a. Upload the file to M3 via the File Management REST API (PUT).
      b. Trigger processing via EVS100MI.ImportFile.

Usage:
    python MIG_SyncTranslData.py
"""

import datetime
import os
import requests
from pathlib import Path
import xlsxwriter
from tqdm import tqdm
from openpyxl import Workbook as OWorkbook
from openpyxl.styles import Font, PatternFill, Alignment

from InforMI import (
    CONFIG,
    get_ion_token,
    post_to_m3,
)
from UserDefaults import load_user_defaults, save_user_defaults

# =============================================================================
# Constants
# =============================================================================

EXPORTMI_SEP = "^"

DEFAULT_BATCH_SIZE = 100

# MBMTRN – Translation header
EXPORTMI_HDR_QUERY = (
    "TRIDTR,TRTRQF,TRMSTD,TRMVRS,TRBMSG,TRIBOB,TRELMP,TRELMD,TRELMC,TRMBMC from MBMTRN"
)
# MBMTRD – Translation details
EXPORTMI_DTL_QUERY = (
    "TDDIVI,TDIDTR,TDTX15,TDMVXP,TDEXTP,TDMVXD,TDMBMD,TDTX40 from MBMTRD"
)

# Mapping from MBMTRN column names → CRS881MI API field names.
HDR_FIELD_MAP: dict[str, str] = {
    "TRIDTR": "IDTR",
    "TRTRQF": "TRQF",
    "TRMSTD": "MSTD",
    "TRMVRS": "MVRS",
    "TRBMSG": "BMSG",
    "TRIBOB": "IBOB",
    "TRELMP": "ELMP",
    "TRELMD": "ELMD",
    "TRELMC": "ELMC",
    "TRMBMC": "MBMC",
}

# Mapping from MBMTRD column names → CRS881MI API field names.
# TDIDTR is the join key only; it is NOT passed to the API.
DTL_FIELD_MAP: dict[str, str] = {
    "TDDIVI": "DIVI",
    "TDTX15": "TX15",
    "TDMVXP": "MVXP",
    "TDEXTP": "EXTP",
    "TDMVXD": "MVXD",
    "TDMBMD": "MBMD",
    "TDTX40": "TX40",
}

# API field lists for the two data sheets.
# AddTranslData prepends CONO so EVS100 appends &cono=X&divi=Y to the URL.
ADD_TRANSLATION_FIELDS = list(HDR_FIELD_MAP.values())   # IDTR TRQF MSTD MVRS BMSG IBOB ELMP ELMD ELMC MBMC
ADD_TRANSL_DATA_FIELDS  = ["CONO"] + ADD_TRANSLATION_FIELDS + list(DTL_FIELD_MAP.values())

EVS100_TO_PROCESS = Path(__file__).parent / "evs100" / "ToProcess"


# =============================================================================
# Tenant helpers
# =============================================================================

def snapshot_config() -> dict:
    return dict(CONFIG)


def restore_config(snapshot: dict) -> None:
    CONFIG.clear()
    CONFIG.update(snapshot)


def _global_url() -> str:
    """
    Build the M3 API URL without &cono / &divi.
    Translation data (MBMTRN / MBMTRD) is global — not scoped to a company or
    division — so omitting those params ensures all records are returned.
    """
    return (
        f"{CONFIG['iu']}/{CONFIG['ti']}/M3/m3api-rest/v2/execute"
        f"?maxrecs=0&extendedresult=true&righttrim=true"
    )


def _apply_global_url() -> None:
    """Override CONFIG['api_url'] with the global (no cono/divi) URL."""
    CONFIG["api_url"] = _global_url()


def _get_ion_token_global() -> None:
    """Refresh the ION token then immediately re-apply the global URL."""
    get_ion_token()
    _apply_global_url()


def _select_ionapi_forced(ionapi_dir: Path, label: str) -> Path:
    files = sorted(
        [f for f in os.listdir(ionapi_dir) if f.endswith(".ionapi")],
        key=str.lower,
    )
    if not files:
        raise FileNotFoundError(f"No .ionapi files found in: {ionapi_dir}")

    print(f"\n  🔹  Select the {label} ION API file:")
    for i, f in enumerate(files, start=1):
        print(f"    {i}. {f}")

    while True:
        try:
            choice = int(input(f"\n  Enter choice (1-{len(files)}): "))
            if 1 <= choice <= len(files):
                return ionapi_dir / files[choice - 1]
            print("  ❌  Invalid choice. Try again.")
        except ValueError:
            print("  ❌  Please enter a number.")


def _prompt_company_division_forced(label: str) -> None:
    defaults = load_user_defaults()
    default_company = defaults.get("company", "100")

    print(f"\n  Configure {label} company / division:")
    company = input(f"    Company  (default: {default_company}):  ").strip()

    CONFIG["company"]  = company if company else default_company
    CONFIG["division"] = ""  # translation data is global — division must be blank

    defaults["company"]  = CONFIG["company"]
    defaults["division"] = ""
    save_user_defaults(defaults)


def setup_tenant(ionapi_dir: Path, label: str) -> dict:
    print(f"\n{'─' * 60}")
    print(f"  📡  Configure {label} tenant")
    print(f"{'─' * 60}")

    CONFIG["tenant"] = _select_ionapi_forced(ionapi_dir, label)
    _prompt_company_division_forced(label)
    _get_ion_token_global()

    snap = snapshot_config()
    print(f"\n  ✅  {label} ready: {Path(CONFIG['tenant']).name}  "
          f"CONO={CONFIG.get('company', '')}  DIVI={CONFIG.get('division', '')}")
    return snap


# =============================================================================
# EXPORTMI.Select → parse REPL rows
# =============================================================================

def fetch_exportmi(
    session: requests.Session,
    query: str,
    label: str,
) -> list[dict]:
    """
    Calls EXPORTMI.Select with the given query and returns a list of dicts,
    one per data row, keyed by the column names found in the HDRS row.
    Handles 401 token refresh automatically.
    """
    list_url = CONFIG["api_url"]

    payload = {
        "program": "EXPORTMI",
        "transactions": [{
            "transaction": "Select",
            "record": {
                "QERY": query,
                "SEPC": EXPORTMI_SEP,
                "HDRS": "1",
            },
            "selectedColumns": ["QERY", "SEPC", "HDRS", "REPL"],
        }],
    }

    headers = {
        "Authorization": f"Bearer {CONFIG['access_token']}",
        "Content-Type": "application/json",
    }

    response = session.post(list_url, json=payload, headers=headers)
    if response.status_code == 401:
        _get_ion_token_global()
        headers["Authorization"] = f"Bearer {CONFIG['access_token']}"
        response = session.post(list_url, json=payload, headers=headers)

    response.raise_for_status()
    data = response.json()

    repl_rows: list[str] = []
    for result in data.get("results", []):
        for record in result.get("records", []):
            repl_val = record.get("REPL", "")
            if repl_val:
                repl_rows.append(repl_val)

    if not repl_rows:
        print(f"  ⚠️   No rows returned from EXPORTMI ({label}).")
        return []

    # First row is the header (HDRS:1)
    col_names = [c.strip() for c in repl_rows[0].rstrip(EXPORTMI_SEP).split(EXPORTMI_SEP)]

    parsed: list[dict] = []
    for raw in repl_rows[1:]:
        values = raw.rstrip(EXPORTMI_SEP).split(EXPORTMI_SEP)
        values += [""] * (len(col_names) - len(values))
        row = {col_names[i]: values[i].strip() for i in range(len(col_names))}
        parsed.append(row)

    print(f"  ✅  {len(parsed):,} records retrieved ({label}).")
    return parsed


# =============================================================================
# Join MBMTRN header + MBMTRD details on IDTR
# =============================================================================

def join_header_details(
    hdr_rows: list[dict],
    dtl_rows: list[dict],
) -> list[dict]:
    """
    Joins MBMTRN header rows and MBMTRD detail rows on IDTR (one-to-many).
    Each detail row is merged with its parent header to produce one combined
    dict.  Orphaned detail rows (no matching header) are silently skipped.
    """
    hdr_index: dict[str, dict] = {r["TRIDTR"]: r for r in hdr_rows}

    combined: list[dict] = []
    for dtl in dtl_rows:
        idtr = dtl.get("TDIDTR", "")
        hdr  = hdr_index.get(idtr)
        if hdr is None:
            continue
        combined.append({**hdr, **dtl})

    return combined


# =============================================================================
# Build API records from combined rows
# =============================================================================

def build_trn_record(hdr_row: dict) -> dict:
    """Build the CRS881MI.AddTranslation record from a header row."""
    rec: dict = {}
    for src_key, api_key in HDR_FIELD_MAP.items():
        val = hdr_row.get(src_key, "")
        if val:
            rec[api_key] = val
    return rec


def build_trd_record(combined_row: dict) -> dict:
    """Build the CRS881MI.AddTranslData record from a combined header + detail row."""
    rec: dict = {"CONO": CONFIG.get("company", "")}
    for src_key, api_key in {**HDR_FIELD_MAP, **DTL_FIELD_MAP}.items():
        val = combined_row.get(src_key, "")
        if val:
            rec[api_key] = val
    return rec


# =============================================================================
# EVS100 Excel export
# =============================================================================

def export_evs100_xlsx(
    trn_records: list[dict],
    trd_records: list[dict],
    out_dir: Path,
) -> Path:
    """
    Writes records to an EVS100 import file with two data sheets.

    Layout
    ------
    Control sheet (tab "Control"):
        Row 1 header : Worksheet | Description | Data
        Row 2 entry  : API_CRS881MI_AddTranslation  | Translation header  | x
        Row 3 entry  : API_CRS881MI_AddTranslData    | Translation details | x

    Data sheet "API_CRS881MI_AddTranslation":
        Row 1 – field names   : MESSAGE | TRQF | MSTD | MVRS | BMSG | IBOB | ELMP | ELMD | ELMC | MBMC
        Row 2 – descriptions  : <blank> | Qualifier | Message standard | …
        Row 3 – required      : no | yes | yes | yes | yes | yes | yes | no | no | no
        Row 4+ – data rows

    Data sheet "API_CRS881MI_AddTranslData":
        Row 1 – field names   : MESSAGE | TRQF | MSTD | … | MVXP | EXTP | MVXD | MBMD | TX40
        Row 2 – descriptions  : <blank> | Qualifier | …
        Row 3 – required      : no | yes | yes | …
        Row 4+ – data rows

    Returns the path of the written file.
    """
    SHEET_TRN = "API_CRS881MI_AddTranslation"
    SHEET_TRD = "API_CRS881MI_AddTranslData"

    DESCS_TRN = [
        None, "Translation ID", "Qualifier", "Message standard", "Message version",
        "Business message", "In/Out", "Element path",
        "Element description", "Element container", "Message context",
    ]
    REQD_TRN = ["no"] + ["yes"] * len(ADD_TRANSLATION_FIELDS)
    COLS_TRN  = ["MESSAGE"] + ADD_TRANSLATION_FIELDS

    DESCS_TRD = [
        None, "Company", "Translation ID", "Qualifier", "Message standard",
        "Message version", "Business message", "In/Out", "Element path",
        "Element description", "Element container", "Message context",
        "Division", "Text 15", "MVX path", "Extension type", "MVX description",
        "Message description", "Text 40",
    ]
    REQD_TRD = ["no"] + ["yes"] * len(ADD_TRANSL_DATA_FIELDS)
    COLS_TRD  = ["MESSAGE"] + ADD_TRANSL_DATA_FIELDS

    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"API_CRS881MI_{ts}.xlsx"

    wb = xlsxwriter.Workbook(str(out_path))

    # ── Control sheet ─────────────────────────────────────────────────────
    ws_ctrl = wb.add_worksheet("Control")
    for col, val in enumerate(["Worksheet", "Description", "Data"]):
        ws_ctrl.write(0, col, val)
    for col, val in enumerate([SHEET_TRN, "Translation header", "x"]):
        ws_ctrl.write(1, col, val)
    for col, val in enumerate([SHEET_TRD, "Translation details", "x"]):
        ws_ctrl.write(2, col, val)

    # ── Helper: write one data sheet ──────────────────────────────────────
    def _write_data_sheet(ws, cols, descs, reqd, records, fields):
        for col, name in enumerate(cols):
            ws.write(0, col, name)
        for col, desc in enumerate(descs):
            if desc is not None:
                ws.write(1, col, desc)
        for col, req in enumerate(reqd):
            ws.write(2, col, req)
        for row_idx, rec in enumerate(records, start=3):
            for col, field in enumerate(fields, start=1):
                val = rec.get(field, "")
                if val:
                    ws.write(row_idx, col, val)

    # ── AddTranslation sheet ──────────────────────────────────────────────
    ws_trn = wb.add_worksheet(SHEET_TRN)
    _write_data_sheet(ws_trn, COLS_TRN, DESCS_TRN, REQD_TRN, trn_records, ADD_TRANSLATION_FIELDS)

    # ── AddTranslData sheet ───────────────────────────────────────────────
    ws_trd = wb.add_worksheet(SHEET_TRD)
    _write_data_sheet(ws_trd, COLS_TRD, DESCS_TRD, REQD_TRD, trd_records, ADD_TRANSL_DATA_FIELDS)

    wb.close()
    return out_path


# =============================================================================
# Upload file to M3 File Management (PUT)
# =============================================================================

def upload_file_to_m3(
    file_path: Path,
    session: requests.Session,
) -> bool:
    """
    Uploads file_path to the M3 FileImport area via the File Management REST API.
    Returns True on success, False on failure.
    """
    filename   = file_path.name
    upload_url = (
        f"{CONFIG['iu']}/{CONFIG['ti']}"
        f"/M3/foundation-rest/file-management/v1/file/FileImport/{filename}"
    )

    headers = {
        "Authorization": f"Bearer {CONFIG['access_token']}",
        "Content-Type": "application/octet-stream",
    }

    with open(file_path, "rb") as fh:
        file_bytes = fh.read()

    response = session.put(upload_url, data=file_bytes, headers=headers)

    if response.status_code == 401:
        get_ion_token()
        headers["Authorization"] = f"Bearer {CONFIG['access_token']}"
        response = session.put(upload_url, data=file_bytes, headers=headers)

    if response.status_code in (200, 201, 204):
        print(f"  ✅  Upload succeeded (HTTP {response.status_code}): {filename}")
        return True
    else:
        print(f"  ❌  Upload failed (HTTP {response.status_code}): {response.text[:200]}")
        return False


# =============================================================================
# Trigger EVS100MI.ImportFile
# =============================================================================

def process_file_in_m3(
    filename: str,
    session: requests.Session,
) -> bool:
    """
    Calls EVS100MI.ImportFile with FNAM=filename so M3 processes the uploaded
    spreadsheet through the EVS100 interface.
    Returns True on success, False on failure.
    """
    payload = {
        "program": "EVS100MI",
        "transactions": [{
            "transaction": "ImportFile",
            "record": {"FNAM": filename},
        }],
    }

    try:
        result = post_to_m3(payload, session)
        for res in result.get("results", []):
            err = res.get("errorMessage", "").strip() if isinstance(res, dict) else ""
            if err:
                print(f"  ❌  EVS100MI.ImportFile error: {err}")
                return False
        print(f"  ✅  EVS100MI.ImportFile processed successfully: {filename}")
        return True
    except Exception as exc:
        print(f"  ❌  EVS100MI.ImportFile failed: {exc}")
        return False


# =============================================================================
# Direct API calls to DEST  (mirrors MIG_SyncPartnerRef pattern)
# =============================================================================

def run_add_translation_batched(
    records: list[dict],
    session: requests.Session,
    batch_size: int,
) -> tuple[list[dict], list[tuple[dict, str]]]:
    """
    Calls CRS881MI.AddTranslation on DEST for each record (batched).
    Returns (successes, failures).
    """
    successes: list[dict] = []
    failures:  list[tuple[dict, str]] = []

    for i in tqdm(
        range(0, len(records), batch_size),
        desc="CRS881MI.AddTranslation",
        unit="batch",
        leave=False,
    ):
        batch = records[i : i + batch_size]
        payload = {
            "program": "CRS881MI",
            "transactions": [
                {
                    "transaction": "AddTranslation",
                    "record": rec,
                    "selectedColumns": ADD_TRANSLATION_FIELDS,
                }
                for rec in batch
            ],
        }
        try:
            result = post_to_m3(payload, session)
            api_results = result.get("results", [])
            for j, res in enumerate(api_results):
                err = res.get("errorMessage", "").strip() if isinstance(res, dict) else ""
                rec = batch[j] if j < len(batch) else {}
                if err:
                    failures.append((rec, err))
                else:
                    successes.append(rec)
            for k in range(len(api_results), len(batch)):
                failures.append((batch[k], "No result returned"))
        except Exception as exc:
            for rec in batch:
                failures.append((rec, str(exc)))

    return successes, failures


def run_add_transl_data_batched(
    records: list[dict],
    session: requests.Session,
    batch_size: int,
) -> tuple[list[dict], list[tuple[dict, str]]]:
    """
    Calls CRS881MI.AddTranslData on DEST for each record (batched).
    Returns (successes, failures).
    """
    successes: list[dict] = []
    failures:  list[tuple[dict, str]] = []

    for i in tqdm(
        range(0, len(records), batch_size),
        desc="CRS881MI.AddTranslData",
        unit="batch",
        leave=False,
    ):
        batch = records[i : i + batch_size]
        payload = {
            "program": "CRS881MI",
            "transactions": [
                {
                    "transaction": "AddTranslData",
                    "record": rec,
                    "selectedColumns": ADD_TRANSL_DATA_FIELDS,
                }
                for rec in batch
            ],
        }
        try:
            result = post_to_m3(payload, session)
            api_results = result.get("results", [])
            for j, res in enumerate(api_results):
                err = res.get("errorMessage", "").strip() if isinstance(res, dict) else ""
                rec = batch[j] if j < len(batch) else {}
                if err:
                    failures.append((rec, err))
                else:
                    successes.append(rec)
            for k in range(len(api_results), len(batch)):
                failures.append((batch[k], "No result returned"))
        except Exception as exc:
            for rec in batch:
                failures.append((rec, str(exc)))

    return successes, failures


def print_summary(
    label: str,
    successes: list[dict],
    failures: list[tuple[dict, str]],
) -> None:
    from collections import Counter
    total = len(successes) + len(failures)
    print(f"\n{'═' * 60}")
    print(f"  {label}")
    print(f"  ✅ Succeeded : {len(successes)}")
    print(f"  ❌ Failed    : {len(failures)}")
    print(f"  📋 Total     : {total}")
    print(f"{'═' * 60}")

    msg_counts: Counter = Counter()
    msg_counts["OK"] = len(successes)
    for _, err in failures:
        msg_counts[err] += 1
    print()
    for msg, count in msg_counts.most_common():
        print(f"  {msg} = {count}")


def export_api_errors_xlsx(
    trn_successes: list[dict],
    trn_failures:  list[tuple[dict, str]],
    trd_successes: list[dict],
    trd_failures:  list[tuple[dict, str]],
    out_dir: Path,
) -> Path:
    """
    Writes three sheets to an xlsx file:
      • Summary             – error-message totals for both steps
      • AddTranslation Err  – one row per failed AddTranslation record
      • AddTranslData Err   – one row per failed AddTranslData record
    Returns the path of the written file.
    """
    from collections import Counter

    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"MIG_SyncTranslData_Errors_{ts}.xlsx"

    wb = OWorkbook()

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", start_color="2F5496")
    ok_fill  = PatternFill("solid", start_color="E2EFDA")
    err_fill = PatternFill("solid", start_color="FCE4D6")
    ctr      = Alignment(horizontal="center")

    def _write_error_sheet(ws, fields, failures):
        for col, f in enumerate(fields, start=1):
            c = ws.cell(row=1, column=col, value=f)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = ctr
        for row_i, (rec, err) in enumerate(failures, start=2):
            for col, f in enumerate(fields, start=1):
                val = err if f == "ERROR" else rec.get(f, "")
                c = ws.cell(row=row_i, column=col, value=val)
                c.font = Font(name="Arial")
        for col, f in enumerate(fields, start=1):
            max_len = max(
                len(f),
                *(len(str(rec.get(f, "") if f != "ERROR" else err))
                  for rec, err in failures),
            )
            ws.column_dimensions[
                ws.cell(row=1, column=col).column_letter
            ].width = min(max_len + 2, 60)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    for col, heading in enumerate(["Step", "Message", "Count"], start=1):
        c = ws_sum.cell(row=1, column=col, value=heading)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = ctr

    row_i = 2
    for step_label, successes, failures in [
        ("AddTranslation", trn_successes, trn_failures),
        ("AddTranslData",  trd_successes, trd_failures),
    ]:
        counts: Counter = Counter()
        counts["OK"] = len(successes)
        for _, err in failures:
            counts[err] += 1
        for msg, count in counts.most_common():
            fill = ok_fill if msg == "OK" else err_fill
            for col, val in enumerate([step_label, msg, count], start=1):
                c = ws_sum.cell(row=row_i, column=col, value=val)
                c.fill = fill
                c.font = Font(name="Arial")
                if col == 3:
                    c.alignment = ctr
            row_i += 1

    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 80
    ws_sum.column_dimensions["C"].width = 10

    # ── Sheet 2: AddTranslation errors ────────────────────────────────────
    if trn_failures:
        ws_trn = wb.create_sheet("AddTranslation Err")
        _write_error_sheet(ws_trn, ADD_TRANSLATION_FIELDS + ["ERROR"], trn_failures)

    # ── Sheet 3: AddTranslData errors ─────────────────────────────────────
    if trd_failures:
        ws_trd = wb.create_sheet("AddTranslData Err")
        _write_error_sheet(ws_trd, ADD_TRANSL_DATA_FIELDS + ["ERROR"], trd_failures)

    wb.save(out_path)
    return out_path


# =============================================================================
# Main
# =============================================================================

def sync_transl_data() -> None:
    ionapi_dir = Path(__file__).parent / "ionapi"
    batch_size = DEFAULT_BATCH_SIZE

    print(f"\n{'═' * 60}")
    print(f"  MIG_SyncTranslData")
    print(f"{'═' * 60}")

    # ------------------------------------------------------------------ #
    # Step 1 – SOURCE: EXPORTMI.Select MBMTRN (header)                   #
    # ------------------------------------------------------------------ #
    source_snap = setup_tenant(ionapi_dir, "SOURCE")
    _apply_global_url()

    print("\n🔍  Step 1 – EXPORTMI.Select MBMTRN (header) …")
    with requests.Session() as session:
        hdr_rows = fetch_exportmi(session, EXPORTMI_HDR_QUERY, label="MBMTRN")

        if not hdr_rows:
            print("⚠️   No header records found in SOURCE. Exiting.")
            return

        # ------------------------------------------------------------------ #
        # Step 2 – SOURCE: EXPORTMI.Select MBMTRD (details)                  #
        # ------------------------------------------------------------------ #
        print("\n🔍  Step 2 – EXPORTMI.Select MBMTRD (details) …")
        dtl_rows = fetch_exportmi(session, EXPORTMI_DTL_QUERY, label="MBMTRD")

    if not dtl_rows:
        print("⚠️   No detail records found in SOURCE. Exiting.")
        return

    # ------------------------------------------------------------------ #
    # Step 3 – Join header + details on IDTR; build API records          #
    # ------------------------------------------------------------------ #
    print("\n🔗  Step 3 – Joining MBMTRN + MBMTRD on IDTR …")
    combined = join_header_details(hdr_rows, dtl_rows)

    if not combined:
        print("⚠️   No records produced after join. Exiting.")
        return

    # Deduplicate headers by TRIDTR (one AddTranslation row per unique header)
    seen_idtr: set[str] = set()
    trn_records: list[dict] = []
    for row in hdr_rows:
        idtr = row.get("TRIDTR", "")
        if idtr not in seen_idtr:
            seen_idtr.add(idtr)
            trn_records.append(build_trn_record(row))

    trd_records = [build_trd_record(row) for row in combined]

    print(f"\n  📊  Headers (AddTranslation) : {len(trn_records):,}")
    print(f"  📊  Details (AddTranslData)  : {len(trd_records):,}")

    # Preview the first 10 detail records
    preview = trd_records[:10]
    fields  = ADD_TRANSL_DATA_FIELDS
    col_w   = {
        f: max(len(f), max((len(str(r.get(f, ""))) for r in preview), default=0))
        for f in fields
    }
    hdr_line = "  " + "  ".join(f.ljust(col_w[f]) for f in fields)
    sep_line = "  " + "  ".join("─" * col_w[f] for f in fields)

    print(f"\n{'═' * max(len(hdr_line), 60)}")
    print(f"  📋  Plan: {len(trd_records):,} AddTranslData record(s) to export")
    print(f"{'═' * max(len(hdr_line), 60)}")
    print(hdr_line)
    print(sep_line)
    for rec in preview:
        print("  " + "  ".join(str(rec.get(f, "")).ljust(col_w[f]) for f in fields))
    if len(trd_records) > 10:
        print(f"  … and {len(trd_records) - 10:,} more record(s)")
    print(f"{'═' * max(len(hdr_line), 60)}")

    # ------------------------------------------------------------------ #
    # Step 4 – Export to EVS100 Excel file                               #
    # ------------------------------------------------------------------ #
    EVS100_TO_PROCESS.mkdir(parents=True, exist_ok=True)
    
    print(f"\n▶   Step 4 – Exporting to EVS100 Excel file …")
    xlsx_path = export_evs100_xlsx(trn_records, trd_records, EVS100_TO_PROCESS)
    print(f"  📄  Saved to: evs100/ToProcess/{xlsx_path.name}")

    # ------------------------------------------------------------------ #
    # Step 5 – Configure DEST tenant                                      #
    # ------------------------------------------------------------------ #
    restore_config(source_snap)
    dest_snap    = setup_tenant(ionapi_dir, "DEST")
    dest_company = CONFIG.get("company", "")

    # Build DEST-specific trd_records: identical to SOURCE set but with
    # CONO overridden to the DEST company number.
    trd_records_dest = [{**rec, "CONO": dest_company} for rec in trd_records]

    # ------------------------------------------------------------------ #
    # Step 6 – Direct API calls: CRS881MI.AddTranslation + AddTranslData  #
    # ------------------------------------------------------------------ #
    confirm = input(
        "\n  Call CRS881MI APIs directly against DEST? [y/N]: "
    ).strip().lower()

    trn_successes: list[dict] = []
    trn_failures:  list[tuple[dict, str]] = []
    trd_successes: list[dict] = []
    trd_failures:  list[tuple[dict, str]] = []

    if confirm == "y":
        restore_config(dest_snap)
        get_ion_token()

        with requests.Session() as session:
            print(f"\n▶   Step 6a – CRS881MI.AddTranslation ({len(trn_records):,} records) …")
            trn_successes, trn_failures = run_add_translation_batched(
                trn_records, session, batch_size
            )

            print(f"\n▶   Step 6b – CRS881MI.AddTranslData ({len(trd_records_dest):,} records) …")
            trd_successes, trd_failures = run_add_transl_data_batched(
                trd_records_dest, session, batch_size
            )

        print_summary("AddTranslation [DEST]", trn_successes, trn_failures)
        print_summary("AddTranslData  [DEST]", trd_successes, trd_failures)

        if trn_failures or trd_failures:
            err_path = export_api_errors_xlsx(
                trn_successes, trn_failures,
                trd_successes, trd_failures,
                Path(__file__).parent,
            )
            print(f"\n  📄  Error report saved to: {err_path.name}")
    else:
        print("  ℹ️   API calls skipped.")

    # ------------------------------------------------------------------ #
    # Step 7 – Run Summary                                                 #
    # ------------------------------------------------------------------ #
    print(f"\n{'═' * 60}")
    print(f"  MIG_SyncTranslData – Run Summary")
    print(f"{'═' * 60}")
    print(f"  📋  Header rows (MBMTRN)    : {len(hdr_rows):,}")
    print(f"  📋  Detail rows (MBMTRD)    : {len(dtl_rows):,}")
    print(f"  📄  AddTranslation records  : {len(trn_records):,}")
    print(f"  📄  AddTranslData records   : {len(trd_records):,}")
    print(f"  📁  Output file             : {xlsx_path.name}")
    if confirm == "y":
        print(f"  ✅  AddTranslation  [API]   : {len(trn_successes):,} ok  /  {len(trn_failures):,} failed")
        print(f"  ✅  AddTranslData   [API]   : {len(trd_successes):,} ok  /  {len(trd_failures):,} failed")
    print(f"{'═' * 60}")

    # ------------------------------------------------------------------ #
    # Step 8 – Optional: upload and process in M3 via EVS100              #
    # ------------------------------------------------------------------ #
    restore_config(dest_snap)
    get_ion_token()

    send = input("\n  Also upload this file to M3 via EVS100? [y/N]: ").strip().lower()
    if send != "y":
        print("  ℹ️   File not sent. Done.")
        return

    with requests.Session() as session:
        print(f"\n▶   Step 8a – Uploading {xlsx_path.name} to M3 …")
        uploaded = upload_file_to_m3(xlsx_path, session)

        if uploaded:
            trigger = input(
                "\n  Process file in M3 (EVS100MI.ImportFile)? [y/N]: "
            ).strip().lower()
            if trigger == "y":
                print(f"\n▶   Step 8b – Triggering EVS100MI.ImportFile …")
                process_file_in_m3(xlsx_path.name, session)
            else:
                print("  ℹ️   File uploaded but not processed. Done.")


# =============================================================================
# Entry Point
# =============================================================================

if __name__ == "__main__":
    sync_transl_data()
