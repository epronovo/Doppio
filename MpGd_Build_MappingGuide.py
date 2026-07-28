# MpGd_Build_MappingGuide.py
# -----------------------------------------------------------------------
# PURPOSE
#   Queries SQLite for all table/column/API/FieldHelp/DocBits/BOD data
#   associated with a given API and produces a formatted Excel mapping
#   guide workbook.  The output uses a two-row grouped header with
#   color-coded column sections (Table, Program, API, BOD, Herff Jones
#   mapping, Help, Doppio mapping), frozen panes at column D / row 3,
#   auto-fitted column widths, and an Excel table style for easy filtering.
#   A UNION query covers both table-mapped fields and any API-only fields
#   present in HerffJonesMappingGuide but not yet linked to a table column.
#
# INPUTS
#   - m3TableCols, m3Api2Table, m3FieldHelp, m3Prompts,
#     DoppioGuide, HerffJonesMappingGuide, CSYRPL,
#     m3Docbits, cmifld, m3Bod2Api  (all SQLite)
#
# OUTPUTS
#   - Excel .xlsx mapping guide file written to the guides/ output folder
#
# DEPENDENCIES
#   - config.py, openpyxl
# -----------------------------------------------------------------------
import logging
import sqlite3
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import get_sqlite_db_path

logger = logging.getLogger(__name__)

def build_mapping_guide(api_value, output_file_name):

    # --- Paths ---
    db_path = get_sqlite_db_path()
    OUTPUT_DIR = "/Users/ericpronovost/Doppio/MappingGuides/mapping/guides/" 
    output_path = os.path.join(OUTPUT_DIR, f"{output_file_name}.xlsx")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # --- SQL Query ---
    query = """ 
    SELECT DISTINCT tc.Sequence,t2a.TableName,tc.ColumnName,tc.Description,tc.DataType,tc.Length,tc.Decimals,CASE WHEN instr(tc.Indexes, '00') > 0 THEN 'Yes' ELSE '' END AS PrimaryKey
            ,''[Table_8]
            ,t2a.Program,t2a.Panel,CASE WHEN r1prtf is null then '' ELSE 'yes' END AS [IDM],CASE WHEN DocBitsField is null then '' ELSE 'yes' END AS [DocBits]
            ,''[Pgm_4]
            ,CASE WHEN t2a.API IS NULL THEN hj.API ELSE t2a.API END [API]
            ,CASE WHEN t2a.TransactionName IS NULL THEN hj.TransactionName ELSE t2a.TransactionName END [TransactionName]
            ,CASE WHEN t2a.FieldName IS NULL THEN hj.FieldName ELSE CASE WHEN t2a.TransactionName='' then '' else t2a.FieldName END END [FieldName]
            ,CASE WHEN MAND = 1 then 'yes' ELSE '' END Mandatory
            ,''[API_4]
            ,COALESCE(XPath,'') AS XPath
            ,''[BOD_1]
            ,COALESCE(hj.MappingNotes,'')MappingNotes,COALESCE(hj.Responsible,'')Responsible,COALESCE(hj.Required,'')Required,COALESCE(hj.DefaultValue,'')DefaultValue
            ,''[Herff Jones_4]
            ,CASE WHEN fh2.Definition NOT NULL THEN COALESCE(fh2.Definition,'') ELSE COALESCE(fh1.Definition,'') END AS [Definition]
            ,CASE WHEN fh2.Alternatives NOT NULL THEN COALESCE(REPLACE(fh2.Alternatives, CHAR(13), CHAR(10)),'') ELSE COALESCE(REPLACE(fh1.Alternatives, CHAR(13), CHAR(10)),'') END AS [Alternatives]
            ,COALESCE(p.PromptProgram,'') AS [PromptPgm]
            ,''[Help_3]
            ,COALESCE(dg.MappingNotes,'')DG_Notes,COALESCE(dg.Responsible,'')DG_Responsible,COALESCE(dg.Required,'')DG_Required,COALESCE(dg.DefaultValue,'')DG_Example
            ,''[Doppio_4]
        FROM m3TableCols tc
        LEFT JOIN m3Api2Table t2a ON substr(t2a.TableName, 1, 6) = tc.TableName AND t2a.ColumnName = tc.ColumnName
        LEFT JOIN m3Bod2Api b2a ON b2a.API = t2a.API AND b2a.TransactionName = t2a.TransactionName AND b2a.FieldName = t2a.FieldName 
        LEFT JOIN m3FieldHelp fh2 ON fh2.FieldHelpID = tc.ColumnName
        LEFT JOIN m3FieldHelp fh1 ON fh1.FieldHelpID = SUBSTR(tc.ColumnName,-4)
        LEFT JOIN m3Prompts p ON p.FieldName = t2a.FieldName 
        LEFT JOIN DoppioGuide dg ON dg.TableName = t2a.TableName AND dg.ColumnName = t2a.ColumnName
        LEFT JOIN HerffJonesMappingGuide hj ON hj.API = t2a.API AND hj.TransactionName = t2a.TransactionName AND hj.FieldName = t2a.FieldName
        LEFT JOIN CSYRPL ON R1OBJC = tc.ColumnName
        LEFT JOIN m3Docbits db ON db.API = t2a.API AND db.FieldName = SUBSTR(tc.ColumnName,-4)
        LEFT JOIN cmifld fld ON MINM = t2a.API AND TRNM = t2a.TransactionName AND TRTP = 'I' AND FLNM = t2a.FieldName 
        WHERE 1=1
        AND tc.TableName IN (SELECT DISTINCT tc.TableName FROM m3TableCols tc LEFT JOIN m3Api2Table t2a on substr(t2a.TableName, 1, 6) = tc.TableName AND t2a.ColumnName = tc.ColumnName WHERE API = :api)
        AND t2a.TableName IN (SELECT DISTINCT TableName FROM m3Api2Table WHERE API = :api)
    UNION
    SELECT 0 Sequence,''TableName,''ColumnName,FLDS AS [Description],CASE WHEN TYPE = 'A' THEN 'String' ELSE 'Decimal' END AS [DataType],LENG AS [Length],''Decimals,CASE WHEN instr(tc.Indexes, '00') > 0 THEN 'Yes' ELSE '' END AS PrimaryKey
        ,''[Table]
        ,''Program,''Panel,'',''
        ,''[Pgm]
        ,CASE WHEN t2a.API IS NULL THEN hj.API ELSE t2a.API END [MI]
        ,CASE WHEN t2a.TransactionName IS NULL THEN hj.TransactionName ELSE t2a.TransactionName END [Transaction]
        ,CASE WHEN t2a.FieldName IS NULL THEN hj.FieldName ELSE t2a.FieldName END [Field]
        ,CASE WHEN MAND = 1 then 'yes' ELSE '' END Mandatory
        ,''[API]
        ,COALESCE(XPath,'') AS XPath
        ,''[BOD]
        ,COALESCE(hj.MappingNotes,'')MappingNotes,COALESCE(hj.Responsible,'')Responsible,COALESCE(hj.Required,'')Required,COALESCE(hj.DefaultValue,'')DefaultValue
        ,''[Herff Jones]
        ,COALESCE(fh1.Definition,'') AS [Definition]
        ,COALESCE(REPLACE(fh1.Alternatives, CHAR(13), CHAR(10)),'') AS [Alternatives]
        ,COALESCE(p.PromptProgram,'') AS [Prompt]
        ,''[Help]
        ,''DG_Notes,''DG_Responsible,''DG_Required,''DG_Example
        ,''[Doppio]
    FROM HerffJonesMappingGuide hj
    LEFT JOIN m3Api2Table t2a ON hj.API = t2a.API AND hj.TransactionName = t2a.TransactionName AND hj.FieldName = t2a.FieldName
    LEFT JOIN m3TableCols tc ON t2a.TableName = tc.TableName AND t2a.ColumnName = tc.ColumnName
    LEFT JOIN m3Bod2Api b2a ON b2a.API = t2a.API AND b2a.TransactionName = t2a.TransactionName AND b2a.FieldName = t2a.FieldName 
    LEFT JOIN cmifld fld ON MINM = hj.API AND TRNM = hj.TransactionName AND TRTP = 'I' AND FLNM = hj.FieldName 
    LEFT JOIN m3FieldHelp fh1 ON fh1.FieldHelpID = CASE WHEN t2a.FieldName IS NULL THEN hj.FieldName ELSE t2a.FieldName END 
    LEFT JOIN m3Prompts p ON p.FieldName = CASE WHEN t2a.FieldName IS NULL THEN hj.FieldName ELSE t2a.FieldName END 
    WHERE t2a.TableName is null
    AND hj.API = :api
    ORDER BY 1,2
    """
    
    # --- Run query with parameter ---
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(query, {"api": api_value}).fetchall()
    conn.close()

    if not rows:
        raise ValueError("No data returned from query.")

    col_names = rows[0].keys()
    col_count = len(col_names)

    # --- Workbook setup ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping Guide"

    # --- Define group colors (dark colors with bold white font) ---
    group_colors = {
        "Table": "8164A2",         # plum
        "Pgm": "1A6B25",           # dark green
        "API": "156082",           # dark teal
        "BOD": "E97131",           # orange
        "Herff Jones": "0F9ED5",   # turquoise
        "Help": "0F2841",          # dark blue
        "Doppio": "C00000"         # dark red
    }

    # --- Build headers ---
    row1 = [""] * col_count
    row2 = list(col_names)
    merge_commands = []

    i = 0
    while i < col_count:
        col = col_names[i]
        if "_" in col:
            base, suffix = col.rsplit("_", 1)
            if suffix.isdigit():
                num = int(suffix)
                start_col = max(0, i - num)
                end_col = i
                merge_commands.append((start_col, end_col, base))
                row1[start_col] = base
                for j in range(start_col + 1, end_col + 1):
                    row1[j] = ""
                i += 1
                continue
        i += 1

    # --- Write header rows ---
    ws.append(row1)  # Row 1: grouped headers
    ws.append(row2)  # Row 2: column names

    # --- Merge group headers and apply colors ---
    for start, end, title in merge_commands:
        # Merge Row 1
        ws.merge_cells(start_row=1, start_column=start+1, end_row=1, end_column=end+1)
        cell = ws.cell(row=1, column=start+1)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Avenir", size=12, bold=True, color="FFFFFF")  # bold white font for Row 1
        fill_color = group_colors.get(title, None)
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color,
                                    end_color=fill_color,
                                    fill_type="solid")
        
        # Apply same background & white font to Row 2 for all columns in group
        for col_idx in range(start, end+1):
            cell2 = ws.cell(row=2, column=col_idx+1)
            if fill_color:
                cell2.fill = PatternFill(start_color=fill_color,
                                        end_color=fill_color,
                                        fill_type="solid")
            # Last column in group (_N) → font matches background to blend
            if col_idx == end:
                cell2.font = Font(name="Avenir", size=12, bold=True, color=fill_color)
            else:
                cell2.font = Font(name="Avenir", size=12, bold=True, color="FFFFFF")

    # --- Write data rows (no manual shading, table handles it) ---
    for r_idx, r in enumerate(rows, start=3):
        ws.append(list(r))

    # --- Autofit columns with max width 60 ---
    for i, col in enumerate(col_names, start=1):
        max_length = max(
            [len(str(ws.cell(row=r, column=i).value)) if ws.cell(row=r, column=i).value is not None else 0 for r in range(1, ws.max_row+1)]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(60, max(15, max_length + 2))

    # --- Freeze panes at row 3 ---
    ws.freeze_panes = "D3"

    # --- Create Excel Table (Plum TableStyleMedium20) over Row 2 + data ---
    table_ref = f"A2:{get_column_letter(col_count)}{ws.max_row}"
    table = Table(displayName="MappingGuide", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium19",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # --- Set all fonts to Avenir, size 12 ---
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            current_font = cell.font
            # Preserve color, bold, italic, underline, etc., but change name and size
            cell.font = Font(
                name="Avenir",
                size=12,
                bold=current_font.bold,
                italic=current_font.italic,
                vertAlign=current_font.vertAlign,
                underline=current_font.underline,
                strike=current_font.strike,
                color=current_font.color
            )

            
    # --- Save workbook ---
    wb.save(output_path)
    logger.info(f"Workbook created: {output_path}")

# Optional: allow running from command line
if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%H:%M:%S"
    )
    build_mapping_guide("MMS200MI", "MappingGuide_MMS200MI_Items.xlsx")