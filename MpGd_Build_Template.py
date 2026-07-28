# MpGd_Build_Template.py
# -----------------------------------------------------------------------
# PURPOSE
#   For each required API transaction in a selected mapping guide table,
#   generates a transposed Excel template workbook intended for data-entry
#   use.  Each workbook contains a Control sheet listing its worksheets,
#   and one worksheet per transaction where fields become columns and the
#   rows represent: field name, description+type, required flag, mapping
#   value, and example value.  FieldHelp is attached as cell comments.
#   Also calls MpGd_Build_MappingGuide to produce the companion full
#   reference mapping guide for each API.  The last-used guide table name
#   is remembered across runs via a hidden .last_mapping_guide file.
#
# INPUTS
#   - Selected GUIDE_ mapping guide table (SQLite, prompted interactively)
#   - cmipgm, cmitrn, cmifld   (SQLite) — API/transaction/field metadata
#   - m3TableCols, m3FieldHelp (SQLite) — column metadata and help text
#
# OUTPUTS
#   - Excel .xlsx template workbooks written to mapping/templates/{guide}/
#   - Companion mapping guide workbooks (via MpGd_Build_MappingGuide)
#
# DEPENDENCIES
#   - config.py, MpGd_Build_MappingGuide.py, openpyxl
# -----------------------------------------------------------------------

import logging
import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.comments import Comment
from MpGd_Build_MappingGuide import build_mapping_guide

from config import get_sqlite_db_path, BASE_DIR

logger = logging.getLogger(__name__)

# ============================================================
# CONFIG
# ============================================================

DB_PATH = get_sqlite_db_path()
BASE_OUTPUT_DIR = BASE_DIR / "mapping/templates"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LAST_GUIDE_FILE = os.path.join(SCRIPT_DIR, ".last_mapping_guide")

# ============================================================
# HELPERS
# ============================================================

def autofit_ws(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)


def add_comment(ws, row, col, text):
    if not text or text.strip() == "":
        return

    lines = text.split("\n")
    max_line_length = max(len(line) for line in lines)

    comment = Comment(text, "System")
    comment.width = max(200, min(max_line_length * 10, 1000))
    comment.height = max(80, min(len(lines) * 20, 1000))

    ws.cell(row=row, column=col).comment = comment


def table_exists(cursor, table_name):
    cursor.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,)
    )
    return cursor.fetchone() is not None


def load_last_guide():
    if os.path.exists(LAST_GUIDE_FILE):
        with open(LAST_GUIDE_FILE, "r") as f:
            return f.read().strip()
    return None


def save_last_guide(guide_table):
    with open(LAST_GUIDE_FILE, "w") as f:
        f.write(guide_table)


def prompt_for_guide(cursor):
    last_guide = load_last_guide()

    while True:
        prompt = (
            f"Mapping guide table [{last_guide}]: "
            if last_guide else
            "Mapping guide table: "
        )

        user_input = input(prompt).strip()
        guide_table = user_input or last_guide

        if not guide_table:
            logger.error("❌ Please enter a table name.")
            continue

        if not table_exists(cursor, guide_table):
            logger.error(f"❌ Table '{guide_table}' does not exist in the database.")
            continue

        save_last_guide(guide_table)
        return guide_table

# ============================================================
# PARAMETERIZED WORKER
# ============================================================

def export_templates_for_guide(guide_table):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    if not table_exists(cursor, guide_table):
        conn.close()
        raise ValueError(f"Table '{guide_table}' does not exist.")

    save_last_guide(guide_table)

    output_dir = os.path.join(BASE_OUTPUT_DIR, guide_table)
    os.makedirs(output_dir, exist_ok=True)

    logger.info(f"Using mapping guide table: {guide_table}")
    logger.info(f"Output directory: {output_dir}")

    # master_sql = f"""
    # SELECT 
    #     'API_' || API || '_' || replace(trim(substr(OBNM, instr(OBNM, ':') + 1)),'IDI -  ','') AS Workbook,
    #     'API_' || API || '_' || TransactionName AS Worksheet,
    #     TRDS AS Description,
    #     'x' AS Data,
    #     API,
    #     TransactionName
    # FROM {guide_table} d
    # LEFT JOIN GUIDE_API a ON a.sequence = d.sequence 
    # JOIN cmipgm p ON p.MINM = API 
    # JOIN cmitrn t ON t.MINM = API AND t.TRNM = TransactionName
    # WHERE required = 'yes'
    # GROUP BY API, OBNM, TransactionName
    # ORDER BY API
    # """
    
    master_sql = f"""
    SELECT 
        'API_' || API || '_' ||
        REPLACE(
            REPLACE(
                TRIM(SUBSTR(OBNM, INSTR(OBNM, ':') + 1)),
                'IDI -  ',
                ''
            ),
            '/',
            '_'
        ) AS Workbook,
        'API_' || API || '_' || TransactionName AS Worksheet,
        TRDS AS Description,
        'x' AS Data,
        API,
        TransactionName
    FROM {guide_table} d
    JOIN cmipgm p ON p.MINM = API 
    JOIN cmitrn t ON t.MINM = API AND t.TRNM = TransactionName
    WHERE required = 'yes'
    GROUP BY API, OBNM, TransactionName
    ORDER BY API;
    """

    rows = cursor.execute(master_sql).fetchall()

    workbooks = {}
    for wb_name, ws_name, desc, data, api, trn in rows:
        workbooks.setdefault(wb_name, {"control": [], "tasks": []})
        workbooks[wb_name]["control"].append((ws_name, desc, data))
        workbooks[wb_name]["tasks"].append((api, trn, ws_name))

    for wb_name, content in workbooks.items():
        logger.info(f"Building workbook: {wb_name}.xlsx")

        api_value = wb_name.split("_")[1]
        output_file_name = wb_name.replace("API_", "MappingGuide_")
        build_mapping_guide(api_value, output_file_name)

        wb = Workbook()
        control_ws = wb.active
        control_ws.title = "Control"
        control_ws.append(["Worksheet", "Description", "Data"])

        for row in content["control"]:
            control_ws.append(list(row))

        autofit_ws(control_ws)

        for api, trn, worksheet_name in content["tasks"]:
            # detail_sql = f"""
            # SELECT
            #     FieldName AS Row1,
            #     FLDS || ' (' || TYPE || ':' || LENG || ')' AS Row2,
            #     required AS Row3,
            #     mapping AS Row4,
            #     example AS Row5,
            #     CASE 
            #         WHEN fh2.Definition IS NOT NULL THEN
            #             COALESCE(
            #                 fh2.HeadingText || CHAR(10)||CHAR(10) ||
            #                 fh2.Definition || CHAR(10) || fh2.Alternatives,''
            #             )
            #         ELSE
            #             COALESCE(
            #                 fh1.HeadingText || CHAR(10)||CHAR(10) ||
            #                 fh1.Definition || CHAR(10) || fh1.Alternatives,''
            #             )
            #     END AS Help
            # FROM {guide_table} d 
            # LEFT JOIN GUIDE_Table t ON t.sequence = d.sequence 
            # LEFT JOIN GUIDE_API a ON a.sequence = d.sequence 
            # JOIN cmifld fld ON MINM = API AND TRNM = TransactionName AND TRTP = 'I' AND FLNM = FieldName
            # LEFT JOIN m3TableCols tc ON tc.TableName = substr(t.TableName, 1, 6) AND tc.ColumnName = t.ColumnName
            # LEFT JOIN m3FieldHelp fh2 ON fh2.FieldHelpID = tc.ColumnName
            # LEFT JOIN m3FieldHelp fh1 ON fh1.FieldHelpID = SUBSTR(tc.ColumnName,-4)
            # WHERE lower(required) not in ('no')
            #   AND API = ?
            #   AND TransactionName = ?
            # ORDER BY CAST(FRPO AS INTEGER)
            # """

            detail_sql = f"""
            SELECT Row1, Row2, Row3, Row4, Row5, Help
            FROM (

                -- Fields already covered by the mapping guide
                SELECT
                    FieldName                                                  AS Row1,
                    fld.FLDS || ' (' || fld.TYPE || ':' || fld.LENG || ')'    AS Row2,
                    required                                                   AS Row3,
                    mapping                                                    AS Row4,
                    example                                                    AS Row5,
                    CASE
                        WHEN fh2.Definition IS NOT NULL THEN
                            COALESCE(
                                fh2.HeadingText || CHAR(10)||CHAR(10) ||
                                fh2.Definition  || CHAR(10) || fh2.Alternatives, '')
                        ELSE
                            COALESCE(
                                fh1.HeadingText || CHAR(10)||CHAR(10) ||
                                fh1.Definition  || CHAR(10) || fh1.Alternatives, '')
                    END                                                        AS Help,
                    CAST(fld.FRPO AS INTEGER)                                  AS sort_key
                FROM {guide_table} d
                JOIN cmifld fld
                    ON  fld.MINM = d.API
                    AND fld.TRNM = d.TransactionName
                    AND fld.TRTP = 'I'
                    AND fld.FLNM = d.FieldName
                LEFT JOIN m3TableCols tc
                    ON  tc.TableName  = substr(d.TableName, 1, 6)
                    AND tc.ColumnName = d.ColumnName
                LEFT JOIN m3FieldHelp fh2 ON fh2.FieldHelpID = tc.ColumnName
                LEFT JOIN m3FieldHelp fh1 ON fh1.FieldHelpID = SUBSTR(tc.ColumnName, -4)
                WHERE lower(d.required) NOT IN ('no')
                  AND d.API             = ?
                  AND d.TransactionName = ?

                UNION ALL

                -- Mandatory cmifld fields missing from the mapping guide
                SELECT
                    fld.FLNM                                                   AS Row1,
                    fld.FLDS || ' (' || fld.TYPE || ':' || fld.LENG || ')'    AS Row2,
                    'Yes'                                                      AS Row3,
                    ''                                                         AS Row4,
                    ''                                                         AS Row5,
                    CASE
                        WHEN fh2.Definition IS NOT NULL THEN
                            COALESCE(
                                fh2.HeadingText || CHAR(10)||CHAR(10) ||
                                fh2.Definition  || CHAR(10) || fh2.Alternatives, '')
                        ELSE
                            COALESCE(
                                fh1.HeadingText || CHAR(10)||CHAR(10) ||
                                fh1.Definition  || CHAR(10) || fh1.Alternatives, '')
                    END                                                        AS Help,
                    CAST(fld.FRPO AS INTEGER)                                  AS sort_key
                FROM cmifld fld
                -- No mapping guide row available; look up help directly on the field name
                LEFT JOIN m3FieldHelp fh2 ON fh2.FieldHelpID = fld.FLNM
                LEFT JOIN m3FieldHelp fh1 ON fh1.FieldHelpID = SUBSTR(fld.FLNM, -4)
                WHERE fld.MINM = ?
                  AND fld.TRNM = ?
                  AND fld.TRTP = 'I'
                  AND fld.MAND = 1
                  AND NOT EXISTS (
                        SELECT 1
                        FROM {guide_table} d2
                        WHERE d2.API             = ?
                          AND d2.TransactionName = ?
                          AND d2.FieldName       = fld.FLNM
                          AND lower(d2.required) NOT IN ('no')
                  )
            )
            ORDER BY sort_key
            """

            detail_rows = cursor.execute(detail_sql, (api, trn, api, trn, api, trn)).fetchall()
            if not detail_rows:
                continue

            ws = wb.create_sheet(title=worksheet_name)

            # ====================================================
            # Declarative headers
            # ====================================================
            HEADER_DEFINITION = [
                {"label": "MESSAGE", "index": 0, "fill": "ADD8E6", "freeze": True},  # row 1–3 light blue
                {"label": "Message", "index": 1, "fill": "ADD8E6", "freeze": True},
                {"label": "no", "index": 2, "fill": "ADD8E6", "freeze": True},
                {"label": "Mapping", "index": 3, "fill": "D3D3D3", "font": Font(bold=True), "align": "center"},  # row 4 gray
                {"label": "Example", "index": 4, "font": Font(italic=True), "align": "center"}  # row 5 example
            ]
            HELP_INDEX = 5

            # Append headers
            for header in HEADER_DEFINITION:
                ws.append([header["label"]] + [r[header["index"]] for r in detail_rows])

            # Freeze panes automatically
            frozen_rows = sum(1 for h in HEADER_DEFINITION if h.get("freeze"))
            ws.freeze_panes = ws.cell(row=frozen_rows + 1, column=3)

            # Apply styles to all columns including column A
            for row_idx, header in enumerate(HEADER_DEFINITION, start=1):
                for row in ws.iter_rows(
                    min_row=row_idx, max_row=row_idx, min_col=1, max_col=ws.max_column
                ):
                    for cell in row:
                        # Fill
                        if "fill" in header:
                            cell.fill = PatternFill(
                                start_color=header["fill"],
                                end_color=header["fill"],
                                fill_type="solid"
                            )
                        # Font
                        if "font" in header:
                            cell.font = header["font"]
                        # Alignment
                        align = header.get("align", "left")
                        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

            # Add comments from Help column
            for col_idx, note in enumerate([r[HELP_INDEX] for r in detail_rows], start=2):
                add_comment(ws, 1, col_idx, note)

            autofit_ws(ws)

        output_path = os.path.join(output_dir, f"{wb_name}.xlsx")
        wb.save(output_path)
        logger.info(f"Saved: {output_path}")

    conn.close()
    logger.info("All workbooks created successfully.")

# ============================================================
# CLI ENTRY POINT
# ============================================================

def export_templates():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    guide_table = prompt_for_guide(cursor)
    conn.close()

    export_templates_for_guide(guide_table)


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%H:%M:%S"
    )
    export_templates()
