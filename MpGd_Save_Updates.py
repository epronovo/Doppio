# MpGd_Save_Updates.py
# -----------------------------------------------------------------------
# PURPOSE
#   Monitors the mapping/update/ folder for Excel (.xlsx) mapping guide
#   files and upserts their content into dynamically-named GUIDE_ tables
#   in SQLite.  Target table names and column groupings are inferred from
#   a section-marker convention in row 2 of the spreadsheet: a header
#   named "SectionName_N" signals the end of an N-column section that maps
#   to GUIDE_SectionName.  A six-part composite key (Sequence, TableName,
#   ColumnName, API, TransactionName, FieldName) governs conflict handling;
#   existing values are preserved when incoming data is NULL (COALESCE).
#   Processed files are moved to mapping/processed/ on completion.
#
# INPUTS
#   - Excel (.xlsx) files in mapping/update/ — rows 1+ as data, row 2 as
#     column headers with section markers
#
# OUTPUTS
#   - GUIDE_{Section} tables (SQLite) — one table per section found in the
#     spreadsheet header
#   - Processed files moved to mapping/processed/
#
# DEPENDENCIES
#   - config.py, openpyxl
#
# USAGE
#   python MpGd_Save_Updates.py
# -----------------------------------------------------------------------

import logging
import os
import re
import sqlite3
import shutil
from openpyxl import load_workbook
from config import BASE_DIR, get_sqlite_db_path

logger = logging.getLogger(__name__)

# ============================================================
# CONFIG
# ============================================================
DB_PATH = get_sqlite_db_path()
UPDATE_DIR = BASE_DIR / "mapping/update"
PROCESSED_DIR = BASE_DIR / "mapping/processed"

os.makedirs(UPDATE_DIR, exist_ok=True)
os.makedirs(PROCESSED_DIR, exist_ok=True)

SECTION_RE = re.compile(r"(.+?)_(\d+)$")

# ============================================================
# HELPERS
# ============================================================
def normalize_column_name(col):
    """
    Strip prefix before first underscore.
    DG_required -> required
    """
    if isinstance(col, str) and "_" in col:
        return col.split("_", 1)[1]
    return col

def make_unique(names):
    """
    Ensure DB column names are unique
    """
    seen = {}
    out = []

    for name in names:
        base = name
        if base not in seen:
            seen[base] = 0
            out.append(base)
        else:
            seen[base] += 1
            new = f"{base}_{seen[base]}"
            while new in seen:
                seen[base] += 1
                new = f"{base}_{seen[base]}"
            seen[new] = 0
            out.append(new)

    return out

# ============================================================
# SECTION DISCOVERY (RAW HEADERS ONLY)
# ============================================================
def discover_sections(raw_headers):
    if "Sequence" not in raw_headers:
        raise ValueError("Missing required 'Sequence' column")

    # Define the keys we are already creating manually in the table
    reserved_keys = {"Sequence", "TableName", "ColumnName", "API", "TransactionName", "FieldName"}

    sections = []
    start_idx = raw_headers.index("Sequence") + 1
    buffer = []

    for h in raw_headers[start_idx:]:
        buffer.append(h)
        match = SECTION_RE.match(str(h))
        
        if match:
            base, count = match.groups()
            count = int(count)

            if count > len(buffer) - 1:
                cols_raw = buffer[:-1]
            else:
                cols_raw = buffer[-(count + 1):-1]

            # FILTER: Only include normalized names if they aren't in our reserved_keys
            cols_norm = make_unique([
                normalize_column_name(c) for c in cols_raw 
                if normalize_column_name(c) not in reserved_keys
            ])

            # Adjust cols_raw to match the filtered cols_norm
            # We only keep the raw columns that survived the normalization filter
            filtered_cols_raw = [
                c for c in cols_raw if normalize_column_name(c) not in reserved_keys
            ]

            sections.append({
                "table": f"GUIDE_{base.replace(' ', '')}",
                "columns_raw": filtered_cols_raw,
                "columns_norm": cols_norm
            })
            buffer = []

    return sections

# ============================================================
# DB SETUP
# ============================================================
def ensure_tables_exist(conn, sections):
    cur = conn.cursor()
    for s in sections:
        col_defs = ", ".join([f"[{c}] TEXT" for c in s["columns_norm"]])
        
        # We define the 6-part Composite Primary Key here
        sql = f"""
            CREATE TABLE IF NOT EXISTS {s["table"]} (
                Sequence INTEGER,
                TableName TEXT,
                ColumnName TEXT,
                API TEXT,
                TransactionName TEXT,
                FieldName TEXT,
                {col_defs},
                PRIMARY KEY (Sequence, TableName, ColumnName, API, TransactionName, FieldName)
            )
        """
        cur.execute(sql)
    conn.commit()
    
# ============================================================
# UPSERT
# ============================================================
def upsert_section(cur, section, keys, row, header_idx):
    cols_raw = section["columns_raw"]
    cols_norm = section["columns_norm"]

    key_values = [
        keys['Sequence'], keys['TableName'], keys['ColumnName'], 
        keys['API'], keys['TransactionName'], keys['FieldName']
    ]
    
    data_values = []
    for raw in cols_raw:
        idx = header_idx.get(raw)
        data_values.append(row[idx].value if idx is not None else None)

    # REMOVE OR MODIFY THE 'ANY' CHECK
    # We want to proceed even if data_values is [None] or [""] 
    # because the row itself (defined by the keys) is important.
    
    all_values = key_values + data_values
    placeholders = ", ".join(["?"] * len(all_values))
    
    # We use COALESCE to ensure we don't overwrite existing data with NULL 
    # unless that's specifically what you want.
    updates = ", ".join([f"[{c}] = COALESCE(excluded.[{c}], [{c}])" for c in cols_norm])

    sql = f"""
        INSERT INTO {section["table"]} (
            Sequence, TableName, ColumnName, API, TransactionName, FieldName, 
            {", ".join([f"[{c}]" for c in cols_norm])}
        )
        VALUES ({placeholders})
        ON CONFLICT(Sequence, TableName, ColumnName, API, TransactionName, FieldName)
        DO UPDATE SET {updates}
    """
    cur.execute(sql, all_values)
        
# ============================================================
# FILE PROCESSING
# ============================================================
def process_mapping_file(path):
    logger.info(f"Processing: {path}")

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    raw_headers = [cell.value for cell in ws[2]]
    header_idx = {h: i for i, h in enumerate(raw_headers)}

    sections = discover_sections(raw_headers)

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    ensure_tables_exist(conn, sections)

    # for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    #     sequence = row[header_idx["Sequence"]].value
    #     if not sequence:
    #         continue

    #     for section in sections:
    #         upsert_section(cur, section, sequence, row, header_idx)
    
    # Inside the row loop in process_mapping_file:
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        # Create a dictionary of your 6 keys
        keys = {
            "Sequence": row[header_idx["Sequence"]].value,
            "TableName": row[header_idx["TableName"]].value,
            "ColumnName": row[header_idx["ColumnName"]].value,
            "API": row[header_idx["API"]].value,
            "TransactionName": row[header_idx["TransactionName"]].value,
            "FieldName": row[header_idx["FieldName"]].value
        }

        if not keys["Sequence"]:
            continue

        for section in sections:
            upsert_section(cur, section, keys, row, header_idx)

    conn.commit()
    conn.close()

    logger.info(f"Finished updating database from: {path}")

# ============================================================
# FOLDER MONITOR
# ============================================================
def monitor_update_folder():
    files = [f for f in os.listdir(UPDATE_DIR) if f.lower().endswith(".xlsx")]

    if not files:
        logger.info("No files to process.")
        return

    for f in files:
        src = UPDATE_DIR / f
        process_mapping_file(src)

        dst = PROCESSED_DIR / f
        shutil.move(src, dst)
        logger.info(f"Moved to processed: {dst}")

    logger.info("All mapping updates applied.")

# ============================================================
# ENTRY
# ============================================================
if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%H:%M:%S"
    )
    monitor_update_folder()
