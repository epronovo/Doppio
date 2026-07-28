# MpGd_PopulateSheets.py
"""
MpGd_PopulateSheets.py
---------------------
Monitors the input folder for Excel workbooks and populates their data-migration
sheets with records fetched from an M3 SOURCE tenant via EXPORTMI.Select.

Workbook layout expected (per tab, starting at tab 3):
  Row 3           : Table definition  –  "XXXXXX: Description"
                    (6-char table name followed by a colon)
                    OR "CSYTAB_XXXX: Description" for System Tables
  Burgundy row    : 4-char field names (column headers, burgundy fill)
  Green row       : Field length / type row (data begins on the NEXT row)
  Data rows       : Existing records (preserved) + missing ones added here

Steps
-----
1.  Prompt for SOURCE tenant (always prompts – no saved default re-use).
2.  Monitor the input/ folder for .xlsx / .xlsm workbooks.
3.  For each workbook, skip sheets 1 & 2, then for each remaining sheet:
      a.  Parse table name from the "XXXXXX: Description" row.
      b.  Locate the burgundy header row → extract 4-char field names + columns.
      c.  Locate the green length row → data starts on the next row.

      Regular tables (e.g. MITARE):
      d.  Query doppio.db (m3TableCols) for the 2-char column prefix.
      e.  Build the EXPORTMI QERY (prefix + 4-char → 6-char field names).
      f.  Call EXPORTMI.Select; key on first field for deduplication.

      CSYTAB tables (e.g. CSYTAB_MODL):
      d.  Extract the STCO code from the suffix (e.g. 'MODL').
      e.  Run the fixed CSYTAB QERY filtering on CTSTCO = <stco>.
      f.  Parse CTPARM into additional fields using doppio.db m3DataStructures
          (DataStructure = 'sDS<stco>'); key on STKY for deduplication.

      g.  Append missing rows as italic text strings.
4.  Save the workbook in place.

Usage:
    python MpGd_PopulateSheets.py
"""

from copy import copy
import os
import re
import sqlite3
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from InforMI import CONFIG, get_ion_token
from UserDefaults import load_user_defaults, save_user_defaults
from config import get_sqlite_db_path

# =============================================================================
# Constants
# =============================================================================

EXPORTMI_SEP = "^"
INPUT_FOLDER = Path(__file__).parent / "input"
DOPPIO_DB    = get_sqlite_db_path()   # ~/sqlite/doppio.db via config.py

# =============================================================================
# Color helpers
# =============================================================================

def _cell_rgb(cell) -> str:
    """
    Return the foreground fill colour of *cell* as a 6-char hex string (RRGGBB).
    Returns an empty string when the colour cannot be determined.

    openpyxl stores colours as 8-char ARGB (e.g. 'FF800000') when the type is
    'rgb'.  Theme/indexed colours are not converted – they are left as ''.
    """
    try:
        fg = cell.fill.fgColor
        if fg.type == "rgb" and fg.rgb and fg.rgb != "00000000":
            argb = fg.rgb  # 8 chars: AARRGGBB
            return argb[-6:]  # drop alpha → RRGGBB
    except Exception:
        pass
    return ""


def _is_burgundy(rgb6: str) -> bool:
    """
    True when the RRGGBB hex string represents a burgundy / dark-red shade.
    Covers the two most common M3 template burgundy values:
      800000 (standard dark red) and C00000 (accent-2 dark red).
    Heuristic: R ≥ 100, G < 80, B < 80.
    """
    if len(rgb6) != 6:
        return False
    try:
        r = int(rgb6[0:2], 16)
        g = int(rgb6[2:4], 16)
        b = int(rgb6[4:6], 16)
        return r >= 100 and g < 80 and b < 80
    except ValueError:
        return False


def _is_green(rgb6: str) -> bool:
    """
    True when the RRGGBB hex string represents a green shade.
    Heuristic: G is dominant (G > R and G > B) and G ≥ 80.
    Covers Excel greens like 00B050, 92D050, 70AD47, etc.
    """
    if len(rgb6) != 6:
        return False
    try:
        r = int(rgb6[0:2], 16)
        g = int(rgb6[2:4], 16)
        b = int(rgb6[4:6], 16)
        return g > r and g > b and g >= 80
    except ValueError:
        return False


# =============================================================================
# Tenant setup – always prompts
# =============================================================================

def _snapshot() -> dict:
    return dict(CONFIG)


def _restore(snapshot: dict) -> None:
    CONFIG.clear()
    CONFIG.update(snapshot)


def _pick_ionapi(ionapi_dir: Path, label: str) -> Path:
    """Display a numbered list of .ionapi files and return the chosen path."""
    files = sorted(
        [f for f in os.listdir(ionapi_dir) if f.endswith(".ionapi")],
        key=str.lower,
    )
    if not files:
        raise FileNotFoundError(f"No .ionapi files found in: {ionapi_dir}")

    print(f"\n  🔹  Select the {label} ION API file:")
    for i, f in enumerate(files, start=1):
        print(f"    {i}. {f}")

    while True:
        try:
            choice = int(input(f"\n  Enter choice (1-{len(files)}): "))
            if 1 <= choice <= len(files):
                return ionapi_dir / files[choice - 1]
            print("  ❌  Invalid choice. Try again.")
        except ValueError:
            print("  ❌  Please enter a number.")


def _prompt_company_division(label: str) -> None:
    """Prompt for company and division, saving the defaults."""
    defaults        = load_user_defaults()
    default_company  = defaults.get("company",  "100")
    default_division = defaults.get("division", "")

    print(f"\n  Configure {label} company / division:")
    raw_company  = input(f"    Company  (default: {default_company}):  ")
    raw_division = input(f"    Division (default: {repr(default_division)}): ")

    CONFIG["company"]  = raw_company.strip() if raw_company.strip() else default_company
    # Preserve spaces – a blank/space division is valid in M3.
    # Empty input (just Enter) means keep the default; anything else (including spaces) is used as-is.
    CONFIG["division"] = raw_division if raw_division != "" else default_division

    defaults["company"]  = CONFIG["company"]
    defaults["division"] = CONFIG["division"]
    save_user_defaults(defaults)

def bold_existing_records(ws, data_start_row: int, key_col: int, col_indices: list[int]) -> int:
    """
    Scan rows from data_start_row onward and apply bold font to the tracked
    columns for any row that already contains a key value.
    """
    bolded_count = 0
    for row_idx in range(data_start_row, ws.max_row + 1):
        val = str(ws.cell(row=row_idx, column=key_col).value or "").strip()
        if val:
            for col_idx in col_indices:
                cell = ws.cell(row=row_idx, column=col_idx)
                # Safely copy existing font properties (like color/size) and set bold
                new_font = copy(cell.font) if cell.font else Font()
                new_font.bold = True
                cell.font = new_font
            bolded_count += 1
    return bolded_count

def remove_italicized_rows(ws, data_start_row: int, key_col: int) -> int:
    """
    Iterate backwards from the bottom of the sheet to data_start_row.
    Deletes any row where the key column's cell is formatted as italic.
    Iterating backwards prevents row indices from shifting during deletion.
    """
    deleted_count = 0
    
    for row_idx in range(ws.max_row, data_start_row - 1, -1):
        cell = ws.cell(row=row_idx, column=key_col)
        # Check if the font object exists and has the italic property set to True
        if cell.font and cell.font.italic:
            ws.delete_rows(row_idx, 1)
            deleted_count += 1
            
    return deleted_count

def autofit_columns(ws) -> None:
    """
    Iterate over all columns with data in the worksheet, calculate the maximum 
    string length of the cell contents, and set the column width to fit.
    """
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        
        for cell in ws[col_letter]:
            if cell.value:
                # Calculate length of the string representation of the cell value
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
                    
        if max_length > 0:
            # Add a little padding (e.g., +2) to ensure it doesn't look cramped
            ws.column_dimensions[col_letter].width = max_length + 2
            
def setup_source_tenant(ionapi_dir: Path) -> None:
    """Configure the SOURCE tenant.  Always prompts – no saved-session shortcuts."""
    label = "SOURCE"
    print(f"\n{'─' * 60}")
    print(f"  📡  Configure {label} tenant")
    print(f"{'─' * 60}")

    CONFIG["tenant"] = _pick_ionapi(ionapi_dir, label)
    _prompt_company_division(label)
    get_ion_token()

    print(
        f"\n  ✅  {label} ready: {Path(CONFIG['tenant']).name}  "
        f"CONO={CONFIG.get('company', '')}  DIVI={CONFIG.get('division', '')}"
    )


# =============================================================================
# SQLite: resolve 2-char column prefix for a given M3 table
# =============================================================================

def get_column_prefix(table_name: str) -> str:
    """
    Queries doppio.db:
        SELECT DISTINCT substr(ColumnName,1,2) prefix
        FROM   m3TableCols
        WHERE  TableName = ?

    Returns the first prefix found, or '' if the table/db is not available.
    The prefix is the 2-char identifier prepended to every column name in the
    EXPORTMI QERY (e.g. 'MN' → MNWHLO, MNSLTP …).
    """
    if not DOPPIO_DB.exists():
        print(f"  ⚠️   doppio.db not found at {DOPPIO_DB}")
        return ""
    try:
        conn = sqlite3.connect(str(DOPPIO_DB))
        cur  = conn.cursor()
        cur.execute(
            "SELECT DISTINCT substr(ColumnName,1,2) prefix "
            "FROM   m3TableCols "
            "WHERE  TableName = ?",
            (table_name,),
        )
        rows = cur.fetchall()
        conn.close()
        if rows:
            return rows[0][0]
        print(f"  ⚠️   No column prefix found for table: {table_name}")
        return ""
    except sqlite3.Error as exc:
        print(f"  ❌  SQLite error while looking up prefix: {exc}")
        return ""


# =============================================================================
# SQLite: CSYTAB PARM structure
# =============================================================================

def get_parm_structure(stco_code: str) -> list[dict]:
    """
    Returns the PARM field layout for a given CSYTAB STCO code by querying
    doppio.db:

        SELECT FieldName, Length, dsFrom, dsTo
        FROM   m3DataStructures
        WHERE  DataStructure = 'sDS<stco_code>'
        ORDER  BY dsFrom

    Returns a list of dicts with keys: FieldName, Length, dsFrom, dsTo.
    dsFrom / dsTo are 1-based character positions within the CTPARM string.
    Returns an empty list if the table/structure is unavailable.
    """
    ds_name = f"sDS{stco_code}"
    if not DOPPIO_DB.exists():
        print(f"  ⚠️   doppio.db not found at {DOPPIO_DB}")
        return []
    try:
        conn = sqlite3.connect(str(DOPPIO_DB))
        cur  = conn.cursor()
        cur.execute(
            "SELECT FieldName, Length, dsFrom, dsTo "
            "FROM   m3DataStructures "
            "WHERE  DataStructure = ? "
            "ORDER  BY dsFrom",
            (ds_name,),
        )
        rows = cur.fetchall()
        conn.close()
        if not rows:
            print(f"  ⚠️   No PARM structure found for DataStructure '{ds_name}'.")
            return []
        result = [
            {"FieldName": r[0], "Length": r[1], "dsFrom": r[2], "dsTo": r[3]}
            for r in rows
        ]
        print(f"  🗂️   PARM structure '{ds_name}': {len(result)} field(s)")
        return result
    except sqlite3.Error as exc:
        print(f"  ❌  SQLite error reading PARM structure: {exc}")
        return []


def fetch_csytab(
    session   : requests.Session,
    stco_code : str,
) -> list[dict]:
    """
    Fetches CSYTAB records for a specific STCO code via EXPORTMI.Select, then
    parses the CTPARM field into individual sub-fields using the data structure
    defined in doppio.db (m3DataStructures).

    EXPORTMI QERY used:
        CTDIVI,CTSTCO,CTSTKY,CTLNCD,CTTX40,CTTX15,CTPARM from CSYTAB
        where CTSTCO = <stco_code>

    CTPARM is parsed according to the 'sDS<stco_code>' data structure.

    The returned records use 4-char short-name keys (DIVI, STCO, STKY, LNCD,
    TX40, TX15, plus the parsed PARM sub-fields).  CTPARM is not included.

    Returns [] on failure or when no records are found.
    """
    label    = f"CSYTAB/{stco_code}"
    division = CONFIG.get("division", "")
    if division.strip():
        qery = (
            f"CTDIVI,CTSTCO,CTSTKY,CTLNCD,CTTX40,CTTX15,CTPARM "
            f"from CSYTAB where CTSTCO = {stco_code} and CTDIVI = {division}"
        )
    else:
        qery = (
            f"CTDIVI,CTSTCO,CTSTKY,CTLNCD,CTTX40,CTTX15,CTPARM "
            f"from CSYTAB where CTSTCO = {stco_code}"
        )
    print(f"    🔍  CSYTAB QERY: {qery}")

    # ── Fetch raw rows ────────────────────────────────────────────────────────
    raw_records = fetch_exportmi(session, qery, label=label)
    if not raw_records:
        return []

    # ── Look up PARM structure ────────────────────────────────────────────────
    parm_fields = get_parm_structure(stco_code)  # [{FieldName, dsFrom, dsTo}, …]

    # Mapping from the 6-char CT-prefixed EXPORTMI key → 4-char short name
    CT_FIELD_MAP = {
        "CTDIVI": "DIVI",
        "CTSTCO": "STCO",
        "CTSTKY": "STKY",
        "CTLNCD": "LNCD",
        "CTTX40": "TX40",
        "CTTX15": "TX15",
    }

    normalised: list[dict] = []
    for raw in raw_records:
        rec: dict[str, str] = {}

        # Map direct fields (CT-prefix → 4-char)
        for ct_key, short_key in CT_FIELD_MAP.items():
            rec[short_key] = str(raw.get(ct_key, "")).strip()

        # Parse CTPARM into sub-fields
        parm = raw.get("CTPARM", "")
        for pf in parm_fields:
            fn       = pf["FieldName"]
            ds_from  = int(pf["dsFrom"])
            ds_to    = int(pf["dsTo"])
            
            # FIX 1: Strip the 2-char prefix from DB field name (YMTEL1 -> TEL1)
            # so it maps correctly to the 4-char Excel column headers.
            short_fn = fn[2:] if len(fn) >= 6 else fn
            
            # FIX 2: Use standard 0-based Python slicing. The db is already 0-based.
            # Python naturally handles strings that are shorter than the slice index.
            val = parm[ds_from:ds_to]
            rec[short_fn] = val.strip()

        normalised.append(rec)

    return normalised

# =============================================================================
# EXPORTMI.Select → list[dict]
# =============================================================================

def fetch_exportmi(
    session: requests.Session,
    query: str,
    label: str,
) -> list[dict]:
    """
    Calls EXPORTMI.Select with the given *query* string and returns a list of
    dicts keyed by the column names found in the HDRS row.

    Handles 401 token refresh transparently.
    """
    list_url = CONFIG["api_url"]

    payload = {
        "program": "EXPORTMI",
        "transactions": [
            {
                "transaction": "Select",
                "record": {
                    "QERY": query,
                    "SEPC": EXPORTMI_SEP,
                    "HDRS": "1",
                },
                "selectedColumns": ["QERY", "SEPC", "HDRS", "REPL"],
            }
        ],
    }

    auth_headers = {
        "Authorization": f"Bearer {CONFIG['access_token']}",
        "Content-Type":  "application/json",
    }

    resp = session.post(list_url, json=payload, headers=auth_headers)
    if resp.status_code == 401:
        get_ion_token()
        auth_headers["Authorization"] = f"Bearer {CONFIG['access_token']}"
        resp = session.post(list_url, json=payload, headers=auth_headers)

    resp.raise_for_status()
    data = resp.json()

    repl_rows: list[str] = []
    for result in data.get("results", []):
        for record in result.get("records", []):
            repl_val = record.get("REPL", "")
            if repl_val:
                repl_rows.append(repl_val)

    if not repl_rows:
        print(f"  ⚠️   No rows returned from EXPORTMI ({label}).")
        return []

    # First REPL row is the column-name header (HDRS:1 was requested)
    col_names = [
        c.strip()
        for c in repl_rows[0].rstrip(EXPORTMI_SEP).split(EXPORTMI_SEP)
    ]

    parsed: list[dict] = []
    for raw in repl_rows[1:]:
        values = raw.rstrip(EXPORTMI_SEP).split(EXPORTMI_SEP)
        values += [""] * (len(col_names) - len(values))
        row = {col_names[i]: values[i].strip() for i in range(len(col_names))}
        parsed.append(row)

    print(f"  ✅  {len(parsed):,} record(s) retrieved from M3 ({label}).")
    return parsed


# =============================================================================
# Sheet inspection helpers
# =============================================================================

def find_table_definition(ws) -> tuple[int, str, str]:
    """
    Scan the worksheet for a cell matching one of two patterns:

      Regular  : 'XXXXXX: Description'          (6 uppercase letters)
      CSYTAB   : 'CSYTAB_XXXX: Description'     (CSYTAB + underscore + STCO code)

    Searches row 3 first (the documented position), then falls back to scanning
    all rows in case of layout variations.

    Returns (row_1based, table_name, description) or (0, '', '').
    Table name is returned as-is (e.g. 'MITARE' or 'CSYTAB_MODL').
    """
    # Matches "XXXXXX:" OR "CSYTAB_XXXX:" at the start of the cell value.
    pattern = re.compile(r"^([A-Z]{6}(?:_[A-Z0-9]+)?):\s*(.*)$")

    def _check_row(row_idx: int) -> tuple[int, str, str]:
        for cell in ws[row_idx]:
            val = str(cell.value or "").strip()
            m   = pattern.match(val)
            if m:
                return row_idx, m.group(1), m.group(2).strip()
        return 0, "", ""

    # Preferred location: row 3
    if ws.max_row >= 3:
        result = _check_row(3)
        if result[0]:
            return result

    # Fallback: scan all rows
    for row_idx in range(1, ws.max_row + 1):
        if row_idx == 3:
            continue  # already checked
        result = _check_row(row_idx)
        if result[0]:
            return result

    return 0, "", ""


def find_field_header_row(
    ws,
    after_row: int,
) -> tuple[int, list[str], list[int]]:
    """
    Scan rows from *after_row + 1* onward for the field-name header row.

    Detection rule: a row where EVERY non-empty cell contains a value that is
    3–4 characters, uppercase, and alphanumeric (A-Z, 0-9), with a minimum of
    2 qualifying cells.

    Using isalnum() instead of isalpha() allows field names that contain digits
    such as TX15 and TX40.  isupper() blocks pure-number strings like '10' and
    '15' (they have no cased letters so Python's isupper() returns False), as
    well as mixed-case description rows.  The length range 3–4 excludes single
    type-code characters like 'A' and 'N'.

    Always checks from row 6 (the fixed field-header position).

    Returns (row_1based, [field_names], [col_indices_1based]).
    Returns (0, [], []) if no matching row is found.
    """
    for row_idx in range(6, ws.max_row + 1):
        non_empty: list[tuple[str, int]] = []   # (value, col_index)

        for cell in ws[row_idx]:
            raw = str(cell.value or "").strip()
            if raw:                              # skip truly empty cells
                non_empty.append((raw, cell.column))

        if len(non_empty) < 2:
            continue  # too sparse — definitely not the field row

        # FIX: Changed v.isalpha() to v.isalnum() to allow numbers in the field names
        if all(3 <= len(v) <= 4 and v.isalnum() and v.isupper() for v, _ in non_empty):
            fields   = [v   for v, _ in non_empty]
            col_idxs = [col for _, col in non_empty]
            return row_idx, fields, col_idxs

    return 0, [], []


def find_green_length_row(ws, after_row: int) -> int:
    """
    Scan rows from *after_row + 1* onward for the first row that contains at
    least one cell with a green fill.  Returns the 1-based row index, or 0.
    """
    for row_idx in range(after_row + 1, ws.max_row + 1):
        for cell in ws[row_idx]:
            if _is_green(_cell_rgb(cell)):
                return row_idx
    return 0


def collect_existing_keys(ws, data_start_row: int, key_col: int) -> set[str]:
    """
    Return the set of non-empty string values in *key_col* starting at
    *data_start_row*.  Used to identify records already present in the sheet.
    """
    keys: set[str] = set()
    for row_idx in range(data_start_row, ws.max_row + 1):
        val = str(ws.cell(row=row_idx, column=key_col).value or "").strip()
        if val:
            keys.add(val)
    return keys


# =============================================================================
# Writing missing rows
# =============================================================================

def append_missing_rows(
    ws,
    missing_records : list[dict],
    field_names_4   : list[str],   # 4-char names (sheet column headers)
    col_indices     : list[int],   # 1-based column positions
    field_map_4to6  : dict[str, str],  # 4-char → 6-char (EXPORTMI key)
    data_start_row  : int,
) -> int:
    """
    Append *missing_records* to the worksheet, one per row, immediately after
    the last existing data row.

    Every cell is written as:
      • italic font
      • text number format (@) so numeric-looking values stay as strings

    Returns the count of rows written.
    """
    # Find the last row that has any data in the tracked columns
    last_data_row = data_start_row - 1
    for row_idx in range(data_start_row, ws.max_row + 1):
        if any(ws.cell(row=row_idx, column=c).value for c in col_indices):
            last_data_row = row_idx

    next_write = last_data_row + 1
    written    = 0

    for rec in missing_records:
        for field_4, col_idx in zip(field_names_4, col_indices):
            field_6 = field_map_4to6.get(field_4, "")
            val     = str(rec.get(field_6, "")).strip()
            cell    = ws.cell(row=next_write, column=col_idx)
            cell.value         = val
            cell.font          = Font(italic=True)
            cell.number_format = "@"   # store as text
        next_write += 1
        written    += 1

    return written


# =============================================================================
# Single-workbook processor
# =============================================================================

def process_workbook(wb_path: Path, session: requests.Session) -> None:
    """
    Open *wb_path*, skip the first two sheets, then for every subsequent sheet:
      1. Locate the table definition, burgundy headers, and green length row.
      2. Fetch all records from M3 via EXPORTMI.
      3. Append any records missing from the sheet (italic, text format).
    Saves the workbook in place if any changes were made.
    """
    print(f"\n{'─' * 60}")
    print(f"  📗  Processing workbook: {wb_path.name}")
    print(f"{'─' * 60}")

    wb     = openpyxl.load_workbook(str(wb_path))
    sheets = wb.sheetnames

    if len(sheets) < 3:
        print(
            f"  ⚠️   Workbook has only {len(sheets)} sheet(s) "
            f"(need at least 3). Skipping."
        )
        return

    modified = False

    for sheet_name in sheets[2:]:          # tabs 1 & 2 skipped
        ws = wb[sheet_name]
        print(f"\n  📄  Sheet: '{sheet_name}'")

        # ── 1. Table definition ───────────────────────────────────────────
        tbl_row, tbl_name, tbl_desc = find_table_definition(ws)
        if not tbl_name:
            print("    ⚠️   No 'XXXXXX: Description' row found. Skipping sheet.")
            continue
        print(f"    📋  Table : {tbl_name}  –  {tbl_desc}")

        # ── 2. Field header row (all cells exactly 4 uppercase alpha chars) ──
        hdr_row, field_names_4, col_indices = find_field_header_row(ws, tbl_row)
        if not field_names_4:
            print("    ⚠️   No 4-char field header row found. Skipping sheet.")
            continue
        print(f"    🔵  Fields ({len(field_names_4)}): {' '.join(field_names_4)}")

        # ── 3. Fixed layout: field row 6 → length row 8 → data from row 9 ───
        length_row     = 8
        data_start_row = 9
        print(f"    📏  Length row at row {length_row}; data from row {data_start_row}")

        # ── Freeze pane at A9 (keeps header rows visible while scrolling) ──
        ws.freeze_panes = "A9"
        modified = True  # freeze pane change counts as a modification

        # ── 4-7.  Route: CSYTAB vs regular table ─────────────────────────────
        is_csytab = tbl_name.startswith("CSYTAB_")

        if is_csytab:
            # ------------------------------------------------------------------
            # CSYTAB path
            # ------------------------------------------------------------------
            stco_code = tbl_name[7:]  # e.g. 'CSYTAB_MODL' → 'MODL'
            print(f"    🗃️   CSYTAB detected – STCO code: '{stco_code}'")

            try:
                m3_records = fetch_csytab(session, stco_code)
            except Exception as exc:
                print(f"    ❌  CSYTAB fetch failed: {exc}  – Skipping sheet.")
                continue

            if not m3_records:
                print(f"    ℹ️   No records returned from CSYTAB/{stco_code}.")
                continue

            # Records already use 4-char keys → identity map
            field_map_4to6 = {f: f for f in field_names_4}

            # For CSYTAB the STCO value is the same for every row (it IS the
            # filter); use STKY (the second burgundy field) as the unique key.
            # Fallback: if 'STKY' is not in the header, use the second field.
            if "STKY" in field_names_4:
                key_field_4 = "STKY"
                key_col_idx = col_indices[field_names_4.index("STKY")]
            elif len(field_names_4) > 1:
                key_field_4 = field_names_4[1]
                key_col_idx = col_indices[1]
            else:
                key_field_4 = field_names_4[0]
                key_col_idx = col_indices[0]

        else:
            # ------------------------------------------------------------------
            # Regular table path
            # ------------------------------------------------------------------

            # 4. Column prefix from doppio.db
            prefix = get_column_prefix(tbl_name)
            if not prefix:
                print(f"    ⚠️   No column prefix found for {tbl_name}. Skipping sheet.")
                continue
            print(f"    🔑  Column prefix: '{prefix}'")

            # 5. Build QERY: prefix + 4-char → 6-char EXPORTMI field names
            field_map_4to6 = {f: (prefix + f) for f in field_names_4}
            qery = (
                ",".join(field_map_4to6[f] for f in field_names_4)
                + f" from {tbl_name}"
            )
            print(f"    🔍  QERY: {qery}")

            # 6. Call EXPORTMI.Select
            try:
                m3_records = fetch_exportmi(session, qery, label=tbl_name)
            except Exception as exc:
                print(f"    ❌  EXPORTMI call failed: {exc}  – Skipping sheet.")
                continue

            if not m3_records:
                print(f"    ℹ️   No records returned from M3 for {tbl_name}.")
                continue

            # Key = first field
            key_field_4 = field_names_4[0]
            key_col_idx = col_indices[0]

        # ── 6b. Clean up previously added italicized rows ─────────────────────
        sheet_modified = False  # Start tracking modifications here

        deleted_italics = remove_italicized_rows(ws, data_start_row, key_col_idx)
        if deleted_italics > 0:
            print(f"    🧹  Removed {deleted_italics} previously added italicized row(s).")
            sheet_modified = True

        # ── 7. Identify missing records & format existing ones ────────────────
        existing_keys = collect_existing_keys(ws, data_start_row, key_col_idx)
        print(
            f"    📊  Existing: {len(existing_keys)} record(s)  |  "
            f"M3: {len(m3_records)} record(s)"
        )

        # Apply bold to existing records
        if existing_keys:
            bolded = bold_existing_records(ws, data_start_row, key_col_idx, col_indices)
            if bolded > 0:
                sheet_modified = True
                print(f"    🖋️   Bolded {bolded} existing row(s).")

        lookup_key = field_map_4to6.get(key_field_4, key_field_4)
        missing = [
            r for r in m3_records
            if str(r.get(lookup_key, "")).strip() not in existing_keys
        ]
        print(f"    ➕  Missing records to add: {len(missing)}")

        # ── 8. Write missing rows & Autofit ───────────────────────────────
        if missing:
            written = append_missing_rows(
                ws, missing, field_names_4, col_indices, field_map_4to6, data_start_row
            )
            print(f"    ✅  Added {written} row(s) in italic text format.")
            sheet_modified = True

        if sheet_modified:
            autofit_columns(ws)
            print("    ↔️   Auto-fitted column widths.")
            modified = True  # Tells the overall script to save the workbook at the very end
        elif not missing:
            print("    ✅  All M3 records are already present and no formatting changed.")

    # ── 9. Save the workbook (ADD THIS BLOCK BACK IN) ─────────────────────────
    if modified:
        wb.save(str(wb_path))
        print(f"\n  💾  Workbook saved: {wb_path.name}")
    else:
        print(f"\n  ℹ️   No changes were needed – workbook not modified.")

# =============================================================================
# Input-folder monitor
# =============================================================================

def process_input_folder(session: requests.Session) -> None:
    """
    Processes all .xlsx / .xlsm workbooks currently in INPUT_FOLDER once, then exits.
    """
    INPUT_FOLDER.mkdir(parents=True, exist_ok=True)

    print(f"\n{'═' * 60}")
    print(f"  📁  Input folder: {INPUT_FOLDER}")
    print(f"{'═' * 60}")

    workbooks = sorted(
        f for f in INPUT_FOLDER.iterdir()
        if f.suffix.lower() in (".xlsx", ".xlsm")
    )

    if not workbooks:
        print(f"  ℹ️   No Excel workbooks found in input folder. Exiting.")
        return

    print(f"  📋  Found {len(workbooks)} workbook(s) to process.")

    for wb_path in workbooks:
        try:
            process_workbook(wb_path, session)
        except Exception as exc:
            print(f"  ❌  Unexpected error processing {wb_path.name}: {exc}")


# =============================================================================
# Entry point
# =============================================================================

def populate_sheets() -> None:
    ionapi_dir = Path(__file__).parent / "ionapi"

    print(f"\n{'═' * 60}")
    print(f"  MIG_PopulateSheets")
    print(f"{'═' * 60}")

    # Always prompt for SOURCE tenant
    setup_source_tenant(ionapi_dir)

    with requests.Session() as session:
        process_input_folder(session)


if __name__ == "__main__":
    populate_sheets()
