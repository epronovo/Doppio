import os
import shutil
from config import BASE_DIR
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# donwnload all the price lists from the sharepoint into the PriceLists folder
# and then run this script to process them into an Add file
# after the file is created the processed files will be moved to the ToProcess folder

# Constants
FOLDER_PATH = BASE_DIR / "PriceLists"
OUTPUT_FOLDER = BASE_DIR / "evs100/ToProcess"
SHEET_NAME = "API_OIS017MI_AddBasePrice"
CONTROL_SHEET_NAME = "Control"
DATA_SHEET_NAME = "API_OIS017MI_AddPriceList"

results = []

# Ensure output folder exists
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Process files
for filename in os.listdir(FOLDER_PATH):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(FOLDER_PATH, filename)

        try:
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
            if SHEET_NAME not in workbook.sheetnames:
                workbook.close()
                continue

            sheet = workbook[SHEET_NAME]
            total_rows = sheet.max_row

            # Get values from row 5, columns B, C, D
            for i, row in enumerate(sheet.iter_rows(min_row=5, max_row=5, min_col=2, max_col=4, values_only=True), start=5):
                b5, c5, d5 = row
                results.append({
                    "PRRF": b5,
                    "CUCD": c5,
                    "FVDT": d5
                })
                break

            workbook.close()

        except Exception as e:
            print(f"Error reading {filename}: {e}")

# Create new workbook
wb = Workbook()
ws = wb.active
ws.title = DATA_SHEET_NAME

# Define columns (header row)
columns = ["MESSAGE", "PRRF", "CUCD", "FVDT", "LVDT", "TX40", "CRTP", "TX15", "SCMU", "PCTP", "SCMO", "WHLO"]

# === Row 1: Header ===
ws.append(columns)

# === Row 2: Blank row ===
ws.append(["" for _ in columns])

# === Row 3: 'yes' under all except MESSAGE ===
yes_row = ["no" if col == "MESSAGE" else "yes" for col in columns]
ws.append(yes_row)

# === Row 4 onward: Data ===
for entry in results:
    prrf = entry["PRRF"]
    cucd = entry["CUCD"]
    fvdt = entry["FVDT"]
    ws.append([
        "",                        # MESSAGE
        prrf,
        cucd,
        fvdt,
        "20251231",                # LVDT
        f"{prrf} Price List",      # TX40
        "1",                       # CRTP
        f"{prrf} Price List",      # TX15
        "1",                       # SCMU
        "3",                       # PCTP
        "COST",                    # SCMO
        "US1"                      # WHLO
    ])

# --- Control Sheet ---
ws_control = wb.create_sheet(CONTROL_SHEET_NAME)
wb.move_sheet(ws_control, offset=-len(wb.sheetnames) + 1)

sheet_name_cleaned = DATA_SHEET_NAME

ws_control.cell(row=1, column=1, value="Worksheet")
ws_control.cell(row=1, column=2, value="Description")
ws_control.cell(row=1, column=3, value="Data")
ws_control.cell(row=2, column=1, value=sheet_name_cleaned)
ws_control.cell(row=2, column=2, value=sheet_name_cleaned)
ws_control.cell(row=2, column=3, value="x")

# Auto-size columns
for col_idx in range(1, 4):
    col_letter = get_column_letter(col_idx)
    ws_control.column_dimensions[col_letter].width = 20

# Save the workbook
output_path = os.path.join(OUTPUT_FOLDER, DATA_SHEET_NAME + ".xlsx")
wb.save(output_path)
print(f"Workbook saved to: {output_path}")

# Move all processed .xlsx files from PriceLists to ToProcess
for filename in os.listdir(FOLDER_PATH):
    if filename.endswith(".xlsx"):
        src_path = os.path.join(FOLDER_PATH, filename)
        dst_path = os.path.join(OUTPUT_FOLDER, filename)
        try:
            shutil.move(src_path, dst_path)
            print(f"Moved: {filename}")
            # Delete the file from the source folder if it still exists
            if os.path.exists(src_path):
                os.remove(src_path)
                print(f"Deleted: {filename} from source folder")
        except Exception as e:
            print(f"Failed to move {filename}: {e}")

