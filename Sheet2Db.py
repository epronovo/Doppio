# Sheet2Db.py
# -----------------------------------------------------------------------
# PURPOSE
#   Scans the input folder for .xlsm / .xlsx workbooks.
#   For every sheet that contains API batch data (identified by the
#   presence of an API name in A2, transaction in G4, environment in I2,
#   and column headers starting at B8), the data rows (B9 onwards) are
#   loaded into a temporary SQLite database so Sql2Api.py can process
#   them.
#
#   SHEET LAYOUT CONVENTION
#     A2  → API / program name  (e.g. "CRS610MI")
#     G4  → Transaction name    (e.g. "ChgOrderInfo")
#     B8+ → Parameter column headers
#     B9+ → Data rows
#
#   TABLE NAMING
#     Each sheet gets its own table: API_{api}_{transaction}
#     e.g.  API_CRS610MI_ChgOrderInfo
#     The table is DROPPED and recreated on every run.
#
#   COLUMNS IN EACH TABLE
#     _api    — API / program name from A2
#     _trnm   — transaction name from G4
#     <col>…  — one column per header found in row 8
#
#   COMPATIBLE WITH Sql2Api.py
#     Use the generated SELECT (printed at the end) which returns:
#       col[0] = minm   (_api)
#       col[1] = trnm   (_trnm)
#       col[2+]= field values
#
# INPUTS
#   ./input/*.xlsm  ./input/*.xlsx
#
# OUTPUTS
#   ~/sqlite/sheet2db_temp.db   (default; override with --db)
#
# USAGE
#   python Sheet2Db.py [--input-folder PATH] [--db PATH] [--verbose]
# -----------------------------------------------------------------------

import argparse
import glob
import logging
import os
import re
import sqlite3
import sys
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
SCRIPT_DIR           = os.path.dirname(os.path.abspath(__file__))
DEFAULT_INPUT_FOLDER = os.path.join(SCRIPT_DIR, "input")
SQLITE_DIR           = Path.home() / "sqlite"
DEFAULT_DB_PATH      = str(SQLITE_DIR / "doppio.db")

# Utility sheet names that are never treated as API data sheets
_SKIP_SHEETS = {
    "master", "settings", "logos", "versions", "help",
    "environments", "transactions", "availablemis", "control",
}

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _sanitize(name: str) -> str:
    """Convert an arbitrary string into a valid SQLite identifier."""
    name = str(name).strip()
    # Replace spaces and special chars with underscores
    name = re.sub(r"[^A-Za-z0-9_]", "_", name)
    # Collapse multiple underscores
    name = re.sub(r"_+", "_", name).strip("_")
    # Must not start with a digit
    if name and name[0].isdigit():
        name = "T_" + name
    return name or "col"


def _cell(ws, row: int, col: int):
    """Return stripped string value of a cell, or empty string."""
    row_data = next(ws.iter_rows(min_row=row, max_row=row, min_col=col, max_col=col, values_only=True), (None,))
    v = row_data[0] if row_data else None
    return str(v).strip() if v is not None else ""


def _read_headers(ws) -> list[str]:
    """Read column headers from row 8 starting at column B (col 2)."""
    row8 = next(ws.iter_rows(min_row=8, max_row=8, min_col=2, values_only=True), ())
    headers = []
    for v in row8:
        v = str(v).strip() if v is not None else ""
        if not v:
            break
        headers.append(v)
    return headers


def _read_data_rows(ws, num_cols: int) -> list[list]:
    """Read data rows from row 9 onwards; stop at first fully empty row."""
    rows = []
    for raw in ws.iter_rows(min_row=9, min_col=2, max_col=1 + num_cols, values_only=True):
        values = [str(v).strip() if v is not None else "" for v in raw]
        if all(v == "" for v in values):
            break
        rows.append(values)
    return rows


def _parse_api_sheet(ws, sheet_name: str):
    """Return (api, trnm, headers) if this is an API data sheet, else None."""
    if sheet_name.lower() in _SKIP_SHEETS:
        return None
    api  = _cell(ws, 2, 1)   # A2
    trnm = _cell(ws, 4, 7)   # G4
    if not api or not trnm:
        return None
    headers = _read_headers(ws)
    return (api, trnm, headers) if headers else None


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def _ensure_db_dir(db_path: str) -> None:
    d = os.path.dirname(db_path)
    if d:
        os.makedirs(d, exist_ok=True)


def _table_name(api: str, trnm: str) -> str:
    return f"API_{_sanitize(api)}_{_sanitize(trnm)}"


def load_sheet_to_db(conn: sqlite3.Connection,
                     api: str,
                     trnm: str,
                     headers: list[str],
                     rows: list[list],
                     source_file: str,
                     sheet_name: str) -> str:
    """
    Drop (if exists) and recreate a table for this api/transaction,
    then insert all rows.  Returns the table name used.
    """
    tbl = _table_name(api, trnm)

    # Sanitize column names (keep originals for logging)
    safe_headers = [_sanitize(h) for h in headers]

    # Deduplicate column names (append _2, _3 … if needed)
    seen: dict[str, int] = {}
    unique_headers = []
    for h in safe_headers:
        if h in seen:
            seen[h] += 1
            unique_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 1
            unique_headers.append(h)

    # Drop existing table
    conn.execute(f'DROP TABLE IF EXISTS "{tbl}"')

    # Build CREATE TABLE
    meta_cols  = '"_api" TEXT, "_trnm" TEXT, "_source_file" TEXT, "_sheet_name" TEXT'
    data_cols  = ", ".join(f'"{h}" TEXT' for h in unique_headers)
    conn.execute(f'CREATE TABLE "{tbl}" ({meta_cols}, {data_cols})')

    # Insert rows
    placeholders = ", ".join(["?"] * (4 + len(unique_headers)))
    meta = [api, trnm, os.path.basename(source_file), sheet_name]
    conn.executemany(
        f'INSERT INTO "{tbl}" VALUES ({placeholders})',
        (meta + row for row in rows),
    )
    conn.commit()
    return tbl


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------

def process_workbook(wb_path: str, conn: sqlite3.Connection) -> list[dict]:
    """
    Open one workbook, detect API sheets, load each into the db.
    Returns list of metadata dicts for each table created.
    """
    log.info(f"Opening: {os.path.basename(wb_path)}")
    wb = openpyxl.load_workbook(wb_path, read_only=True, keep_vba=True,
                                data_only=True)
    tables_created = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        parsed = _parse_api_sheet(ws, sheet_name)
        if parsed is None:
            log.debug(f"  Skipping sheet: {sheet_name}")
            continue

        api, trnm, headers = parsed
        rows = _read_data_rows(ws, len(headers))

        log.info(
            f"  Sheet '{sheet_name}' → {api}/{trnm}  "
            f"cols={len(headers)}  rows={len(rows)}"
        )

        if not rows:
            log.warning(f"    No data rows found — skipping.")
            continue

        tbl = load_sheet_to_db(
            conn, api, trnm, headers, rows,
            wb_path, sheet_name,
        )
        log.info(f"    → table '{tbl}'  ({len(rows)} rows)")

        tables_created.append({
            "table":  tbl,
            "api":    api,
            "trnm":   trnm,
            "cols":   headers,
            "rows":   len(rows),
            "sheet":  sheet_name,
            "source": os.path.basename(wb_path),
        })

    wb.close()
    return tables_created


def build_sql2api_query(meta: dict) -> str:
    """
    Generate a SELECT that Sql2Api.py can use directly.
    col[0]=minm, col[1]=trnm, col[2+]=field values.
    Empty-string fields are excluded by Sql2Api itself.
    """
    tbl      = meta["table"]
    api      = meta["api"].replace("'", "''")
    trnm     = meta["trnm"].replace("'", "''")
    col_list = ", ".join(f'"{_sanitize(c)}"' for c in meta["cols"])
    return (
        f"SELECT '{api}' AS minm, '{trnm}' AS trnm, {col_list}\n"
        f'FROM "{tbl}"'
    )


def process(input_folder: str, db_path: str) -> None:
    patterns = [
        os.path.join(input_folder, "*.xlsm"),
        os.path.join(input_folder, "*.xlsx"),
    ]
    files = []
    for pat in patterns:
        files.extend(glob.glob(pat))
    # Skip temp/lock files (names starting with ~$)
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    files = sorted(files)

    if not files:
        log.error(f"No .xlsm / .xlsx files found in: {input_folder}")
        sys.exit(1)

    log.info(f"Found {len(files)} workbook(s) in: {input_folder}")

    _ensure_db_dir(db_path)
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA synchronous=OFF")

    all_tables: list[dict] = []
    for wb_path in files:
        tables = process_workbook(wb_path, conn)
        all_tables.extend(tables)

    conn.close()

    if not all_tables:
        log.warning("No API data sheets found across all workbooks.")
        sys.exit(0)

    log.info("")
    log.info(f"Temp database : {db_path}")
    log.info(f"Tables created: {len(all_tables)}")
    log.info("")
    log.info("=" * 60)
    log.info("Sql2Api.py compatible SELECT statements")
    log.info("=" * 60)
    for meta in all_tables:
        q = build_sql2api_query(meta)
        log.info(f"\n-- {meta['source']}  |  sheet: {meta['sheet']}  |  rows: {meta['rows']}")
        log.info(q)
    log.info("")
    log.info(
        "Run with:\n"
        f"  python Sql2Api.py --sql \"<paste SELECT above>\""
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Load Excel API batch sheets into a temp SQLite db "
            "for processing by Sql2Api.py"
        )
    )
    parser.add_argument(
        "--input-folder", "-i",
        default=DEFAULT_INPUT_FOLDER,
        help=f"Folder containing *.xlsm / *.xlsx files (default: {DEFAULT_INPUT_FOLDER})",
    )
    parser.add_argument(
        "--db",
        default=DEFAULT_DB_PATH,
        help=f"Path for the temp SQLite database (default: {DEFAULT_DB_PATH})",
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable DEBUG logging",
    )
    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    log.info("=" * 60)
    log.info("Sheet2Db  —  Excel API sheets → SQLite temp db")
    log.info("=" * 60)

    if not os.path.isdir(args.input_folder):
        log.error(f"Input folder not found: {args.input_folder}")
        sys.exit(1)

    process(args.input_folder, args.db)
    log.info("Done.")


if __name__ == "__main__":
    main()
