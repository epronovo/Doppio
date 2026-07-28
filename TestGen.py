import sqlite3
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime

def build_api_excel(db_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Pull all relevant rows
    query = """
        SELECT DISTINCT API,replace(OBNM,' ','_')[API_Name],MIN([Transaction])[Transaction],ShortName,Description,case when Data_Type = 'Decimal' then 'N' else 'A' end AS DataType,Length 
        FROM Final_Mapping_Guide 
        LEFT JOIN CMIPGM ON MINM = API
        WHERE (API='MMS200MI')
        GROUP BY API,API_Name,ShortName,Description,DataType,Length
        ORDER BY 1,2,3
    """
    rows = cursor.execute(query).fetchall()
    conn.close()

    if not rows:
        print("No results found.")
        return

    # Extract primary info for filename
    API, API_Name = rows[0][0], rows[0][1]
    file_name = f"{API}_{API_Name}_{datetime.now().strftime('%y%m%d')}.xlsx"

    # Create workbook
    wb = Workbook()
    # Remove default sheet if needed
    if "Sheet" in wb.sheetnames:
        std = wb["Sheet"]
        wb.remove(std)

    # Group rows by Transaction
    from collections import defaultdict
    transactions = defaultdict(list)
    for r in rows:
        transactions[r[2]].append(r)  # key = Transaction

    # Light blue fill
    blue60 = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    for transaction, trans_rows in transactions.items():
        sheet_name = f"API_{API}_{transaction}"
        ws = wb.create_sheet(title=sheet_name)

        # Row 1: MESSAGE + ShortNames
        row1 = ["MESSAGE"] + [r[3] for r in trans_rows]
        # Row 2: blank + Description(DataType:Length)
        row2 = [""] + [f"{r[4]} ({r[5]}:{r[6]})" for r in trans_rows]
        # Row 3: no + yes
        row3 = ["no"] + ["yes"] * len(trans_rows)

        ws.append(row1)
        ws.append(row2)
        ws.append(row3)

        # Apply light blue to columns 2 → N
        num_cols = len(row1)
        for col_idx in range(2, num_cols + 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            for row in range(1, 4):
                ws[f"{col_letter}{row}"].fill = blue60

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_name)
    print(f"Workbook created: {file_name}")


# ---- RUN ROUTINE ----
build_api_excel("/Users/ericpronovost/sqlite/doppio.db")