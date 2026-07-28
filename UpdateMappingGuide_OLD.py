# UpdateMappingGuide.py
import os
import sqlite3
import shutil
from openpyxl import load_workbook
from config import BASE_DIR


DB_PATH = "/Users/ericpronovost/sqlite/doppio.db"
UPDATE_DIR = BASE_DIR / "mapping/update"      # folder to monitor
PROCESSED_DIR = BASE_DIR / "mapping/processed"  # folder for completed files

# Ensure folders exist
os.makedirs(UPDATE_DIR, exist_ok=True)
os.makedirs(PROCESSED_DIR, exist_ok=True)


def ensure_table_exists():
    """Create HerffJonesGuide table if missing."""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS HerffJonesGuide (
            Sequence INTEGER,
            MappingNotes TEXT,
            Responsible TEXT,
            Required TEXT,
            DefaultValue TEXT,
            PRIMARY KEY (Sequence)
        )
    """)

    conn.commit()
    conn.close()


def process_mapping_file(path):
    """Read a MappingGuide Excel file and update DB rows."""
    print(f"Processing: {path}")

    wb = load_workbook(path, data_only=True)
    ws = wb.active  # MappingGuide sheet

    # Locate required columns
    header_row = {cell.value: idx+1 for idx, cell in enumerate(ws[2])}

    required_columns = [
        "Sequence", "MappingNotes", "Responsible", "Required", "DefaultValue"
    ]

    for col in required_columns:
        if col not in header_row:
            print(f"ERROR: Column '{col}' missing in file: {path}")
            return

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # Process data rows (row 3 and down)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        sequence = row[header_row["Sequence"] - 1].value

        # Skip blank key rows
        if not sequence:
            continue

        # Extract editable fields
        mapping_notes = row[header_row["MappingNotes"] - 1].value
        responsible = row[header_row["Responsible"] - 1].value
        required = row[header_row["Required"] - 1].value
        default_value = row[header_row["DefaultValue"] - 1].value

        # Skip update if ALL editable fields are blank
        if not any([mapping_notes, responsible, required, default_value]):
            continue

        # UPSERT the row
        cur.execute("""
            INSERT INTO HerffJonesGuide
                (Sequence, MappingNotes, Responsible, Required, DefaultValue)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(Sequence)
            DO UPDATE SET
                MappingNotes = COALESCE(excluded.MappingNotes, MappingNotes),
                Responsible = COALESCE(excluded.Responsible, Responsible),
                Required = COALESCE(excluded.Required, Required),
                DefaultValue = COALESCE(excluded.DefaultValue, DefaultValue)
        """, (sequence, mapping_notes, responsible, required, default_value))

    conn.commit()
    conn.close()

    print(f"Finished updating database from: {path}")


def monitor_update_folder():
    """Process all .xlsx files placed in UPDATE_DIR."""
    ensure_table_exists()

    files = [f for f in os.listdir(UPDATE_DIR) if f.lower().endswith(".xlsx")]

    if not files:
        print("No files to process.")
        return

    for f in files:
        full_path = os.path.join(UPDATE_DIR, f)

        process_mapping_file(full_path)

        # Move to processed folder
        new_path = PROCESSED_DIR / f
        shutil.move(full_path, new_path)

        print(f"Moved to processed folder: {new_path}")

    print("All mapping updates applied.")


# Run the monitor
if __name__ == "__main__":
    monitor_update_folder()
