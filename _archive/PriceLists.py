import os
from openpyxl import load_workbook
import openpyxl

# Folder containing XLSX files
FOLDER_PATH = r"C:\\ASRaymond\\PriceLists"
SHEET_NAME = "API_OIS017MI_AddBasePrice"

results = []

for filename in os.listdir(FOLDER_PATH):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(FOLDER_PATH, filename)

        try:
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
            if SHEET_NAME not in workbook.sheetnames:
                continue

            sheet = workbook[SHEET_NAME]
            total_rows = sheet.max_row

            for i, row in enumerate(sheet.iter_rows(min_row=5, max_row=5, min_col=2, max_col=4, values_only=True), start=5):
                b5, c5, d5 = row
                results.append({
                    "filename": filename,
                    "B5": b5,
                    "C5": c5,
                    "D5": d5,
                    "total_rows": total_rows
                })
                break

        except Exception as e:
            print(f"Error reading {filename}: {e}")

# for entry in results:
#     print(entry)

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AddPriceList"

# Write data to the worksheet
# Write header
ws.append(["filename", "PRRF", "CUCD", "FVDT", "total_rows"])
# Write the extracted results
for entry in results:
    ws.append([entry["filename"], entry["B5"], entry["C5"], entry["D5"], entry["total_rows"]])

# Save the workbook
wb.save(os.path.join(r"C:\ASRaymond", "AddPriceList.xlsx"))