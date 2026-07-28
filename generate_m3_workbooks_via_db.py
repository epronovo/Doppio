"""
M3 Configuration Workbook Generator
====================================
Generates Excel workbooks for each Area in the M3 system with DataStructure support.

Requirements:
- pandas
- openpyxl
- sqlite3

Input Files:
- claude.db (SQLite database containing all M3 data)
- HJ_M3_Template.xlsx (master template)

Output:
- One Excel workbook per Area containing:
  1. [Area] Config sheet with program list and hyperlinks
  2. ColourCode sheet with formatting reference
  3. Program sheets with table columns and DataStructure fields
"""

import pandas as pd
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.utils import get_column_letter
import os

# Get the directory where this script is located
base_path = os.path.dirname(os.path.abspath(__file__))
DB_PATH = "/Users/ericpronovost/sqlite/doppio.db"

def get_path(filename):
    """Helper function to get full path for a file in the script directory"""
    return os.path.join(base_path, filename)

# Connect to SQLite database
print("Connecting to SQLite database...")
# conn = sqlite3.connect(get_path('claude.db'))
conn = sqlite3.connect(DB_PATH)

# Read data from database tables
print("Loading data from database...")
extract_rules = pd.read_sql_query("SELECT * FROM ExtractRules", conn)
m3_programs = pd.read_sql_query("SELECT * FROM m3Programs", conn)
m3_table_cols = pd.read_sql_query("SELECT * FROM m3TableCols", conn)
m3_data_structures = pd.read_sql_query("SELECT * FROM m3DataStructures", conn)
m3_reference_fields = pd.read_sql_query("SELECT * FROM m3ReferenceFields", conn)
m3_tables = pd.read_sql_query("SELECT * FROM m3Tables", conn)

conn.close()

# Load master template
master_wb = load_workbook(get_path('HJ_M3_Template.xlsx'))
master_color_sheet = master_wb['ColourCode']

# Check which config sheet exists in template
if 'Config' in master_wb.sheetnames:
    master_config_sheet = master_wb['Config']
elif 'Distribution Config' in master_wb.sheetnames:
    master_config_sheet = master_wb['Distribution Config']
else:
    raise ValueError("No Config or Distribution Config sheet found in template")

# Extract cell styles from ColourCode sheet (rows 3-6) with FONT SIZE 12
cell_styles = {}
for row in range(3, 7):
    src = master_color_sheet[f'A{row}']
    cell_styles[row] = {
        'fill': PatternFill(
            start_color=src.fill.start_color, 
            end_color=src.fill.end_color, 
            fill_type=src.fill.fill_type
        ) if src.fill else None,
        'font': Font(
            name=src.font.name, 
            size=12,  # Force size 12 for all program sheet cells
            bold=src.font.bold, 
            italic=src.font.italic, 
            color=src.font.color
        ) if src.font else None,
        'alignment': Alignment(
            horizontal=src.alignment.horizontal, 
            vertical=src.alignment.vertical, 
            wrap_text=src.alignment.wrap_text
        ) if src.alignment else None,
        'border': Border(
            left=src.border.left, 
            right=src.border.right, 
            top=src.border.top, 
            bottom=src.border.bottom
        ) if src.border else None
    }

# Merge ExtractRules with m3Programs and m3Tables to get descriptions
merged = extract_rules.merge(m3_programs, left_on='Pgm', right_on='programName', how='left')
merged['Tbl_Short'] = merged['Tbl'].str[:6]

# Join with m3Tables to get table descriptions
merged = merged.merge(
    m3_tables,
    left_on='Tbl_Short',
    right_on='tableName',
    how='left'
)

# Extract tableDescription (chars 5-40)
merged['tableDescription_Short'] = merged['tableDescription'].apply(
    lambda x: str(x)[4:44] if pd.notna(x) else ''
)

query_result = merged[['Area', 'Pgm', 'programDescription', 'Tbl_Short', 'DataStructure', 'Tbl', 'tableDescription_Short']].copy()
query_result.rename(columns={'Tbl': 'tableLongName'}, inplace=True)

# Get unique areas
areas = [str(a) for a in extract_rules['Area'].unique() if pd.notna(a)]

# Create output directory
os.makedirs(get_path('output'), exist_ok=True)

# Generate one workbook per area
for area in areas:
    print(f"\nProcessing area: {area}")
    
    # Filter data for this area (exact match, preserving case)
    area_data = query_result[query_result['Area'].astype(str) == area].copy()
    
    # Create new workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # ===== SHEET 1: Config Sheet =====
    config_sheet_name = f"{area} Config"
    config_sheet = wb.create_sheet(config_sheet_name, 0)
    
    # Copy all formatting from master config sheet
    for row in master_config_sheet.iter_rows():
        for cell in row:
            new_cell = config_sheet.cell(row=cell.row, column=cell.column)
            if cell.has_style:
                new_cell.font = Font(
                    name=cell.font.name, 
                    size=cell.font.size, 
                    bold=cell.font.bold, 
                    italic=cell.font.italic, 
                    color=cell.font.color
                )
                new_cell.fill = PatternFill(
                    start_color=cell.fill.start_color, 
                    end_color=cell.fill.end_color, 
                    fill_type=cell.fill.fill_type
                )
                new_cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal, 
                    vertical=cell.alignment.vertical, 
                    wrap_text=cell.alignment.wrap_text
                )
                new_cell.border = Border(
                    left=cell.border.left, 
                    right=cell.border.right, 
                    top=cell.border.top, 
                    bottom=cell.border.bottom
                )

    # Set config title (merged A2:D2)
    config_sheet.merge_cells('A2:D2')
    title_cell = config_sheet['A2']
    title_cell.value = f"{area.upper()} CONFIGURATION TEMPLATE"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    
    # Set config headers (Row 4)
    headers = ['Sr. Nr.', 'Program', 'M3 Program Code']
    for i, header in enumerate(headers, 1):
        cell = config_sheet.cell(row=4, column=i)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Get unique programs for this area
    unique_programs = area_data.drop_duplicates(subset=['Pgm']).reset_index(drop=True)
    sheet_map = {}  # Maps program name to sanitized sheet name

    # ===== PROGRAM SHEETS: Create one sheet per program =====
    for _, row in unique_programs.iterrows():
        pgm = str(row['Pgm'])
        
        # Sanitize sheet name (max 31 chars, no special characters)
        sanitized_name = pgm.replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_')[:31]
        
        try:
            pgm_sheet = wb.create_sheet(sanitized_name)
        except:
            # If name collision, use fallback
            sanitized_name = f"Pgm_{_}"
            pgm_sheet = wb.create_sheet(sanitized_name)
        
        sheet_map[pgm] = sanitized_name
        
        # Row 1: Back link to config sheet
        pgm_sheet['A1'] = f"← Back to {area} Config"
        pgm_sheet['A1'].hyperlink = f"#'{config_sheet_name}'!A1"
        pgm_sheet['A1'].font = Font(color="0563C1", underline="single", bold=True, size=12)
        
        # Row 2: Program title
        pgm_sheet['A2'] = f"{pgm}: {str(row['programDescription']).upper()}"
        pgm_sheet['A2'].font = Font(bold=True, size=12)
        
        # Row 3: Table name and description
        table_long_name = row['tableLongName'] if pd.notna(row['tableLongName']) else ''
        table_description = row['tableDescription_Short'] if pd.notna(row['tableDescription_Short']) else ''
        
        if table_long_name:
            pgm_sheet['A3'] = f"{table_long_name}: {table_description}"
            pgm_sheet['A3'].font = Font(size=12)
        
        # Row 4: Blank

        tbl = row['Tbl_Short']
        data_structure = row['DataStructure']
        
        # Initialize combined columns list
        all_columns = []
        
        # --- GET TABLE COLUMNS ---
        if pd.notna(tbl):
            # Query table columns from m3TableCols
            table_cols = m3_table_cols[m3_table_cols['TableName'] == tbl].copy()
            
            # Apply standard filters to exclude system columns
            table_cols = table_cols[
                ~table_cols['ColumnName'].str.contains('CONO', na=False) &
                ~table_cols['ColumnName'].str.contains('DIVI', na=False) &
                ~table_cols['ColumnName'].str.contains('TXID', na=False) &
                ~table_cols['ColumnName'].str.contains('PRTX', na=False) &
                ~table_cols['ColumnName'].str.contains('RGDT', na=False) &
                ~table_cols['ColumnName'].str.contains('RGTM', na=False) &
                ~table_cols['ColumnName'].str.contains('LMDT', na=False) &
                ~table_cols['ColumnName'].str.contains('CHNO', na=False) &
                ~table_cols['ColumnName'].str.contains('CHID', na=False) &
                ~table_cols['ColumnName'].str.contains('LMTS', na=False)
            ]
            
            # If DataStructure exists, also filter out PARM, DTID, and MGSQ
            # (These fields are replaced by DataStructure data)
            if pd.notna(data_structure) and str(data_structure).strip() != '':
                table_cols = table_cols[
                    ~table_cols['ColumnName'].str.contains('PARM', na=False) &
                    ~table_cols['ColumnName'].str.contains('DTID', na=False) &
                    ~table_cols['ColumnName'].str.contains('MGSQ', na=False)
                ]
            
            # Sort by sequence and prepare column data
            table_cols = table_cols.sort_values('Sequence')
            table_cols['ColumnName_Short'] = table_cols['ColumnName'].str[2:6]  # Extract chars 3-6
            table_cols['DataType_Code'] = table_cols['DataType'].apply(
                lambda x: 'A' if x == 'String' else 'N' if x == 'Decimal' else ''
            )
            
            # Add table columns to combined list
            for _, col_row in table_cols.iterrows():
                all_columns.append({
                    'Description': col_row['Description'],
                    'ColumnName': col_row['ColumnName_Short'],
                    'DataType': col_row['DataType_Code'],
                    'Length': col_row['Length']
                })
        
        # --- GET DATASTRUCTURE COLUMNS ---
        if pd.notna(data_structure) and str(data_structure).strip() != '':
            # Query DataStructure fields
            ds_fields = m3_data_structures[
                m3_data_structures['DataStructure'] == str(data_structure)
            ].copy()
            
            if len(ds_fields) > 0:
                # Extract short field name (chars 3-6) for joining
                ds_fields['FieldName_Short'] = ds_fields['FieldName'].str[2:6]
                
                # Join with reference fields to get descriptions and metadata
                merged_ds = ds_fields.merge(
                    m3_reference_fields,
                    left_on='FieldName_Short',
                    right_on='referenceFieldName',
                    how='left'
                )
                
                # Sort by dsFrom (field position in structure)
                merged_ds = merged_ds.sort_values('dsFrom')
                
                # Add DataStructure columns to combined list
                for _, ds_row in merged_ds.iterrows():
                    # Extract first character of category for DataType
                    data_type = ''
                    if pd.notna(ds_row.get('referenceFieldCategory')):
                        data_type = str(ds_row['referenceFieldCategory'])[0:1]
                    
                    all_columns.append({
                        'Description': ds_row.get('referenceFieldDescription', ''),
                        'ColumnName': ds_row.get('referenceFieldName', ''),
                        'DataType': data_type,
                        'Length': ds_row.get('Length', '')
                    })
        
        # --- WRITE COLUMNS TO SHEET ---
        if all_columns:
            column_widths = {}
            
            for col_idx, col_data in enumerate(all_columns, start=1):
                # Data mapping: rows 5-8 contain Description, ColumnName, DataType, Length
                data_map = {
                    5: col_data['Description'],
                    6: col_data['ColumnName'],
                    7: col_data['DataType'],
                    8: col_data['Length']
                }
                
                # Apply styles from template (rows 3-6) to workbook rows (5-8)
                for workbook_row in range(5, 9):
                    template_style_row = workbook_row - 2  # Row 5→Style 3, Row 6→Style 4, etc.
                    cell = pgm_sheet.cell(row=workbook_row, column=col_idx)
                    val = data_map[workbook_row]
                    cell.value = val
                    
                    # Apply template styles
                    style = cell_styles[template_style_row]
                    if style['fill']: cell.fill = style['fill']
                    if style['font']: cell.font = style['font']
                    if style['alignment']: cell.alignment = style['alignment']
                    if style['border']: cell.border = style['border']
                    
                    # Track max width for this column
                    length = len(str(val)) if val else 0
                    column_widths[col_idx] = max(column_widths.get(col_idx, 0), length)

            # Set dynamic column widths (with max of 50)
            for col_idx, max_len in column_widths.items():
                adjusted_width = (max_len + 2) * 1.2
                pgm_sheet.column_dimensions[get_column_letter(col_idx)].width = min(adjusted_width, 50)

    # ===== POPULATE CONFIG SHEET WITH PROGRAM LINKS =====
    for idx, row in unique_programs.iterrows():
        row_num = idx + 5  # Data starts at row 5
        pgm_name = str(row['Pgm'])
        
        # Column A: Serial number
        config_sheet.cell(row=row_num, column=1).value = idx + 1
        
        # Column B: Program description
        config_sheet.cell(row=row_num, column=2).value = row['programDescription'] if pd.notna(row['programDescription']) else ''
        
        # Column C: Program code with hyperlink
        cell_link = config_sheet.cell(row=row_num, column=3)
        cell_link.value = pgm_name
        if pgm_name in sheet_map:
            cell_link.hyperlink = f"#'{sheet_map[pgm_name]}'!A1"
            cell_link.font = Font(color="0563C1", underline="single")

    # Copy column widths from template
    for col_name, col_dim in master_config_sheet.column_dimensions.items():
        config_sheet.column_dimensions[col_name].width = col_dim.width

    # ===== SHEET 2: ColourCode Reference Sheet =====
    color_code_sheet = wb.create_sheet('ColourCode', 1)
    
    # Copy all content and formatting from template
    for row in master_color_sheet.iter_rows():
        for cell in row:
            new_cell = color_code_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = Font(
                    name=cell.font.name, 
                    size=cell.font.size, 
                    bold=cell.font.bold, 
                    color=cell.font.color
                )
                new_cell.fill = PatternFill(
                    start_color=cell.fill.start_color, 
                    end_color=cell.fill.end_color, 
                    fill_type=cell.fill.fill_type
                )
    
    # Save workbook
    output_file = get_path(f"output/{area} Configuration.xlsx")
    wb.save(output_file)
    print(f"  ✓ Created: {area} Configuration.xlsx")

print("\n✅ All workbooks created successfully!")
print(f"   Output directory: {get_path('output')}")