from pathlib import Path
from openpyxl import load_workbook

# -----------------------------------------------------------------------------
# Folder setup
# -----------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
EVS100_DIR = BASE_DIR / "evs100"
TO_PREPARE_DIR = EVS100_DIR / "ToPrepare"
TO_PROCESS_DIR = EVS100_DIR / "ToProcess"

TO_PREPARE_DIR.mkdir(parents=True, exist_ok=True)
TO_PROCESS_DIR.mkdir(parents=True, exist_ok=True)


# -----------------------------------------------------------------------------
# Helper function
# -----------------------------------------------------------------------------
def prepare_evs100_workbook(file_path: Path):
    filename_no_ext = file_path.stem

    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames

    # Use the first sheet as the main data sheet
    original_sheet_name = sheet_names[0]
    ws_data = wb[original_sheet_name]

    # Rename first sheet to match filename
    new_sheet_name = filename_no_ext
    ws_data.title = new_sheet_name

    # Insert MESSAGE column at the first position
    ws_data.insert_cols(1)
    ws_data.cell(row=1, column=1, value="MESSAGE")  # header cell

    # Add two new rows at the top (shift existing data down)
    ws_data.insert_rows(2, 2)  # inserts two empty rows at positions 2 and 3

    # Fill third row: "no" in first column, "yes" in all other columns
    max_col = ws_data.max_column
    ws_data.cell(row=3, column=1, value="no")
    for col in range(2, max_col + 1):
        ws_data.cell(row=3, column=col, value="yes")

    # Remove Control sheet if it exists
    if "Control" in wb.sheetnames:
        del wb["Control"]

    # Add Control sheet at the beginning
    ws_control = wb.create_sheet("Control", 0)
    ws_control.append(["Worksheet", "Description", "Data"])
    ws_control.append([new_sheet_name, "", "x"])

    # Save to ToProcess folder
    output_path = TO_PROCESS_DIR / f"{filename_no_ext}.xlsx"
    wb.save(output_path)
    wb.close()

    print(f"✅ Prepared: {file_path.name} → {output_path.name}")


# -----------------------------------------------------------------------------
# Main execution
# -----------------------------------------------------------------------------
def main():
    excel_files = list(TO_PREPARE_DIR.glob("*.xlsx"))

    if not excel_files:
        print("📂 No Excel files found in evs100/ToPrepare/")
        return

    print(f"🔍 Found {len(excel_files)} file(s) to prepare...\n")

    for file_path in excel_files:
        try:
            prepare_evs100_workbook(file_path)
        except Exception as e:
            print(f"❌ Error processing {file_path.name}: {e}")

    print("\n✅ All files processed. Ready in evs100/ToProcess/")


if __name__ == "__main__":
    main()